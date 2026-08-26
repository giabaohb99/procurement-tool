"""Celery task chạy import ở worker: cập nhật trạng thái + đếm + log, xong bắn chuông."""
import traceback
from datetime import datetime
from io import BytesIO, StringIO

import app.core.all_models  # noqa: F401 — đăng ký toàn bộ mapper
from app.core.celery_app import celery_app
from app.core.database import SessionLocal
from app.core.storage import download_bytes
from app.modules.attachment.model import StoredFile
from app.modules.notification.model import Notification

from . import catalog_import, doc_import
from .model import ImportBatch, ImportMode, ImportModule, ImportStatus

_MODULE_LABEL = {
    ImportModule.SURVEY: "Khảo sát",
    ImportModule.PURCHASE_ORDER: "Đơn mua hàng",
    ImportModule.COMPANY: "Công ty",
    ImportModule.DEPARTMENT: "Phòng ban",
    ImportModule.EMPLOYEE: "Nhân sự",
    ImportModule.SUPPLIER: "Nhà cung cấp",
    ImportModule.PRODUCT: "Sản phẩm & Vật tư",
    ImportModule.UNIT: "Đơn vị tính",
    ImportModule.ITEM_GROUP: "Phân loại VTBB",
    ImportModule.WAREHOUSE: "Danh mục kho",
    ImportModule.SURVEY_REQUEST: "Yêu cầu báo giá",
    ImportModule.PURCHASE_REQUEST: "Yêu cầu mua hàng",
}


@celery_app.task(name="import_tool.run_import")
def run_import(batch_id: int) -> dict:
    db = SessionLocal()
    try:
        batch = db.get(ImportBatch, batch_id)
        if not batch:
            return {"error": "batch not found"}
        batch.status = ImportStatus.RUNNING
        batch.started_at = datetime.utcnow()
        db.commit()

        sf = db.get(StoredFile, batch.file_id)
        if not sf:
            raise RuntimeError("Không tìm thấy file đã lưu")
        wb = _load_workbook(sf)

        # Khảo sát + ĐMH nay đi MẪU CHUẨN (doc_import), không còn Misa (CR-176).
        if catalog_import.is_catalog_module(batch.module):
            catalog_import.run(db, batch, wb, apply=(batch.mode == ImportMode.APPLY))
        elif doc_import.is_doc_module(batch.module):
            doc_import.run(db, batch, wb, apply=(batch.mode == ImportMode.APPLY))
        else:
            raise RuntimeError("Module import chưa hỗ trợ")

        db.refresh(batch)
        _notify(db, batch, ok=True)
        return {"status": "done", "created": batch.created_count, "updated": batch.updated_count,
                "error": batch.error_count}
    except catalog_import.ImportValidationError as e:
        # Chọn nhầm phân hệ/bảng: chỉ ghi thông điệp thân thiện, KHÔNG kèm traceback.
        db.rollback()
        b = db.get(ImportBatch, batch_id)
        if b:
            b.status = ImportStatus.FAILED
            b.error_summary = str(e)
            b.finished_at = datetime.utcnow()
            db.commit()
            _notify(db, b, ok=False)
        return {"status": "failed", "error": str(e)}
    except Exception as e:  # noqa: BLE001
        db.rollback()
        b = db.get(ImportBatch, batch_id)
        if b:
            b.status = ImportStatus.FAILED
            b.error_summary = f"{e}\n{traceback.format_exc()[-2000:]}"
            b.finished_at = datetime.utcnow()
            db.commit()
            _notify(db, b, ok=False)
        return {"status": "failed", "error": str(e)}
    finally:
        db.close()


def workbook_from_bytes(filename: str, raw: bytes):
    """Bytes -> openpyxl Workbook. `.csv` dựng workbook một sheet để các adapter
    (vốn đọc theo ô/worksheet, có nhánh lùi về `worksheets[0]`) chạy y nguyên."""
    import openpyxl
    if (filename or "").lower().endswith(".csv"):
        return _csv_to_workbook(raw)
    return openpyxl.load_workbook(BytesIO(raw), data_only=True, read_only=False)


def _load_workbook(sf):
    """Đọc file đã lưu (theo file_key) thành openpyxl Workbook."""
    return workbook_from_bytes(sf.filename or "", download_bytes(sf.file_key))


def precheck_headers(module: int, wb) -> None:
    """Đối chiếu dòng tiêu đề của file với bảng đã chọn — sai thì raise
    `ImportValidationError`. Dùng để chặn NGAY tại bước upload (đồng bộ), không
    phải đợi chạy nền rồi mới biết chọn nhầm phân hệ/bảng."""
    if catalog_import.is_catalog_module(module):
        adapter = catalog_import.ADAPTERS[module]
        fields = adapter["fields"]
    elif doc_import.is_doc_module(module):
        adapter = doc_import.DOC_ADAPTERS[module]
        fields = [adapter["code"], *adapter["header_fields"], *adapter["line_fields"]]
    else:
        return  # module chưa hỗ trợ — để bước sau báo
    sheet = adapter["sheet"]
    ws = wb[sheet] if sheet in wb.sheetnames else wb.worksheets[0]
    field_col = catalog_import.build_field_col(ws, fields)
    catalog_import.ensure_file_matches(sheet, fields, field_col)


def _csv_to_workbook(raw: bytes):
    """CSV -> Workbook một sheet (bỏ BOM).

    Dấu phân cách chọn theo dấu ÁP ĐẢO ở DÒNG ĐẦU (`;` / tab / `,`) — bền hơn
    `csv.Sniffer` ở đúng bộ dữ liệu này: nhãn cột có sẵn dấu phẩy
    (vd «VAT (tỉ lệ, vd 0.08)») và Excel VN xuất bằng `;` vì `,` là dấu thập phân
    («0,1»); để Sniffer tự đoán là nó cắt nhầm ngay tại dấu phẩy trong nhãn.
    """
    import csv

    import openpyxl
    text = raw.decode("utf-8-sig", errors="replace")
    header = next((ln for ln in text.splitlines() if ln.strip()), "")
    delim = max((";", "\t", ","), key=header.count)
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in csv.reader(StringIO(text), delimiter=delim):
        ws.append(row)
    return wb


def _notify(db, batch, ok: bool) -> None:
    label = _MODULE_LABEL.get(batch.module, "")
    mode = "chạy thử" if batch.mode == ImportMode.DRY_RUN else "ghi"
    if ok:
        title = f"Import {label} ({mode}) xong"
        body = f"{batch.created_count} tạo · {batch.updated_count} cập nhật · {batch.error_count} lỗi"
    else:
        title = f"Import {label} lỗi"
        body = "Xem chi tiết log để xử lý"
    db.add(Notification(user_id=batch.created_by, title=title, body=body,
                        link=f"/import-batches/{batch.id}", created_by=batch.created_by))
    db.commit()
