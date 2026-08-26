"""Import danh mục nền (Công ty · Phòng ban · Nhân sự) qua khung batch dùng chung.

Cùng cơ chế với `survey_import` / `po_import`:
- Tích luỹ log / changes / counts trong bộ nhớ; cuối cùng `apply` thì `db.commit()`,
  `dry-run` thì `db.rollback()` — rồi `_persist` ghi batch + log (và changes NẾU apply)
  ở một transaction riêng, nên **bản chạy thử vẫn giữ được log để xem trước**.
- Mỗi thực thể khai một *adapter* (cột file → thuộc tính model, khoá trùng, tham chiếu).
  Thêm danh mục import mới về sau = thêm một adapter, không đụng thân khung.

Phạm vi v1 (Đ-13d): các cột cốt lõi 1 bảng. KHÔNG tạo tài khoản đăng nhập cho nhân sự,
KHÔNG xử lý kiêm nhiệm nhiều phòng ban (chỉ set `department_id` chính), KHÔNG đụng
người đại diện pháp luật của công ty — các thứ đó mở việc riêng.
"""
import json
from datetime import datetime
from io import BytesIO

from sqlalchemy import inspect as sa_inspect
from sqlalchemy.orm import Session

from app.modules.catalog.model import ItemGroup, Unit, Warehouse
from app.modules.company.model import Company
from app.modules.department.model import Department
from app.modules.employee.model import Employee
from app.modules.product.model import Product
from app.modules.supplier.model import Supplier

from .model import (ImportBatch, ImportChange, ImportLog, ImportModule,
                    ImportStatus, LogLevel)

HEADER_ROW = 1
DATA_START = 2

# Ô dữ liệu ghi đúng chuỗi này -> XÓA giá trị trường (về rỗng/0/false), khác với
# để trống (giữ nguyên giá trị cũ khi cột không có trong file).
EMPTY_SENTINEL = "__/empty_value/__"
# Ô cột "Hành động" ghi chuỗi này -> XÓA bản ghi (theo Mã) thay vì tạo/sửa.
DELETE_SENTINEL = "__/delete/__"
# Tiêu đề (đã chuẩn hoá) được nhận là cột hành động.
_ACTION_HEADERS = {"hành động", "hanh dong", "action", "thao tác", "thao tac"}
# Tiền tố mã TỰ SINH khi dòng thiếu Mã — mỗi bảng một tiền tố cho dễ đọc.
_AUTO_PREFIX = {
    ImportModule.COMPANY: "CTY", ImportModule.DEPARTMENT: "PB",
    ImportModule.EMPLOYEE: "NV", ImportModule.SUPPLIER: "NCC",
    ImportModule.PRODUCT: "SP", ImportModule.UNIT: "DVT",
    ImportModule.ITEM_GROUP: "PL", ImportModule.WAREHOUSE: "KHO",
}


def _empty_for_kind(kind: str):
    """Giá trị RỖNG theo kiểu trường (dùng cho sentinel xóa dữ liệu ô)."""
    return {"int": 0, "float": 0.0, "bool": False, "ref": 0}.get(kind, "")


def _gen_code(db: Session, module: int, model, dedupe: str, used: set) -> str:
    """Sinh mã mới, DUY NHẤT (không trùng DB lẫn các mã vừa sinh trong batch)."""
    prefix = _AUTO_PREFIX.get(module, "AUTO")
    n = db.query(model).count() + 1
    while True:
        cand = f"{prefix}{n:05d}"
        if cand not in used and not db.query(model).filter(getattr(model, dedupe) == cand).first():
            used.add(cand)
            return cand
        n += 1


# ---------- chuẩn hoá ----------
def _s(v) -> str:
    return "" if v is None else str(v).strip()


def _norm(v) -> str:
    """Chuẩn hoá tiêu đề cột để so khớp: bỏ dấu sao, gộp khoảng trắng, viết thường."""
    return " ".join(_s(v).replace("*", "").lower().split())


class ImportValidationError(Exception):
    """File tải lên không khớp bảng đã chọn (sai phân hệ/bảng hoặc nhầm file mẫu)."""


def _field_norms(f: dict) -> list[str]:
    """Các tiêu đề (đã chuẩn hoá) mà một field CHẤP NHẬN: tiêu đề chính + alias.

    Alias để khớp cả file xuất ra có nhãn cột hơi khác template (vd export ghi
    «Ngày QĐ (có sẵn)» còn template ghi «Số ngày QĐ (có sẵn)»)."""
    return [_norm(f["header"]), *(_norm(a) for a in f.get("aliases") or [])]


def _match_col(header_col: dict, f: dict):
    """Chỉ số cột của field trong file — thử tiêu đề chính rồi tới alias."""
    for n in _field_norms(f):
        if n in header_col:
            return header_col[n]
    return None


def build_field_col(ws, fields: list[dict]) -> dict:
    """attr -> chỉ số cột theo dòng tiêu đề của worksheet (None nếu file thiếu cột)."""
    header_col: dict[str, int] = {}
    for col in range(1, (ws.max_column or 0) + 1):
        key = _norm(ws.cell(row=HEADER_ROW, column=col).value)
        if key:
            header_col[key] = col
    return {f["attr"]: _match_col(header_col, f) for f in fields}


def ensure_file_matches(sheet_label: str, fields: list[dict], field_col: dict) -> None:
    """Kiểm tra file có đúng cho bảng đang nhập không — chặn sớm trước khi ghi.

    Luật: MỌI cột BẮT BUỘC của bảng phải có mặt ở dòng tiêu đề của file. Cột bắt
    buộc mang nhãn riêng từng bảng («Tên pháp lý», «Tên kho», «Mã NV», «Mã
    phiếu»…) nên thiếu chúng gần như chắc chắn là dùng nhầm file — trong khi vẫn
    KHÔNG chặn oan file đúng bảng nhưng lược bớt cột tuỳ chọn.

    `field_col`: attr -> chỉ số cột (None nếu file không có cột đó).
    """
    missing = [_s(f["header"]).replace("*", "").strip()
               for f in fields
               if f.get("required") and field_col.get(f["attr"]) is None]
    if not missing:
        return
    raise ImportValidationError(
        f"File không khớp bảng «{sheet_label}» — thiếu cột bắt buộc: "
        f"{', '.join(missing)}. "
        "Kiểm tra lại phân hệ/bảng đã chọn, hoặc tải đúng file mẫu của bảng này."
    )


def _to_int(v, default: int = 0) -> int:
    try:
        return int(float(_s(v)))
    except (ValueError, TypeError):
        return default


def _to_float(v, default: float = 0.0) -> float:
    try:
        return float(_s(v).replace(",", "."))
    except (ValueError, TypeError):
        return default


def _to_bool(v, default: bool = True) -> bool:
    s = _s(v).lower()
    if s in ("0", "false", "không", "khong", "ẩn", "an", "no", "n", "ngưng", "ngung", "tắt", "tat"):
        return False
    if s in ("1", "true", "có", "co", "hiện", "hien", "yes", "y", "hoạt động", "hoat dong", "bật", "bat"):
        return True
    return default


# ---------- khai báo adapter ----------
# Mỗi field: header (nhãn cột trong file mẫu), attr (thuộc tính model),
# kind: str | int | bool | ref, required, ref (company|department|employee|self),
# default.
def _f(header, attr, kind="str", required=False, ref=None, default=None, aliases=None):
    return {"header": header, "attr": attr, "kind": kind, "required": required,
            "aliases": aliases or [],
            "ref": ref, "default": default}


_REF_MODEL = {"company": Company, "department": Department, "employee": Employee}

ADAPTERS: dict[int, dict] = {
    ImportModule.COMPANY: {
        "label": "Công ty",
        "model": Company,
        "sheet": "Cong ty",
        "dedupe": "code",
        "fields": [
            _f("Mã *", "code", required=True),
            _f("Tên công ty *", "name", required=True),
            _f("Tên viết tắt", "short_name"),
            _f("Mã hiệu văn bản", "issue_code"),
            _f("Cấp (1 tập đoàn/2 công ty/3 trực thuộc)", "level", kind="int", default=2),
            _f("Mã số thuế", "tax_code"),
            _f("Địa chỉ", "address"),
            _f("Email nhận hóa đơn", "invoice_email"),
            _f("Công ty mẹ (mã)", "parent", kind="ref", ref="self"),
            _f("Hoạt động (1/0)", "is_active", kind="bool", default=True),
        ],
    },
    ImportModule.DEPARTMENT: {
        "label": "Phòng ban",
        "model": Department,
        "sheet": "Phong ban",
        "dedupe": "code",
        "fields": [
            _f("Mã *", "code", required=True),
            _f("Tên phòng ban *", "name", required=True),
            _f("Mã hiệu văn bản", "issue_code"),
            _f("Loại (1 chức năng/2 kinh doanh/3 dự án)", "kind", kind="int", default=1),
            _f("Công ty (mã)", "company_id", kind="ref", ref="company"),
            _f("Phòng cấp trên (mã)", "parent", kind="ref", ref="self"),
            _f("Trưởng bộ phận (mã NV)", "manager_id", kind="ref", ref="employee"),
            _f("Hoạt động (1/0)", "is_active", kind="bool", default=True),
        ],
    },
    ImportModule.EMPLOYEE: {
        "label": "Nhân sự",
        "model": Employee,
        "sheet": "Nhan su",
        "dedupe": "code",
        "fields": [
            _f("Mã NV *", "code", required=True),
            _f("Họ tên *", "full_name", required=True),
            _f("Email", "email"),
            _f("Điện thoại", "phone"),
            _f("Công ty (mã)", "company_id", kind="ref", ref="company"),
            _f("Phòng ban (mã)", "department_id", kind="ref", ref="department"),
            _f("Chức vụ", "position"),
            _f("Trạng thái (official/probation/…)", "status", default="official"),
            _f("Hoạt động (1/0)", "is_active", kind="bool", default=True),
        ],
    },
    ImportModule.SUPPLIER: {
        "label": "Nhà cung cấp",
        "model": Supplier,
        "sheet": "Nha cung cap",
        "dedupe": "code",
        "fields": [
            _f("Mã *", "code", required=True),
            _f("Tên pháp lý *", "name", required=True),
            _f("Mã số thuế", "tax_code"),
            _f("Địa chỉ", "address"),
            _f("Loại NCC (goods/transport)", "supplier_type", default="goods"),
            _f("Người liên hệ", "contact_person"),
            _f("Điện thoại", "phone"),
            _f("Hình thức thanh toán", "payment_terms"),
            _f("Số tài khoản", "bank_account"),
            _f("Ngân hàng", "bank_name"),
            _f("Tên tài khoản", "bank_account_name"),
            _f("VAT (tỉ lệ, vd 0.08)", "vat", kind="float", default=0.08),
            _f("Hoạt động (1/0)", "is_active", kind="bool", default=True),
        ],
    },
    ImportModule.PRODUCT: {
        "label": "Sản phẩm & Vật tư",
        "model": Product,
        "sheet": "San pham",
        "dedupe": "code",
        "fields": [
            _f("Mã VTBB/NL *", "code", required=True),
            _f("Tên VTBB/NL *", "name", required=True),
            _f("Tên trên hóa đơn", "invoice_name"),
            _f("Tên pháp lý", "legal_name"),
            _f("Phân loại (mã/tên)", "item_group"),
            _f("Đơn vị tính", "unit"),
            _f("Mã HH", "hh_code"),
            _f("Tên HH", "hh_name"),
            _f("Thông số kỹ thuật", "specs"),
            _f("Hoạt động (1/0)", "is_active", kind="bool", default=True),
        ],
    },
    ImportModule.UNIT: {
        "label": "Đơn vị tính",
        "model": Unit,
        "sheet": "Don vi tinh",
        "dedupe": "code",
        "fields": [
            _f("Mã *", "code", required=True),
            _f("Tên *", "name", required=True),
            _f("Hoạt động (1/0)", "is_active", kind="bool", default=True),
        ],
    },
    ImportModule.ITEM_GROUP: {
        "label": "Phân loại VTBB",
        "model": ItemGroup,
        "sheet": "Phan loai",
        "dedupe": "code",
        # name là UNIQUE: dòng thiếu Mã đối chiếu theo tên -> re-import cập nhật,
        # không tạo bản trùng làm vỡ ràng buộc unique.
        "natural_key": "name",
        "fields": [
            _f("Mã *", "code", required=True),
            _f("Tên *", "name", required=True),
            # alias: khớp cả nhãn cột ngắn của file EXPORT («Ngày QĐ (có sẵn)»).
            _f("Số ngày QĐ (có sẵn)", "std_days", aliases=["Ngày QĐ (có sẵn)"]),
            _f("Số ngày QĐ (không sẵn)", "std_days_unavail", aliases=["Ngày QĐ (không sẵn)"]),
            _f("Ngày áp dụng", "apply_date"),
            _f("Ghi chú", "note"),
            _f("Hoạt động (1/0)", "is_active", kind="bool", default=True, aliases=["Hoạt động"]),
        ],
    },
    ImportModule.WAREHOUSE: {
        "label": "Danh mục kho",
        "model": Warehouse,
        "sheet": "Kho",
        "dedupe": "code",
        "fields": [
            _f("Mã *", "code", required=True),
            _f("Tên kho *", "name", required=True),
            _f("Địa chỉ", "address"),
            _f("Hoạt động (1/0)", "is_active", kind="bool", default=True),
        ],
    },
}


def is_catalog_module(module: int) -> bool:
    return module in ADAPTERS


# ---------- chạy import ----------
def run(db: Session, batch: ImportBatch, wb, apply: bool) -> None:
    adapter = ADAPTERS[batch.module]
    model = adapter["model"]
    ws = wb[adapter["sheet"]] if adapter["sheet"] in wb.sheetnames else wb.worksheets[0]

    # Bản đồ tiêu đề chuẩn hoá -> chỉ số cột.
    header_col: dict[str, int] = {}
    for col in range(1, (ws.max_column or 0) + 1):
        key = _norm(ws.cell(row=HEADER_ROW, column=col).value)
        if key:
            header_col[key] = col
    field_col = {f["attr"]: _match_col(header_col, f) for f in adapter["fields"]}
    ensure_file_matches(adapter["sheet"], adapter["fields"], field_col)

    # Cột "Hành động" (tùy chọn) — không thuộc bộ trường, dò riêng theo tiêu đề.
    action_col = next((c for k, c in header_col.items() if k in _ACTION_HEADERS), None)
    dedupe = adapter["dedupe"]
    used_codes: set = set()

    counts = {"created": 0, "updated": 0, "deleted": 0, "skipped": 0,
              "warning": 0, "review": 0, "error": 0}
    logs: list[dict] = []
    changes: list[dict] = []

    def log(row_no, level, category, message, ref_key="", target_code=""):
        logs.append({"sheet": adapter["sheet"], "row_no": row_no, "level": int(level),
                     "category": category, "message": message, "ref_key": ref_key,
                     "target_code": target_code})
        if level == LogLevel.WARNING:
            counts["warning"] += 1
        elif level == LogLevel.REVIEW:
            counts["review"] += 1
        elif level == LogLevel.ERROR:
            counts["error"] += 1

    def read(row, attr):
        col = field_col.get(attr)
        return ws.cell(row=row, column=col).value if col else None

    def resolve_ref(ref, code, row_no, header):
        """code -> id. Không thấy -> log REVIEW, trả 0 (để trống)."""
        if not code:
            return 0
        ref_model = model if ref == "self" else _REF_MODEL[ref]
        obj = db.query(ref_model).filter(ref_model.code == code).first()
        if obj:
            return obj.id
        log(row_no, LogLevel.REVIEW, "ref_not_found",
            f"«{header}» = '{code}' không có trong danh mục — để trống", ref_key=str(code))
        return 0

    total = 0
    max_row = ws.max_row or 0
    for r in range(DATA_START, max_row + 1):
        action = _s(ws.cell(row=r, column=action_col).value).lower() if action_col else ""
        code = _s(read(r, dedupe))
        if code == EMPTY_SENTINEL:
            code = ""
        # Bỏ dòng trắng hoàn toàn (không khoá, không hành động, không giá trị nào).
        has_value = any(_s(read(r, f["attr"])) for f in adapter["fields"])
        if not code and not action and not has_value:
            continue
        total += 1

        # XÓA bản ghi: cột Hành động = __/delete/__, xác định theo Mã.
        if action == DELETE_SENTINEL:
            if not code:
                counts["skipped"] += 1
                log(r, LogLevel.ERROR, "delete_no_code",
                    "Xóa cần Mã để xác định dòng — bỏ qua")
                continue
            existing = db.query(model).filter(getattr(model, dedupe) == code).first()
            if not existing:
                counts["skipped"] += 1
                log(r, LogLevel.WARNING, "delete_not_found",
                    f"Không thấy '{code}' để xóa — bỏ qua", ref_key=code)
                continue
            snap = {f["attr"]: getattr(existing, f["attr"], None) for f in adapter["fields"]}
            old_id = existing.id
            try:
                with db.begin_nested():
                    db.delete(existing)
                    db.flush()
            except Exception:  # noqa: BLE001 — thường do khoá ngoại đang tham chiếu
                # log(ERROR) đã cộng counts["error"] — không cộng tay lần nữa.
                log(r, LogLevel.ERROR, "delete_failed",
                    f"Không xóa được '{code}' (có thể đang được tham chiếu) — bỏ qua",
                    ref_key=code, target_code=code)
                continue
            counts["deleted"] += 1
            changes.append({"target_id": old_id, "was_new": 2,
                            "snapshot": json.dumps({"cols": snap}, ensure_ascii=False, default=str)})
            log(r, LogLevel.INFO, "deleted", f"Xóa {adapter['label']} '{code}'",
                ref_key=code, target_code=code)
            continue

        # Dòng thiếu Mã: nếu bảng có KHÓA TỰ NHIÊN duy nhất (vd name) và đã tồn tại
        # bản ghi trùng -> CẬP NHẬT bản ghi đó (re-import không đẻ trùng, không đụng
        # ràng buộc unique); chưa có -> tạo mới với MÃ TỰ SINH.
        auto_code = False
        if not code:
            nk = adapter.get("natural_key")
            nk_val = _s(read(r, nk)) if nk else ""
            existing_nk = (db.query(model).filter(getattr(model, nk) == nk_val).first()
                           if nk and nk_val else None)
            if existing_nk:
                code = getattr(existing_nk, dedupe)
            else:
                code = _gen_code(db, batch.module, model, dedupe, used_codes)
                auto_code = True

        # Kiểm trường bắt buộc (Mã lấy giá trị ĐÃ RESOLVE; sentinel xóa coi như trống).
        missing = []
        for f in adapter["fields"]:
            if not f["required"]:
                continue
            val = code if f["attr"] == dedupe else _s(read(r, f["attr"]))
            if not val or val == EMPTY_SENTINEL:
                missing.append(f["header"].replace(" *", ""))
        if missing:
            counts["skipped"] += 1
            log(r, LogLevel.ERROR, "missing_required",
                f"Thiếu trường bắt buộc: {', '.join(missing)} — bỏ qua dòng", ref_key=code)
            continue

        # Dựng dict giá trị theo kiểu.
        data: dict = {}
        for f in adapter["fields"]:
            attr, kind = f["attr"], f["kind"]
            if field_col.get(attr) is None:
                # Cột không có trong file — bỏ qua để giữ nguyên giá trị cũ khi cập nhật.
                continue
            raw = read(r, attr)
            # Sentinel xóa dữ liệu ô -> ép về giá trị rỗng theo kiểu.
            if _s(raw) == EMPTY_SENTINEL:
                data[attr] = _empty_for_kind(kind)
                continue
            if kind == "int":
                data[attr] = _to_int(raw, f["default"] if f["default"] is not None else 0)
            elif kind == "float":
                data[attr] = _to_float(raw, f["default"] if f["default"] is not None else 0.0)
            elif kind == "bool":
                data[attr] = _to_bool(raw, bool(f["default"]))
            elif kind == "ref":
                data[attr] = resolve_ref(f["ref"], _s(raw), r, f["header"])
            else:
                val = _s(raw)
                if not val and f["default"] is not None:
                    val = f["default"]
                data[attr] = val
        # Mã luôn lấy giá trị đã resolve (kể cả mã tự sinh).
        data[dedupe] = code

        existing = None if auto_code else db.query(model).filter(getattr(model, dedupe) == code).first()
        if existing:
            # Snapshot GIÁ TRỊ CŨ của đúng các cột sắp đổi — để revert.
            old = {k: getattr(existing, k) for k in data}
            for k, v in data.items():
                setattr(existing, k, v)
            existing.updated_by = batch.created_by
            db.flush()
            counts["updated"] += 1
            changes.append({"target_id": existing.id, "was_new": False,
                            "snapshot": json.dumps({"cols": old}, ensure_ascii=False, default=str)})
            log(r, LogLevel.INFO, "updated", f"Cập nhật {adapter['label']} '{code}'",
                ref_key=code, target_code=code)
        else:
            # Chặn TRÙNG KHÓA TỰ NHIÊN (vd name unique) TRƯỚC khi insert -> tránh lỗi
            # 1062 làm sập cả batch; báo lỗi đúng dòng rồi đi tiếp.
            nk = adapter.get("natural_key")
            if nk and data.get(nk):
                clash = db.query(model).filter(getattr(model, nk) == data[nk]).first()
                if clash:
                    # log(ERROR) đã cộng counts["error"] — không cộng tay lần nữa.
                    log(r, LogLevel.ERROR, "duplicate",
                        f"Trùng «{nk}» = '{data[nk]}' đã tồn tại (mã '{getattr(clash, dedupe)}') — bỏ qua dòng",
                        ref_key=code, target_code=code)
                    continue
            obj = model(created_by=batch.created_by, updated_by=batch.created_by, **data)
            db.add(obj)
            db.flush()
            counts["created"] += 1
            changes.append({"target_id": obj.id, "was_new": True, "snapshot": ""})
            msg = f"Tạo {adapter['label']} mới '{code}'" + (" (mã tự sinh)" if auto_code else "")
            log(r, LogLevel.INFO, "created", msg, ref_key=code, target_code=code)

    if apply:
        db.commit()
    else:
        db.rollback()
    _persist(db, batch, counts, logs, changes if apply else [], total, adapter["sheet"])


def _persist(db, batch, counts, logs, changes, total_rows, sheet) -> None:
    b = db.get(ImportBatch, batch.id)
    b.total_rows = total_rows
    b.created_count = counts["created"]; b.updated_count = counts["updated"]; b.skipped_count = counts["skipped"]
    b.deleted_count = counts.get("deleted", 0)
    b.warning_count = counts["warning"]; b.review_count = counts["review"]; b.error_count = counts["error"]
    b.sheet_info = json.dumps({"sheet": sheet, "rows": total_rows}, ensure_ascii=False)
    for lg in logs:
        db.add(ImportLog(batch_id=b.id, sheet=lg["sheet"], row_no=lg["row_no"], level=lg["level"],
                         category=lg["category"], message=lg["message"][:60000], ref_key=lg["ref_key"][:120],
                         target_code=lg["target_code"][:50], created_by=b.created_by))
    for ch in changes:
        # was_new: 0=sửa · 1=tạo mới · 2=đã xóa (revert tạo lại từ snapshot).
        db.add(ImportChange(batch_id=b.id, survey_id=ch["target_id"], was_new=int(ch["was_new"]),
                            snapshot=ch["snapshot"], created_by=b.created_by))
    b.status = ImportStatus.DONE
    b.finished_at = datetime.utcnow()
    db.commit()


# ---------- hoàn tác ----------
def revert(db: Session, module: int, changes, user_id: int) -> tuple[int, int]:
    """Hoàn tác batch danh mục: bản ghi MỚI -> xoá; bản ghi ĐÃ XÓA (was_new=2) ->
    tạo lại từ snapshot; bản ghi CŨ SỬA -> khôi phục cột đã đổi."""
    from .service import _apply_cols
    model = ADAPTERS[module]["model"]
    deleted = restored = 0
    for ch in changes:
        # was_new = 2: bản ghi bị batch XÓA -> tạo lại từ snapshot (mã giữ nguyên).
        if ch.was_new == 2:
            if not ch.snapshot:
                continue
            cols = json.loads(ch.snapshot).get("cols", {})
            keys = {c.key for c in sa_inspect(model).columns}
            data = {k: v for k, v in cols.items() if k in keys and k not in ("id",)}
            db.add(model(created_by=user_id, updated_by=user_id, **data))
            restored += 1
            continue
        obj = db.get(model, ch.survey_id)   # survey_id = id bản ghi (dùng chung)
        if not obj:
            continue
        if ch.was_new:
            db.delete(obj); deleted += 1
            continue
        if not ch.snapshot:
            continue
        snap = json.loads(ch.snapshot)
        _apply_cols(obj, snap.get("cols", {}))
        obj.updated_by = user_id
        restored += 1
    return deleted, restored


# ---------- file mẫu ----------
def build_template(module: int) -> bytes:
    """Sinh file .xlsx mẫu: đúng bộ cột của adapter (một dòng tiêu đề)."""
    import openpyxl
    adapter = ADAPTERS[module]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = adapter["sheet"]
    for i, f in enumerate(adapter["fields"], start=1):
        ws.cell(row=1, column=i, value=f["header"])
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(14, len(f["header"]) + 2)
    # Cột Hành động: để trống = thêm/sửa; ghi __/delete/__ = xóa dòng theo Mã.
    ai = len(adapter["fields"]) + 1
    ws.cell(row=1, column=ai, value="Hành động")
    ws.column_dimensions[ws.cell(row=1, column=ai).column_letter].width = 16
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
