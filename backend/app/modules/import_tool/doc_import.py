"""Import CHỨNG TỪ nhiều dòng: 1 phiếu = 1 header + N dòng, gộp theo MÃ PHIẾU.

Dùng cho Yêu cầu báo giá (survey_request) và Yêu cầu mua hàng (purchase_request)
— hai loại chưa có importer. (Khảo sát và Đơn mua hàng đã có importer Misa riêng
ở `survey_import`/`po_import`.)

Cùng khung batch: dry-run rollback + apply commit + snapshot `ImportChange` để revert.
**v1 chỉ TẠO MỚI** — mã phiếu đã tồn tại thì bỏ qua cả phiếu + cảnh báo (không sửa
đè phiếu cũ, vì hoà dòng của chứng từ đã có là rủi ro cao).

Cấu trúc file mẫu: MỘT sheet, mỗi dòng là một DÒNG HÀNG; cột **Mã phiếu** gộp các
dòng vào cùng một phiếu, các cột ĐẦU PHIẾU lấy từ dòng ĐẦU TIÊN của mỗi mã.
"""
from collections import OrderedDict
from datetime import date, datetime

from sqlalchemy.orm import Session

from app.modules.company.model import Company
from app.modules.department.model import Department
from app.modules.employee.model import Employee
from app.modules.purchase_order.model import POItem, PurchaseOrder
from app.modules.purchase_request.model import PurchaseRequest, PurchaseRequestItem
from app.modules.survey.model import Survey, SurveySupplierLine
from app.modules.survey_request.model import SurveyRequest, SurveyRequestLine

from . import catalog_import
from .catalog_import import _norm, _s, _to_bool, _to_float, _to_int
from .model import ImportModule, LogLevel

HEADER_ROW = 1
DATA_START = 2

_REF_MODEL = {"company": Company, "department": Department, "employee": Employee}
_REF_NAME = {"company": "name", "department": "name", "employee": "full_name"}


def _df(header, attr, kind="str", required=False, ref=None, name_attr=None, default=None):
    """Field khai báo. `name_attr`: khi ref resolve, ghi thêm TÊN vào cột này (bản chụp)."""
    return {"header": header, "attr": attr, "kind": kind, "required": required,
            "ref": ref, "name_attr": name_attr, "default": default}


def _to_date_str(v) -> str:
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return _s(v)[:10]


DOC_ADAPTERS: dict[int, dict] = {
    ImportModule.SURVEY_REQUEST: {
        "label": "Yêu cầu báo giá",
        "sheet": "YCBG",
        "header_model": SurveyRequest,
        "line_model": SurveyRequestLine,
        "line_fk": "survey_request_id",
        "code": _df("Mã phiếu *", "code", required=True),
        "header_fields": [
            _df("Công ty (mã)", "company_id", kind="ref", ref="company"),
            _df("Người yêu cầu", "requester"),
            _df("Phòng ban (mã)", "department_id", kind="ref", ref="department", name_attr="department"),
            _df("Mục đích", "purpose"),
            _df("Ngày yêu cầu", "request_date", kind="date"),
        ],
        "line_fields": [
            _df("Phân loại", "item_group"),
            _df("Yêu cầu kỹ thuật *", "requirement_detail", required=True),
            _df("Số lượng", "request_qty", kind="float"),
            _df("ĐVT", "uom"),
            _df("Giá đề xuất", "proposed_price", kind="float"),
        ],
    },
    ImportModule.PURCHASE_REQUEST: {
        "label": "Yêu cầu mua hàng",
        "sheet": "YCMH",
        "header_model": PurchaseRequest,
        "line_model": PurchaseRequestItem,
        "line_fk": "pr_id",
        "code": _df("Mã phiếu *", "code", required=True),
        "header_fields": [
            _df("Công ty (mã)", "company_id", kind="ref", ref="company"),
            _df("Người yêu cầu", "requester"),
            _df("Phòng ban (mã)", "department_id", kind="ref", ref="department", name_attr="department"),
            _df("Mục đích", "purpose"),
            _df("Ngày yêu cầu", "request_date", kind="date"),
            _df("Ngày cần hàng", "need_date", kind="date"),
        ],
        "line_fields": [
            _df("Mã sản phẩm", "product_code"),
            _df("Tên sản phẩm *", "product_name", required=True),
            _df("Phân loại", "item_group"),
            _df("Số lượng", "qty", kind="float"),
            _df("ĐVT", "unit"),
            _df("Giá đề xuất", "price", kind="float"),
        ],
    },
    # Khảo sát + Đơn mua hàng chuyển sang MẪU CHUẨN (CR-176), thay importer Misa.
    # Revert vẫn qua `_revert_survey`/`_revert_po` (dọn đủ 2 loại dòng / side-effect).
    ImportModule.SURVEY: {
        "label": "Khảo sát",
        "sheet": "Khao sat",
        "header_model": Survey,
        "line_model": SurveySupplierLine,
        "line_fk": "survey_id",
        "code": _df("Mã phiếu *", "code", required=True),
        "header_defaults": {"survey_type": "supplier"},
        "header_fields": [
            _df("Loại (supplier/product)", "survey_type", default="supplier"),
            _df("Phân loại", "item_group"),
            _df("Nội dung chính", "main_content"),
            _df("Yêu cầu kỹ thuật", "requirement_detail"),
            _df("SL dự kiến", "request_qty", kind="float"),
        ],
        "line_fields": [
            _df("Mã NCC", "supplier_code"),
            _df("Tên NCC *", "supplier_name", required=True),
            _df("MST", "tax_code"),
            _df("Người liên hệ", "contact_person"),
            _df("SĐT", "contact_phone"),
            _df("Chính sách hóa đơn", "invoice_policy"),
            _df("Chính sách công nợ", "debt_policy"),
        ],
    },
    ImportModule.PURCHASE_ORDER: {
        "label": "Đơn mua hàng",
        "sheet": "Don mua hang",
        "header_model": PurchaseOrder,
        "line_model": POItem,
        "line_fk": "po_id",
        "code": _df("Mã phiếu *", "code", required=True),
        "header_fields": [
            _df("Mã Misa", "misa_code"),
            _df("Mã YCMH", "pr_code"),
            _df("Công ty (mã)", "company_id", kind="ref", ref="company"),
            _df("Mã NCC", "supplier_code"),
            _df("Tên NCC", "supplier_name"),
            _df("Ngày đặt", "order_date", kind="date"),
        ],
        "line_fields": [
            _df("Mã sản phẩm", "product_code"),
            _df("Tên sản phẩm *", "product_name", required=True),
            _df("Phân loại", "item_group"),
            _df("ĐVT", "unit"),
            _df("SL đặt", "qty_order", kind="float"),
            _df("Đơn giá", "price", kind="float"),
        ],
    },
}


def is_doc_module(module: int) -> bool:
    return module in DOC_ADAPTERS


def run(db: Session, batch, wb, apply: bool) -> None:
    adapter = DOC_ADAPTERS[batch.module]
    header_model = adapter["header_model"]
    line_model = adapter["line_model"]
    line_fk = adapter["line_fk"]
    ws = wb[adapter["sheet"]] if adapter["sheet"] in wb.sheetnames else wb.worksheets[0]

    all_fields = [adapter["code"], *adapter["header_fields"], *adapter["line_fields"]]
    header_col: dict[str, int] = {}
    for col in range(1, (ws.max_column or 0) + 1):
        key = _norm(ws.cell(row=HEADER_ROW, column=col).value)
        if key:
            header_col[key] = col
    field_col = {f["attr"]: header_col.get(_norm(f["header"])) for f in all_fields}
    catalog_import.ensure_file_matches(adapter["sheet"], all_fields, field_col)

    counts = {"created": 0, "updated": 0, "skipped": 0, "warning": 0, "review": 0, "error": 0}
    logs: list[dict] = []
    changes: list[dict] = []

    def log(row_no, level, category, message, ref_key="", target_code=""):
        logs.append({"sheet": adapter["sheet"], "row_no": row_no, "level": int(level),
                     "category": category, "message": message, "ref_key": ref_key,
                     "target_code": target_code})
        if level == LogLevel.WARNING:
            counts["warning"] += 1
        elif level == LogLevel.REVIEW:
            counts["review"] += 1
        elif level == LogLevel.ERROR:
            counts["error"] += 1

    def read(row, attr):
        col = field_col.get(attr)
        return ws.cell(row=row, column=col).value if col else None

    def build(fields, row, target: dict, code_for_log: str):
        """Đổ giá trị các field vào dict target theo kiểu; ref set id + tên snapshot."""
        for f in fields:
            if field_col.get(f["attr"]) is None:
                continue
            raw = read(row, f["attr"])
            kind = f["kind"]
            if kind == "int":
                target[f["attr"]] = _to_int(raw, f["default"] or 0)
            elif kind == "float":
                target[f["attr"]] = _to_float(raw, f["default"] or 0.0)
            elif kind == "bool":
                target[f["attr"]] = _to_bool(raw, bool(f["default"]))
            elif kind == "date":
                target[f["attr"]] = _to_date_str(raw)
            elif kind == "ref":
                target[f["attr"]] = _resolve_ref(db, f, _s(raw), row, code_for_log, log)
                obj = _resolve_ref_obj(db, f["ref"], _s(raw))
                if obj and f.get("name_attr"):
                    target[f["name_attr"]] = getattr(obj, _REF_NAME[f["ref"]], "") or ""
            else:
                val = _s(raw)
                if not val and f["default"] is not None:
                    val = f["default"]
                target[f["attr"]] = val

    # Gộp dòng theo mã phiếu (giữ thứ tự).
    groups: "OrderedDict[str, list[int]]" = OrderedDict()
    total = 0
    max_row = ws.max_row or 0
    for r in range(DATA_START, max_row + 1):
        code = _s(read(r, "code"))
        if not code and not any(_s(read(r, f["attr"])) for f in all_fields):
            continue
        total += 1
        if not code:
            counts["skipped"] += 1
            log(r, LogLevel.ERROR, "missing_code", "Dòng thiếu Mã phiếu — bỏ qua")
            continue
        groups.setdefault(code, []).append(r)

    for code, rows in groups.items():
        if db.query(header_model).filter(header_model.code == code).first():
            counts["skipped"] += len(rows)
            log(rows[0], LogLevel.WARNING, "doc_exists",
                f"Mã phiếu '{code}' đã tồn tại — bỏ qua cả phiếu", ref_key=code)
            continue

        header_data: dict = {"code": code}
        build(adapter["header_fields"], rows[0], header_data, code)
        # Cột đầu phiếu vắng trong file nhưng model đòi non-null (vd Survey.survey_type).
        for k, v in adapter.get("header_defaults", {}).items():
            header_data.setdefault(k, v)
        header = header_model(created_by=batch.created_by, updated_by=batch.created_by, **header_data)
        db.add(header)
        db.flush()

        nlines = 0
        for r in rows:
            missing = [f["header"].replace(" *", "") for f in adapter["line_fields"]
                       if f["required"] and not _s(read(r, f["attr"]))]
            if missing:
                log(r, LogLevel.ERROR, "line_missing",
                    f"Dòng thiếu: {', '.join(missing)} — bỏ dòng", ref_key=code, target_code=code)
                continue
            line_data: dict = {line_fk: header.id}
            build(adapter["line_fields"], r, line_data, code)
            # Bỏ dòng rỗng (không có nội dung nào ngoài khoá ngoại).
            if not any(_s(v) for k, v in line_data.items() if k != line_fk):
                continue
            db.add(line_model(created_by=batch.created_by, updated_by=batch.created_by, **line_data))
            nlines += 1

        db.flush()
        # Phiếu không có DÒNG hợp lệ nào -> không tạo phiếu rỗng, gỡ header vừa thêm.
        if nlines == 0:
            db.delete(header)
            db.flush()
            counts["skipped"] += len(rows)
            log(rows[0], LogLevel.WARNING, "doc_no_line",
                f"Phiếu '{code}' không có dòng hợp lệ — bỏ qua", ref_key=code)
            continue

        counts["created"] += 1
        changes.append({"target_id": header.id, "was_new": True, "snapshot": ""})
        log(rows[0], LogLevel.INFO, "doc_created",
            f"Tạo {adapter['label']} '{code}' với {nlines} dòng", ref_key=code, target_code=code)

    if apply:
        db.commit()
    else:
        db.rollback()
    catalog_import._persist(db, batch, counts, logs, changes if apply else [], total, adapter["sheet"])


def _resolve_ref_obj(db, ref, code):
    if not code:
        return None
    m = _REF_MODEL[ref]
    return db.query(m).filter(m.code == code).first()


def _resolve_ref(db, f, code, row_no, code_for_log, log) -> int:
    if not code:
        return 0
    obj = _resolve_ref_obj(db, f["ref"], code)
    if obj:
        return obj.id
    log(row_no, LogLevel.REVIEW, "ref_not_found",
        f"«{f['header']}» = '{code}' không có trong danh mục — để trống", ref_key=code_for_log)
    return 0


def revert(db: Session, module: int, changes, user_id: int) -> tuple[int, int]:
    """Hoàn tác: xoá phiếu do batch tạo (kèm dòng). v1 create-only nên không có khôi phục."""
    adapter = DOC_ADAPTERS[module]
    header_model = adapter["header_model"]
    line_model = adapter["line_model"]
    line_fk = adapter["line_fk"]
    deleted = 0
    for ch in changes:
        if not ch.was_new:
            continue
        h = db.get(header_model, ch.survey_id)   # survey_id dùng chung = header id
        if not h:
            continue
        db.query(line_model).filter(getattr(line_model, line_fk) == h.id).delete(synchronize_session=False)
        db.delete(h)
        deleted += 1
    return deleted, 0


def build_template(module: int) -> bytes:
    """File .xlsx mẫu: một sheet, cột Mã phiếu + cột đầu phiếu + cột dòng hàng."""
    import openpyxl
    from io import BytesIO
    adapter = DOC_ADAPTERS[module]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = adapter["sheet"]
    cols = [adapter["code"], *adapter["header_fields"], *adapter["line_fields"]]
    for i, f in enumerate(cols, start=1):
        ws.cell(row=1, column=i, value=f["header"])
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(14, len(f["header"]) + 2)
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
