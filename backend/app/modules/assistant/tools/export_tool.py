"""Hai tool XUẤT FILE của Trợ lý AI: `export_report_file` (Word .docx chuẩn DEGO) và
`export_excel_file` (bảng tính .xlsx) — người dùng xin "xuất Excel" mà chỉ có mỗi đường
Word là trợ lý trả sai định dạng, nên hai định dạng tách thành hai tool để model chọn
theo đúng chữ người dùng nói.

Khác các tool loại A (read-only), cặp tool này GHI hai thứ — một object lên storage và một
dòng `tab_file` (StoredFile) — nhưng không đụng chứng từ nghiệp vụ nào. File thuộc về NGƯỜI
HỎI (created_by), chỉ chủ file tải được qua `GET /api/assistant/files/{id}/download`; giao
diện chat đọc khóa `file` trên tool call và hiện nút "Tải báo cáo".

Dữ liệu trong file là chữ model đã tổng hợp TRONG hội thoại (kết quả các tool tra cứu đã
qua apply_scope) — tool này không tự tra thêm gì, nên không mở lối rò dữ liệu mới.
"""
import re
from io import BytesIO

from .base import ToolContext, ToolSpec

MAX_SECTIONS = 15
MAX_PARAS = 20          # đoạn văn mỗi mục
MAX_BULLETS = 30
MAX_COLS = 6
MAX_ROWS = 200
MAX_CELL = 300
MAX_TEXT = 2000

MAX_SHEETS = 5
MAX_XLSX_COLS = 15
MAX_XLSX_ROWS = 500

DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_PARAMS = {
    "type": "object",
    "properties": {
        "filename": {
            "type": "string",
            "description": "Tên tệp không dấu, không đuôi (vd: bao-cao-chi-tieu-thang-7).",
        },
        "title": {"type": "string", "description": "Tiêu đề lớn của báo cáo — bắt buộc."},
        "subtitle": {"type": "string", "description": "Phụ đề ngắn dưới tiêu đề (nếu có)."},
        "meta": {
            "type": "array",
            "description": "Thông tin chung dạng nhãn-giá trị (Kỳ báo cáo, Phạm vi, Người lập...).",
            "items": {
                "type": "object",
                "properties": {"label": {"type": "string"}, "value": {"type": "string"}},
                "required": ["label", "value"],
            },
        },
        "summary": {
            "type": "array",
            "description": "3-6 gạch đầu dòng tóm tắt cho hộp TL;DR đầu báo cáo.",
            "items": {"type": "string"},
        },
        "sections": {
            "type": "array",
            "description": f"Các mục nội dung theo thứ tự (tối đa {MAX_SECTIONS}).",
            "items": {
                "type": "object",
                "properties": {
                    "heading": {"type": "string", "description": "Tên mục — bắt buộc."},
                    "paragraphs": {
                        "type": "array", "items": {"type": "string"},
                        "description": "Các đoạn văn của mục.",
                    },
                    "bullets": {
                        "type": "array", "items": {"type": "string"},
                        "description": "Các gạch đầu dòng của mục.",
                    },
                    "table": {
                        "type": "object",
                        "description": f"Bảng số liệu (tối đa {MAX_COLS} cột x {MAX_ROWS} dòng); "
                                       "cột STT tự thêm, đừng khai.",
                        "properties": {
                            "columns": {"type": "array", "items": {"type": "string"}},
                            "rows": {
                                "type": "array",
                                "items": {"type": "array", "items": {"type": "string"}},
                            },
                        },
                        "required": ["columns", "rows"],
                    },
                },
                "required": ["heading"],
            },
        },
    },
    "required": ["title", "sections"],
}

_DESC = (
    "XUẤT nội dung đã tổng hợp trong hội thoại thành file báo cáo Word (.docx) định dạng chuẩn "
    "DEGO để người dùng tải về. Gọi khi người dùng muốn 'tạo file báo cáo', 'lưu thành Word' — "
    "báo cáo TRÌNH BÀY văn bản có đoạn văn, gạch đầu dòng. Người dùng muốn EXCEL / bảng tính / "
    ".xlsx thì gọi export_excel_file, ĐỪNG dùng tool này. Đổ số liệu THẬT đã tra cứu được trong "
    "hội thoại vào sections (đoạn văn, gạch đầu dòng, bảng) — không bịa số. Trong chuỗi có thể "
    "dùng <strong>chữ đậm</strong> và <br>. Sau khi gọi, báo người dùng bấm nút 'Tải báo cáo' "
    "ngay dưới câu trả lời để lưu file về máy."
)


def _clean(value, limit: int = MAX_TEXT) -> str:
    return str(value or "").strip()[:limit]


def _clean_sections(raw_sections: list) -> list[dict]:
    sections = []
    for raw in raw_sections[:MAX_SECTIONS]:
        if not isinstance(raw, dict):
            continue
        heading = _clean(raw.get("heading"), 200)
        if not heading:
            continue
        sec: dict = {"heading": heading, "paragraphs": [], "bullets": [], "table": None}
        for p in (raw.get("paragraphs") or [])[:MAX_PARAS]:
            text = _clean(p)
            if text:
                sec["paragraphs"].append(text)
        for b in (raw.get("bullets") or [])[:MAX_BULLETS]:
            text = _clean(b)
            if text:
                sec["bullets"].append(text)
        table = raw.get("table")
        if isinstance(table, dict):
            cols = [_clean(c, 100) for c in (table.get("columns") or [])[:MAX_COLS]]
            cols = [c for c in cols if c]
            rows = []
            for row in (table.get("rows") or [])[:MAX_ROWS]:
                if isinstance(row, list):
                    #  Kẹp mỗi dòng về đúng số cột — model hay trả dòng lệch cột.
                    values = [_clean(v, MAX_CELL) for v in row[:len(cols)]]
                    values += [""] * (len(cols) - len(values))
                    rows.append(values)
            if cols and rows:
                sec["table"] = {"columns": cols, "rows": rows}
        if sec["paragraphs"] or sec["bullets"] or sec["table"]:
            sections.append(sec)
    return sections


def render_docx(spec: dict) -> bytes:
    """Dựng file .docx từ spec ĐÃ chuẩn hóa. Tách riêng để test không cần DB/storage."""
    from datetime import datetime

    from ..report import docx_blocks as blk

    doc = blk.new_document()
    blk.header(doc, ["BÁO CÁO NỘI BỘ", "Trích xuất bởi Trợ lý AI",
                     f"Ngày lập: {datetime.now():%d/%m/%Y}"])
    blk.title(doc, spec["title"])
    if spec.get("subtitle"):
        blk.subtitle(doc, spec["subtitle"])
    if spec.get("meta"):
        blk.meta_table(doc, spec["meta"])
    if spec.get("summary"):
        blk.tldr_box(doc, "TÓM TẮT NHANH", spec["summary"])
    for i, sec in enumerate(spec["sections"], start=1):
        blk.section_bar(doc, f"{i}. {sec['heading']}")
        for text in sec["paragraphs"]:
            blk.paragraph(doc, text)
        for text in sec["bullets"]:
            blk.bullet(doc, text)
        if sec["table"]:
            blk.data_table(doc, sec["table"]["columns"], sec["table"]["rows"])
    blk.note(doc, "Số liệu do Trợ lý AI tổng hợp từ dữ liệu hệ thống tại thời điểm lập — "
                  "đối chiếu lại trên màn hình báo cáo khi cần chốt số.")
    blk.footer(doc, "DEGO HOLDING · Tài liệu nội bộ",
               "Sinh tự động bởi Trợ lý AI — Công cụ thu mua")

    buf = BytesIO()
    doc.save(buf)
    return buf.getvalue()


def _store_report_file(ctx: ToolContext, data: bytes, raw_filename, ext: str, mime: str,
                       total: int) -> dict:
    """Lưu file đã dựng lên storage + tạo StoredFile của NGƯỜI HỎI, trả khối kết quả chung.

    Key luôn nằm trong thư mục `assistant-report/` — điều kiện endpoint tải kiểm để không
    thành lối tải chung cho mọi file đính kèm khác của người đó.
    """
    from app.core.storage import dated_key, safe_name, upload_fileobj
    from app.modules.attachment.model import StoredFile

    base = safe_name(_clean(raw_filename, 80) or "bao-cao-tro-ly-ai")
    filename = base if base.lower().endswith(ext) else f"{base}{ext}"
    #  Tạo bản ghi trước (flush lấy id) để đặt key theo cấu trúc chung của storage —
    #  giống hệt _store_one của module attachment.
    sf = StoredFile(filename=filename, file_key="", url="", content_type=mime,
                    size=len(data), created_by=ctx.user.id, updated_by=ctx.user.id)
    ctx.db.add(sf)
    ctx.db.flush()
    key = dated_key("assistant-report", filename, sf.id)
    upload_fileobj(BytesIO(data), key, mime)
    sf.file_key = key
    ctx.db.commit()

    return {
        "status": "created",
        "file": {
            "id": sf.id,
            "filename": filename,
            "size": len(data),
            "download_url": f"/api/assistant/files/{sf.id}/download",
        },
        "total": total,
        # Nhắc model khỏi tự chế link markdown — nút tải do giao diện chat dựng.
        "reminder": "File đã sẵn sàng. Báo người dùng bấm nút 'Tải báo cáo' ngay dưới câu "
                    "trả lời để lưu về máy — ĐỪNG tự chèn đường link.",
    }


def _run(ctx: ToolContext, args: dict) -> dict:
    title = _clean(args.get("title"), 200)
    raw_sections = args.get("sections")
    if not title or not isinstance(raw_sections, list) or not raw_sections:
        return {"error": "Thiếu title hoặc sections — gom đủ nội dung rồi gọi lại."}

    sections = _clean_sections(raw_sections)
    if not sections:
        return {"error": "Không có mục nội dung hợp lệ nào (mỗi mục cần heading + nội dung)."}

    meta = []
    for m in (args.get("meta") or [])[:10]:
        if isinstance(m, dict) and _clean(m.get("label"), 100):
            meta.append((_clean(m.get("label"), 100), _clean(m.get("value"), 500)))

    spec = {
        "title": title,
        "subtitle": _clean(args.get("subtitle"), 300),
        "meta": meta,
        "summary": [_clean(s, 500) for s in (args.get("summary") or [])[:8] if _clean(s)],
        "sections": sections,
    }
    data = render_docx(spec)
    return _store_report_file(ctx, data, args.get("filename"), ".docx", DOCX_MIME,
                              total=len(sections))


EXPORT_REPORT_FILE_SPEC = ToolSpec(
    name="export_report_file",
    description=_DESC,
    parameters=_PARAMS,
    handler=_run,
)


# ── Tool 2: xuất Excel (.xlsx) ─────────────────────────────────────────────────────────
_XLSX_PARAMS = {
    "type": "object",
    "properties": {
        "filename": {
            "type": "string",
            "description": "Tên tệp không dấu, không đuôi (vd: tien-do-mua-hang-thang-8).",
        },
        "sheets": {
            "type": "array",
            "description": f"Các sheet dữ liệu theo thứ tự (tối đa {MAX_SHEETS}).",
            "items": {
                "type": "object",
                "properties": {
                    "name": {"type": "string",
                             "description": "Tên sheet, tối đa 31 ký tự (vd 'Tien do thang 8')."},
                    "columns": {"type": "array", "items": {"type": "string"},
                                "description": f"Tiêu đề cột (tối đa {MAX_XLSX_COLS})."},
                    "rows": {
                        "type": "array",
                        "description": f"Dữ liệu: mỗi dòng một mảng ô theo đúng thứ tự cột "
                                       f"(tối đa {MAX_XLSX_ROWS} dòng). Ô SỐ điền chuỗi số trần "
                                       "kiểu '40555800' — tool tự đổi sang kiểu số của Excel.",
                        # Khai string để cả Claude lẫn Gemini nuốt được schema (Gemini không
                        # nhận union type) — số thật được nhận diện lại ở `_to_cell`.
                        "items": {"type": "array", "items": {"type": "string"}},
                    },
                },
                "required": ["name", "columns", "rows"],
            },
        },
    },
    "required": ["sheets"],
}

_XLSX_DESC = (
    "XUẤT số liệu đã tổng hợp trong hội thoại thành file EXCEL (.xlsx) để người dùng tải về — "
    "mỗi sheet một bảng cột x dòng. Gọi khi người dùng muốn 'xuất Excel', 'file xlsx', 'bảng "
    "tính', 'xuất dữ liệu dạng bảng'. Đổ số liệu THẬT đã tra cứu trong hội thoại vào rows — "
    "không bịa; ô SỐ (tiền, số lượng, đơn giá) điền số trần '40555800', KHÔNG kèm 'đ' hay dấu "
    "ngăn nghìn để Excel cộng/lọc được. Người dùng muốn báo cáo Word trình bày văn bản thì dùng "
    "export_report_file. Sau khi gọi, báo người dùng bấm nút 'Tải báo cáo' ngay dưới câu trả "
    "lời để lưu file về máy."
)

_NUM_RE = re.compile(r"^-?\d+(\.\d+)?$")
_SHEET_BAD = re.compile(r"[\[\]:*?/\\]")   # ký tự Excel cấm trong tên sheet


def _to_cell(v):
    """Ô Excel: giữ số là SỐ (kể cả model gõ '40555800' dạng chuỗi) để cộng/lọc/pivot được."""
    if isinstance(v, bool):
        return "Có" if v else "Không"
    if isinstance(v, (int, float)):
        return v
    s = _clean(v, MAX_CELL)
    if _NUM_RE.match(s):
        return float(s) if "." in s else int(s)
    return s


def _clean_sheets(raw_sheets: list) -> list[dict]:
    sheets: list[dict] = []
    used: set[str] = set()
    for idx, raw in enumerate(raw_sheets[:MAX_SHEETS], start=1):
        if not isinstance(raw, dict):
            continue
        cols = [_clean(c, 100) for c in (raw.get("columns") or [])[:MAX_XLSX_COLS]]
        cols = [c for c in cols if c]
        rows = []
        for row in (raw.get("rows") or [])[:MAX_XLSX_ROWS]:
            if isinstance(row, list):
                #  Kẹp mỗi dòng về đúng số cột — model hay trả dòng lệch cột.
                values = [_to_cell(v) for v in row[:len(cols)]]
                values += [""] * (len(cols) - len(values))
                rows.append(values)
        if not cols or not rows:
            continue
        name = _SHEET_BAD.sub(" ", _clean(raw.get("name"), 60)).strip()[:31] or f"Sheet{idx}"
        if name.lower() in used:   # tên sheet trùng thì openpyxl nổ — đánh số phân biệt
            name = f"{name[:28]}-{idx}"
        used.add(name.lower())
        sheets.append({"name": name, "columns": cols, "rows": rows})
    return sheets


def render_xlsx(sheets: list[dict]) -> bytes:
    """Dựng file .xlsx từ danh sách sheet ĐÃ chuẩn hóa. Tách riêng để test không cần DB/storage."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    from app.core.export_xlsx import HEAD_BG, HEAD_TEXT

    wb = Workbook()
    wb.remove(wb.active)
    head_font = Font(bold=True, color=HEAD_TEXT)
    head_fill = PatternFill("solid", fgColor=HEAD_BG)
    for sheet in sheets:
        ws = wb.create_sheet(title=sheet["name"])
        for i, label in enumerate(sheet["columns"], start=1):
            c = ws.cell(row=1, column=i, value=label)
            c.font = head_font
            c.fill = head_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        widths = [len(str(label)) for label in sheet["columns"]]
        for r, row in enumerate(sheet["rows"], start=2):
            for i, v in enumerate(row, start=1):
                c = ws.cell(row=r, column=i, value=v)
                if isinstance(v, (int, float)):
                    c.number_format = "#,##0.####"   # ngăn nghìn, đơn giá giữ tối đa 4 lẻ
                    c.alignment = Alignment(horizontal="right")
                widths[i - 1] = max(widths[i - 1], len(str(v)))
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 10), 45)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(sheet['columns']))}{len(sheet['rows']) + 1}"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _run_excel(ctx: ToolContext, args: dict) -> dict:
    raw_sheets = args.get("sheets")
    if not isinstance(raw_sheets, list) or not raw_sheets:
        return {"error": "Thiếu sheets — gom đủ số liệu dạng bảng rồi gọi lại."}
    sheets = _clean_sheets(raw_sheets)
    if not sheets:
        return {"error": "Không có sheet hợp lệ nào (mỗi sheet cần columns + rows)."}
    data = render_xlsx(sheets)
    return _store_report_file(ctx, data, args.get("filename"), ".xlsx", XLSX_MIME,
                              total=sum(len(s["rows"]) for s in sheets))


EXPORT_EXCEL_FILE_SPEC = ToolSpec(
    name="export_excel_file",
    description=_XLSX_DESC,
    parameters=_XLSX_PARAMS,
    handler=_run_excel,
)
