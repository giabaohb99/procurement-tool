"""Xuất Excel (.xlsx) cho các màn danh sách — CR-068.

Dùng chung cho Yêu cầu mua hàng, Yêu cầu báo giá, Đơn mua hàng và Tiến độ mua hàng:
mỗi màn khai báo bộ cột (`Col`) rồi gọi `xlsx_response`. Quy ước chung:

- Số giữ nguyên KIỂU SỐ (không kèm chữ "đ") để cộng / lọc / pivot ngay trong Excel.
- Ngày lưu dạng chuỗi 'YYYY-MM-DD' được đổi sang kiểu ngày của Excel (hiện dd/mm/yyyy),
  còn `created_at` là mốc UTC nên phải quy về giờ VN cho khớp màn hình.
- Dòng tiêu đề in đậm, CHỮ ĐEN trên nền xanh nhạt, đóng băng để cuộn vẫn thấy.

Phạm vi dữ liệu và việc che cột nhạy cảm KHÔNG xử lý ở đây — mỗi màn tự dựng câu truy vấn
đã qua `apply_scope` rồi mới đưa dòng vào đây.
"""
from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO
from typing import Any, Iterable, NamedTuple

from fastapi import HTTPException
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MAX_ROWS = 5000
"""Trần số dòng cho một lần xuất — vượt thì bắt lọc bớt hoặc tự tick chọn."""

VN_OFFSET = timedelta(hours=7)

#  DÒNG TIÊU ĐỀ: chữ ĐEN trên nền xanh nhạt (khách báo 25/08/2026).
#  Bản cũ là chữ TRẮNG trên nền xanh đậm `1F3864`. Đẹp khi mọi thứ nguyên vẹn,
#  nhưng chữ trắng chỉ đọc được KHI CÒN NỀN: file đi qua một vòng chuyển đổi
#  (mở bằng Google Sheets / LibreOffice, dán sang bảng khác, xem trước trong
#  Outlook) là nền rơi trước, chữ trắng ở lại — người nhận thấy hàng tiêu đề
#  trắng trơn, phải bấm từng ô đọc thanh công thức mới biết cột nào là cột nào.
#  Chữ đen thì mất nền vẫn đọc được, nên chọn phía hỏng-mà-vẫn-dùng-được.
HEAD_TEXT = "000000"
HEAD_BG = "D9E2F3"

_FORMATS = {
    "money": "#,##0",          # tiền: thành tiền, tổng tiền
    "price": "#,##0.####",     # đơn giá giữ 4 số lẻ (theo migration d4b9e7c1a305)
    "qty": "#,##0.###",        # số lượng giữ 3 số lẻ
    "int": "#,##0",
    "date": "dd/mm/yyyy",
    "datetime": "dd/mm/yyyy hh:mm",
}
_RIGHT = {"money", "price", "qty", "int"}


class Col(NamedTuple):
    """Một cột trong file xuất. `kind` quyết định định dạng ô Excel."""

    key: str
    label: str
    kind: str = "text"   # text | money | price | qty | int | date | datetime | bool
    width: int = 16


def parse_ids(ids: str | None) -> list[int]:
    """'12, 15,x' -> [12, 15]. Rỗng -> [] (nghĩa là xuất theo bộ lọc, không theo tick chọn)."""
    if not ids:
        return []
    return [int(i.strip()) for i in ids.split(",") if i.strip().isdigit()]


def pick_columns(spec: Iterable[Col], wanted: str | None, always: Iterable[str] = ()) -> list[Col]:
    """Lọc bộ cột theo danh sách key người dùng đang hiện trên bảng (`cols=a,b,c`).

    Giữ ĐÚNG THỨ TỰ người dùng thấy trên màn hình. Key lạ thì bỏ qua (bảng đổi cột,
    localStorage cũ...). Không truyền gì -> lấy trọn bộ cột khai báo.
    `always` là các cột luôn xuất dù có hiện trên bảng hay không (vd cụm dòng hàng).
    """
    spec = list(spec)
    by_key = {c.key: c for c in spec}
    keys = [k.strip() for k in (wanted or "").split(",") if k.strip()]
    if not keys:
        return spec
    out = [by_key[k] for k in keys if k in by_key]
    seen = {c.key for c in out}
    out += [by_key[k] for k in always if k in by_key and k not in seen]
    return out or spec


def check_row_limit(n: int) -> None:
    if n > MAX_ROWS:
        raise HTTPException(400, f"Kết quả {n:,} dòng, vượt mức {MAX_ROWS:,} dòng cho một lần xuất — "
                                 f"hãy lọc bớt hoặc tự tick chọn các phiếu cần xuất")


def _to_date(v: Any):
    """Chuỗi 'YYYY-MM-DD' -> date của Excel. Giá trị lạ trả nguyên văn để không nuốt dữ liệu."""
    if isinstance(v, (datetime, date)):
        return v
    s = str(v or "").strip()
    if not s:
        return None
    try:
        return datetime.strptime(s[:10], "%Y-%m-%d").date()
    except ValueError:
        return s


def _to_datetime(v: Any):
    """`created_at` lưu UTC (naive) -> quy về giờ VN cho khớp với màn hình."""
    if isinstance(v, datetime):
        return v + VN_OFFSET
    return _to_date(v)


def cell_value(col: Col, row: dict) -> Any:
    v = row.get(col.key)
    if col.kind in ("money", "price", "qty", "int"):
        if v in (None, ""):
            return 0
        return float(v) if isinstance(v, Decimal) else v
    if col.kind == "date":
        return _to_date(v)
    if col.kind == "datetime":
        return _to_datetime(v)
    if col.kind == "bool":
        return "Có" if v else ""
    return "" if v is None else str(v)


def xlsx_response(filename: str, columns: list[Col], rows: list[dict], sheet_title: str = "") -> Response:
    """Dựng file .xlsx và trả về cho trình duyệt tải xuống.

    `filename` chưa gồm đuôi và ngày — hàm tự gắn `-DDMMYYYY.xlsx` theo ngày xuất (giờ VN),
    cùng lối đặt tên với phiếu in (CR-057).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_title or "Du lieu")[:31]

    head_font = Font(bold=True, color=HEAD_TEXT)
    head_fill = PatternFill("solid", fgColor=HEAD_BG)
    for i, col in enumerate(columns, start=1):
        c = ws.cell(row=1, column=i, value=col.label)
        c.font = head_font
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = col.width

    for r, row in enumerate(rows, start=2):
        for i, col in enumerate(columns, start=1):
            c = ws.cell(row=r, column=i, value=cell_value(col, row))
            fmt = _FORMATS.get(col.kind)
            if fmt:
                c.number_format = fmt
            if col.kind in _RIGHT:
                c.alignment = Alignment(horizontal="right")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(max(len(columns), 1))}{max(len(rows) + 1, 1)}"

    buf = BytesIO()
    wb.save(buf)
    stamp = (datetime.utcnow() + VN_OFFSET).strftime("%d%m%Y")
    name = f"{filename}-{stamp}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{name}"'},
    )
