"""Đồng bộ SẢN PHẨM/VTBB (tab_product) từ Excel sanpham.xlsx.

Mỗi dòng tab_product = 1 Mã VTBB (Nhãn / Chai / Nắp / Thùng...). Mã HH KHÔNG phải
dòng riêng — chỉ là field tham chiếu (hh_code/hh_name) trên dòng VTBB. Sheet HH dùng
làm bảng TRA CỨU để điền Tên SP (HH) + Tên pháp lý cho dòng nhãn (nhãn 1:1 với HH,
số HH nằm trong mã VTBB của nhãn).

Nguồn: /app/_sanpham_tmp.xlsx (bind-mount) hoặc biến SP_XLSX.

Bố cục:
  HH (tra cứu)  r2+ : Mã HH(1) TênNộiBộ(2) TênPhápLý(3)   -> map hh_code -> (ten, phap_ly)
  Nhãn          r3+ : Mã HH(1) Phân loại(2) Mã VTBB(3) Tên(4)
                       -> code=Mã VTBB, name=Tên, item_group=Phân loại,
                          hh_code=Mã HH, hh_name=HH.tênNộiBộ, legal_name=HH.tênPhápLý
  Chai và nắp   r3+ : Mã VTBB(1) Tên(2) Phân loại(3) [LINK ảnh(4) -> script ảnh riêng]
                       -> code/name/item_group (Chai/Nắp)
  Thùng         r3+ : Mã VTBB(1) Tên(2) Phân loại(3) [LINK ảnh(4)]
                       -> code/name/item_group='Thùng'
  NL            r2+ : Mã NL NB(1) Tên NL NB(2) ĐVT(3)
                       -> code/name, item_group='NL' (nguyên liệu), unit=ĐVT

Upsert theo `code` — KHÔNG xoá; chỉ ghi đè cột sheet có cấp (giữ nguyên cột khác).
Re-run an toàn.

Chạy: docker exec -w /app <api> python -m scripts.import_product
"""
import os
import re
import unicodedata

from openpyxl import load_workbook

import app.core.all_models  # noqa: F401  (đăng ký mapper)
from app.core.database import SessionLocal
from app.modules.product.model import Product
from app.modules.user.model import User


def _norm(s) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.replace("đ", "d")
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _txt(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.upper() in ("#N/A", "N/A", "NA", "-", "NULL") else s


# Gộp biến thể hoa/thường của Phân loại về nhãn chuẩn (theo controller product)
_GROUP_CANON = {
    "thung": "Thùng", "hop": "Hộp", "chai": "Chai", "nap": "Nắp", "tui": "Túi",
    "xo": "Xô", "can": "Can", "bao": "Bao", "coc": "Cóc", "lon": "Lon",
    "seal": "Seal", "bang keo": "Băng keo", "tem": "Tem", "nhan": "Nhãn",
    "nhan thung": "Nhãn thùng", "nhan phu": "Nhãn phụ",
    "nl": "NL", "nguyen lieu": "NL", "vt": "VT", "ndt": "NDT", "icare": "ICARE",
}


def _group(v) -> str:
    s = _txt(v)
    if s.startswith("[") and s.endswith("]"):
        s = s[1:-1].strip()
    s = re.sub(r"\s+", " ", s).strip()
    return _GROUP_CANON.get(_norm(s), s)


def _xlsx_path() -> str:
    p = os.environ.get("SP_XLSX", "/app/_sanpham_tmp.xlsx")
    if not os.path.exists(p):
        alt = os.path.join(os.path.dirname(__file__), "..", "_sanpham_tmp.xlsx")
        if os.path.exists(alt):
            return alt
    return p


def _upsert(db, by_code, uid, code, fields, stats):
    """Upsert 1 dòng theo code; chỉ set cột có giá trị (không xoá cột khác)."""
    code = _txt(code)
    if not code:
        return
    p = by_code.get(code)
    if p is None:
        p = Product(code=code, name=_txt(fields.get("name")) or code,
                    is_active=True, created_by=uid, updated_by=uid)
        for k, v in fields.items():
            if v:
                setattr(p, k, v)
        db.add(p)
        by_code[code] = p
        stats["created"] += 1
    else:
        for k, v in fields.items():
            if v:
                setattr(p, k, v)
        p.updated_by = uid
        stats["updated"] += 1


def run():
    path = _xlsx_path()
    wb = load_workbook(path, data_only=True)

    db = SessionLocal()
    try:
        sys_user = db.query(User).order_by(User.id).first()
        uid = sys_user.id if sys_user else 1

        by_code = {p.code: p for p in db.query(Product).all()}
        print(f"tab_product hien co: {len(by_code)} dong")

        # ---- Bảng tra cứu HH: hh_code -> (ten_noi_bo, ten_phap_ly) ----
        ws = wb["HH"]
        hh_map = {}
        for r in range(2, ws.max_row + 1):
            hh = _txt(ws.cell(r, 1).value)
            if hh:
                hh_map[hh] = (_txt(ws.cell(r, 2).value), _txt(ws.cell(r, 3).value))
        print(f"HH tra cuu: {len(hh_map)} ma")

        # ---- Nhãn: có Mã HH -> điền hh_name + legal_name từ HH ----
        ws = wb["Nhãn"]
        st = {"created": 0, "updated": 0}
        miss_hh = 0
        for r in range(3, ws.max_row + 1):
            code = _txt(ws.cell(r, 3).value)
            if not code:
                continue
            hh_code = _txt(ws.cell(r, 1).value)
            hh_name, hh_legal = hh_map.get(hh_code, ("", ""))
            if hh_code and hh_code not in hh_map:
                miss_hh += 1
            _upsert(db, by_code, uid, code, {
                "name": _txt(ws.cell(r, 4).value),
                "item_group": _group(ws.cell(r, 2).value) or "Nhãn",
                "hh_code": hh_code,
                "hh_name": hh_name,
                "legal_name": hh_legal,
            }, st)
        db.commit()
        print(f"[Nhãn]        +{st['created']} moi / {st['updated']} cap nhat"
              + (f"  (|{miss_hh} nhan co Ma HH khong thay trong sheet HH)" if miss_hh else ""))

        # ---- Chai và nắp + Thùng (không có Mã HH trong sheet) ----
        for sheet, default_grp in (("Chai và nắp", ""), ("Thùng", "Thùng")):
            ws = wb[sheet]
            st = {"created": 0, "updated": 0}
            for r in range(3, ws.max_row + 1):
                code = _txt(ws.cell(r, 1).value)
                if not code:
                    continue
                _upsert(db, by_code, uid, code, {
                    "name": _txt(ws.cell(r, 2).value),
                    "item_group": _group(ws.cell(r, 3).value) or default_grp,
                }, st)
            db.commit()
            print(f"[{sheet:11s}] +{st['created']} moi / {st['updated']} cap nhat")

        # ---- NL (Nguyên liệu): Mã NL NB(1) Tên NL NB(2) ĐVT(3), data từ row 2 ----
        if "NL" in wb.sheetnames:
            ws = wb["NL"]
            st = {"created": 0, "updated": 0}
            for r in range(2, ws.max_row + 1):
                code = _txt(ws.cell(r, 1).value)
                if not code:
                    continue
                _upsert(db, by_code, uid, code, {
                    "name": _txt(ws.cell(r, 2).value),
                    "item_group": "NL",
                    "unit": _txt(ws.cell(r, 3).value),
                }, st)
            db.commit()
            print(f"[NL         ] +{st['created']} moi / {st['updated']} cap nhat")

        # ---- Tổng kết ----
        from sqlalchemy import func
        print("=" * 60)
        print("PHAN LOAI sau dong bo:")
        for g, c in db.query(Product.item_group, func.count()).group_by(
                Product.item_group).order_by(func.count().desc()).all():
            print(f"  [{c:5d}] {g!r}")
        print(f"TONG tab_product: {db.query(Product).count()}")
        print("=" * 60)
    finally:
        db.close()


if __name__ == "__main__":
    run()
