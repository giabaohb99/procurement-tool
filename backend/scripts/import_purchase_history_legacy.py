"""import_purchase_history_legacy — Đổ LỊCH SỬ MUA HÀNG CŨ (khaosatsanpham.xlsx) vào
bảng `tab_purchase_history` với `source='legacy'`.

Đây là các lần mua CÓ THẬT của giai đoạn trước khi có hệ thống, không có Đơn mua hàng
tương ứng → ghi thẳng vào bảng lịch sử (po_item_id/po_code để trống), phục vụ tab
"Lịch sử mua hàng" ở chi tiết Sản phẩm và chi tiết NCC.

HAI SHEET, HAI CẤU TRÚC KHÁC NHAU:
  Sheet1 (VTBB — có phân loại): NGÀY GIAO | NCC/KHO XUẤT | PHÂN LOẠI | MÃ VTBB | TÊN |
      ĐVT | SL NHẬP | MOQ | ĐƠN GIÁ CHƯA VAT | VAT (tiền) | ĐƠN GIÁ CÓ VAT
  Sheet2 (Nguyên liệu): TÊN NCC | MÃ HÀNG HÓA | TÊN HÀNG HÓA | SL ĐẶT | ĐƠN GIÁ |
      NGÀY NHẬN HÀNG | ĐVT   (không có VAT, không có MOQ)

NGÀY THÁNG — CẢNH BÁO QUAN TRỌNG (Sheet2):
  File được nhập trên Excel ĐỊNH DẠNG MỸ (m/d/yyyy). Người nhập gõ kiểu Việt (d/m/yyyy),
  nên ô nào ngày <= 12 thì Excel nuốt được nhưng HOÁN ĐỔI ngày<->tháng, ô nào ngày > 12
  thì Excel bó tay và để nguyên dạng chuỗi "17/01/2024".
  Bằng chứng: trong 609 ô kiểu ngày của Sheet2 KHÔNG ô nào có ngày > 12 (xác suất ~0 nếu
  dữ liệu đúng), còn 736/736 ô dạng chuỗi đều có phần đầu > 12. Sau khi hoán đổi lại thì
  không còn ngày nào rơi vào tương lai.
  => Sheet2: ô kiểu datetime phải ĐỔI LẠI ngày<->tháng; ô chuỗi đọc d/m/yyyy.
  => Sheet1 KHÔNG dính lỗi này (3130 ô có ngày > 12 => Excel đọc đúng locale) — giữ nguyên.

BỐN CHỖ DỮ LIỆU BẨN KHÁC ĐÃ XỬ LÝ (chi tiết ở ngay trên đoạn code tương ứng):
  1. Đơn giá Sheet2 ghi lẫn nghìn đồng / đồng  -> quy về đồng theo ngưỡng độ lớn
  2. Số lượng vô lý (>= 1.000.000)             -> BỎ dòng, xuất ls_sl_bat_thuong.csv
  3. Ngày gõ sai đủ kiểu                       -> sửa theo dòng lân cận, ls_ngay_sua.csv
  4. NCC / mã hàng chưa có trong danh mục      -> tạo mới, ls_ncc_tao_moi + ls_ma_tao_moi

PHÁP NHÂN: file không có cột công ty → gán hết cho DEGO (COMPANY_ID=1, theo yêu cầu user).

IDEMPOTENT: mỗi lần chạy XÓA SẠCH dòng `source='legacy'` rồi nạp lại. Dòng do hệ thống tự
chốt (`source='system'`) KHÔNG bị đụng tới. NCC/mã hàng đã tạo ở lần chạy trước thì lần sau
khớp được ngay nên không tạo trùng.

Chạy trong container:
  DRY_RUN=1 docker compose exec -T api python -m scripts.import_purchase_history_legacy
  docker compose exec -T api python -m scripts.import_purchase_history_legacy

ENV: XLSX (mặc định /app/_ks_goc.xlsx), COMPANY_ID (1), DRY_RUN (0/1), REPORT_DIR (/app)
"""
import bisect
import csv
import json
import os
import re
import unicodedata
from collections import Counter
from datetime import datetime

import openpyxl
from sqlalchemy import text

from app.core import all_models  # noqa: F401  (đăng ký mọi mapper)
from app.core.database import SessionLocal
from app.core.utils import generate_code
from app.modules.company.model import Company
from app.modules.product.model import Product
from app.modules.purchase_history.model import PurchaseHistory
from app.modules.supplier.model import Supplier

XLSX = os.getenv("XLSX", "/app/_ks_goc.xlsx")
COMPANY_ID = int(os.getenv("COMPANY_ID", "1"))
DRY_RUN = os.getenv("DRY_RUN", "0") == "1"
REPORT_DIR = os.getenv("REPORT_DIR", "/app")

# Kho/VP nội bộ — KHÔNG phải NCC, bỏ mọi dòng của chúng (giống import_survey_history).
# So khớp qua `sup_key` nên khai báo bằng tên gốc, khóa được tính ở dưới.
SKIP_SUPPLIERS_RAW = ("KHO VTBB SX", "VP B19")

# NCC tạo mới dùng chung tiền tố với import_survey_history — cùng nguồn khaosatsanpham.xlsx,
# nhìn mã là biết NCC sinh ra từ file này chứ không phải người dùng tự thêm trên UI.
NEW_SUPPLIER_PREFIX = "NCCKS"

_BAD = {"#N/A", "N/A", "NA", "-", "NULL", "#REF!", "NONE"}


def nfc(s):
    return unicodedata.normalize("NFC", (str(s) if s is not None else "").strip())


def clean_text(s):
    s = nfc(s)
    return "" if s.upper() in _BAD else s


def name_key(s):
    """Khóa so khớp TÊN hàng hóa — dùng để cứu dòng gõ sai mã (xem phần mã hàng ở main)."""
    return re.sub(r"\s+", " ", nfc(s)).lower()


# Người nhập file hay viết tắt ("CTY TNHH SX TM DV HẢI BÌNH") trong khi danh mục NCC ghi
# đầy đủ ("CÔNG TY TNHH SẢN XUẤT THƯƠNG MẠI DỊCH VỤ HẢI BÌNH"). Bung viết tắt ở CẢ HAI phía
# để so khớp CHÍNH XÁC, an toàn hơn nhiều so với đoán gần đúng theo chuỗi con.
ABBR = {
    "sx": "sản xuất", "tm": "thương mại", "dv": "dịch vụ", "dvu": "dịch vụ",
    "xnk": "xuất nhập khẩu", "mtv": "một thành viên", "cp": "cổ phần",
    "đt": "đầu tư", "kd": "kinh doanh", "tnhh": "", "và": "",
}


def sup_key(s):
    """Khóa so khớp NCC: bỏ loại hình công ty, bung viết tắt, gộp khoảng trắng."""
    s = nfc(s).lower()
    for w in ("công ty trách nhiệm hữu hạn", "công ty", "cty", "c.ty"):
        s = s.replace(w, " ")
    s = re.sub(r"[.,()\-/]", " ", s)
    toks = [ABBR.get(t, t) for t in s.split()]
    return re.sub(r"\s+", " ", " ".join(toks)).strip()


SKIP_SUPPLIERS = {sup_key(x) for x in SKIP_SUPPLIERS_RAW}

# Số lượng vô lý. 6 dòng của NCC "MIA" ghi SL 2.000.000–5.800.000 cho hoạt chất mà chính mã
# hàng đó mua ở NCC khác chỉ 2–600 Kg, ĐVT lại bỏ trống; 6 dòng này chiếm 97,5% tổng giá trị
# của cả file. Dòng lớn thứ hai trong file chỉ 117.400 (nhãn, hợp lý) nên ngưỡng 1.000.000
# tách sạch, không đụng dòng thật. Người dùng đã chốt: BỎ hẳn thay vì đoán lại đơn vị tính.
QTY_MAX = 1_000_000


def parse_money(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = nfc(v).replace("đ", "").replace("₫", "").replace(" ", "")
    if not s:
        return 0.0
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "." in s:
        s = s.replace(".", "")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_date(v, swap_dm=False):
    """Trả 'YYYY-MM-DD' hoặc '' nếu không đọc được. `swap_dm`: xem docstring đầu file."""
    if not v:
        return ""
    if isinstance(v, datetime):
        if swap_dm and v.day <= 12:
            try:
                return datetime(v.year, v.day, v.month).strftime("%Y-%m-%d")
            except ValueError:
                pass
        return v.strftime("%Y-%m-%d")
    s = nfc(v)
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            d = datetime.strptime(s, fmt)
            return d.strftime("%Y-%m-%d") if 2000 <= d.year <= 2100 else ""
        except ValueError:
            pass
    return ""


# ── Sửa ngày gõ sai ──────────────────────────────────────────────────────────────────
# Cột ngày gõ tay nên hỏng đủ kiểu, đã gặp thật trong file:
#   '#VALUE!' · ô trống              -> mất trắng
#   'Dự trù 17/11' · '30/03'         -> thiếu năm
#   '21//01/2026' · '29/03/03/2025'  -> thừa dấu ngăn / thừa cụm
#   '28/012026' · '2602/2026'        -> thiếu dấu ngăn
#   '27/01/25024' · '08/07/0206'     -> gõ lộn phần năm
#   ô ngày Excel ghi năm 2525 / 2015 -> gõ lộn năm ngay trong ô kiểu ngày
# CẢ HAI SHEET ĐỀU SẮP THEO NGÀY (đã kiểm: các dòng kề nhau lệch nhau vài ngày), nên dòng
# lân cận là căn cứ đáng tin để suy ra năm còn thiếu và để lấp ô mất trắng.
YEAR_MIN, YEAR_MAX = 2015, 2030


def _fix_year(y, near):
    """Chuẩn hóa phần năm; `near` = năm của dòng lân cận. Trả None nếu chịu."""
    y = (y or "").strip()
    if y.isdigit():
        if len(y) == 4 and YEAR_MIN <= int(y) <= YEAR_MAX:
            return int(y)
        if len(y) == 2:
            return 2000 + int(y)
        if len(y) == 5:                       # gõ thừa 1 chữ số: '25024' -> 2024
            cand = sorted({int(y[:i] + y[i + 1:]) for i in range(5)}
                          & set(range(YEAR_MIN, YEAR_MAX + 1)))
            if len(cand) == 1:
                return cand[0]
            if cand and near:
                return min(cand, key=lambda c: abs(c - near))
    return near


def _repair_date_str(v, near_year):
    """Đọc lại ô ngày hỏng theo kiểu d/m/y. Trả 'YYYY-MM-DD' hoặc ''."""
    s = re.sub(r"[^\d/\-.]", " ", nfc(v))     # bỏ chữ ('Dự trù'), giữ số và dấu ngăn
    s = re.sub(r"[.\-]", "/", s)
    parts = [p.strip() for p in re.split(r"/+", s) if p.strip()]
    if not parts:
        return ""
    if len(parts) == 1 and len(parts[0]) == 4:          # '2602' -> 26/02, thiếu năm
        parts = [parts[0][:2], parts[0][2:]]
    elif len(parts) == 2 and len(parts[0]) == 4:        # '2602/2026'
        parts = [parts[0][:2], parts[0][2:], parts[1]]
    elif len(parts) == 2 and len(parts[1]) > 2:         # '28/012026'
        parts = [parts[0], parts[1][:2], parts[1][2:]]
    if len(parts) < 2:
        return ""
    year = _fix_year(parts[-1], near_year) if len(parts) > 2 else near_year
    if not year:
        return ""
    try:
        return datetime(year, int(parts[1]), int(parts[0])).strftime("%Y-%m-%d")
    except ValueError:
        return ""


def build_date_map(ws, col, swap):
    """Đọc cả cột ngày của 1 sheet rồi sửa các ô hỏng.

    Trả (map {dòng excel -> 'YYYY-MM-DD'}, map {dòng -> (ô gốc, ngày sau sửa, cách suy)}).

    Căn cứ suy luận LẤY TỪ BẢN ĐỌC THẲNG (chưa sửa) nên không sai dây chuyền, và lấy theo
    CẢ MỘT VÙNG lân cận chứ không lấy một dòng: hai ô gõ sai nằm sát nhau thì nếu chỉ nhìn
    một dòng chúng sẽ "làm chứng" cho nhau (dòng đúng bị kéo theo dòng sai). Lấy năm PHỔ
    BIẾN NHẤT của WINDOW dòng gần nhất thì một vài ô sai không lật được kết quả.
    """
    raw = {i: ws.cell(row=i, column=col).value for i in range(2, ws.max_row + 1)}
    base = {i: parse_date(v, swap_dm=swap) for i, v in raw.items()}
    ok = sorted(i for i, d in base.items() if d and YEAR_MIN <= int(d[:4]) <= YEAR_MAX)
    WINDOW = 7                                       # số dòng lấy về mỗi phía

    def near_rows(i):
        """Các dòng đọc được gần dòng i nhất (KHÔNG tính chính nó), gần trước xa sau."""
        p = bisect.bisect_left(ok, i)
        left = [r for r in ok[max(0, p - WINDOW):p] if r != i]
        right = [r for r in ok[p:p + WINDOW + 1] if r != i]
        return sorted(left + right, key=lambda r: abs(r - i))

    out, fixed = dict(base), {}
    for i, v in raw.items():
        d = base[i]
        near = near_rows(i)
        yc = Counter(int(base[r][:4]) for r in near)
        ny = yc.most_common(1)[0][0] if yc else None
        if d and YEAR_MIN <= int(d[:4]) <= YEAR_MAX and (not ny or abs(int(d[:4]) - ny) < 2):
            continue                                 # đọc thẳng đã hợp lý -> để yên
        if d and ny:
            # đọc được nhưng năm vô lý (2525) hoặc lệch hẳn cả vùng (2015 giữa vùng 2025)
            # -> chỉ sai phần năm, giữ nguyên ngày/tháng người ta gõ
            new, how = f"{ny}{d[4:]}", "sửa năm theo vùng lân cận"
            # ...trừ khi sửa xong vẫn lệch quá xa vùng: cả hai sheet đều sắp theo ngày nên
            # ô nằm giữa một khối cùng ngày mà lại cách khối vài tháng thì phần ngày/tháng
            # cũng gõ sai nốt -> lấy luôn ngày của khối, sai vài ngày còn hơn sai vài tháng
            nr = near[0]
            if abs((datetime.strptime(new, "%Y-%m-%d")
                    - datetime.strptime(base[nr], "%Y-%m-%d")).days) > 45:
                new, how = base[nr], f"lệch quá xa vùng, lấy ngày dòng {nr}"
        else:
            new, how = _repair_date_str(v, ny), "đọc lại ô gõ sai"
            if not new and near:
                # ô mất trắng -> mượn ngày của dòng gần nhất CÓ CÙNG NĂM với cả vùng
                nr = next((r for r in near if int(base[r][:4]) == ny), near[0])
                new, how = base[nr], f"lấy ngày dòng {nr} (dòng gần nhất có ngày)"
        out[i] = new
        if new:
            fixed[i] = (repr(v)[:40], new, how)
    return out, fixed


def read_sheet1(wb):
    """VTBB — có phân loại, có VAT. Bỏ dòng kho nội bộ, dòng giá rỗng, dòng SL vô lý."""
    ws = wb["Sheet1"]
    dmap, dfix = build_date_map(ws, 2, swap=False)
    out, st, bad_qty = [], Counter(), []
    for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not r or all(c is None for c in r[:12]):
            continue
        code, sup = nfc(r[4]), nfc(r[2])
        if not code or not sup:
            continue
        if sup_key(sup) in SKIP_SUPPLIERS:
            st["kho_noi_bo"] += 1
            continue
        price = parse_money(r[9])
        if price <= 0:
            st["khong_gia"] += 1
            continue
        vat_amt = parse_money(r[10])
        vat_pct = round(vat_amt / price * 100) if price else 0
        if vat_pct < 0 or vat_pct > 30:
            vat_pct = 0
        qty = parse_money(r[7])
        if qty >= QTY_MAX:
            bad_qty.append(("Sheet1", i, sup, code, clean_text(r[5]), qty, price))
            continue
        price_vat = parse_money(r[11]) or price * (1 + vat_pct / 100)
        out.append({
            "sheet": "Sheet1", "row": i, "date": dmap.get(i, ""), "date_fix": dfix.get(i),
            "sup": sup, "code": code,
            "name": clean_text(r[5]), "unit": clean_text(r[6]), "qty": qty, "price": price,
            "vat": vat_pct, "amount": round(qty * price_vat, 2),
            "group": clean_text(r[3]), "moq": parse_money(r[8]),
        })
    return out, st, bad_qty


# Sheet2 ghi ĐƠN GIÁ theo hai kiểu lẫn lộn:
#   - phần lớn ghi theo NGHÌN ĐỒNG  : Chelate Cu = 108   (thực tế 108.000 đ/kg)
#   - một số ít gõ đủ số bằng ĐỒNG  : GA3 = "4.900.000"  (thực tế 4.900.000 đ/kg)
# Không phân biệt được theo kiểu ô (ô chữ có cả '15,4' lẫn '4.900.000'), nhưng phân biệt
# được theo ĐỘ LỚN: rà tay toàn bộ 62 dòng có giá thô >= 1.000 thì đều là giá ĐỒNG có thật
# (CaO 5.999 đ/kg, hộp giấy 2.241 đ/cái, tinh dầu đàn hương 22.000.000 đ/kg...), còn 1.294
# dòng dưới 1.000 nhân lên đều ra giá hợp lý (NAA 320 -> 320.000, MKP 32 -> 32.000).
# Không có dòng nào rơi vào vùng nhập nhằng nên ngưỡng này tách sạch.
PRICE_K_THRESHOLD = 1000


def read_sheet2(wb):
    """Nguyên liệu — không VAT, ngày hoán đổi ngày<->tháng, giá quy về ĐỒNG (xem trên)."""
    ws = wb["Sheet2"]
    dmap, dfix = build_date_map(ws, 6, swap=True)
    out, st, bad_qty = [], Counter(), []
    for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not r or all(c is None or not str(c).strip() for c in r[:7]):
            continue
        code, sup = nfc(r[1]), nfc(r[0])
        if not code or not sup:
            continue
        price = parse_money(r[4])
        if price <= 0:
            st["khong_gia"] += 1
            continue
        raw_price = price
        if price < PRICE_K_THRESHOLD:
            price = round(price * 1000, 2)
            st["quy_doi_nghin"] += 1
        qty = parse_money(r[3])
        if qty >= QTY_MAX:
            bad_qty.append(("Sheet2", i, sup, code, clean_text(r[2]), qty, price))
            continue
        out.append({
            "sheet": "Sheet2", "row": i, "date": dmap.get(i, ""), "date_fix": dfix.get(i),
            "sup": sup, "code": code, "name": clean_text(r[2]), "unit": clean_text(r[6]),
            "qty": qty, "price": price, "vat": 0, "amount": round(qty * price, 2),
            "group": "", "moq": 0, "raw_price": raw_price,
        })
    return out, st, bad_qty


def build_supplier_matcher(db):
    sups = db.query(Supplier).all()
    exact = {}
    for s in sups:
        exact.setdefault(sup_key(s.name), s)
        exact.setdefault(sup_key(s.code), s)
    name_keys = [(sup_key(s.name), s) for s in sups if sup_key(s.name)]

    def match(name):
        k = sup_key(name)
        if not k:
            return None, ""
        if k in exact:
            return exact[k], "exact"
        for nk, s in name_keys:
            if len(k) >= 5 and (k in nk or nk in k):
                return s, "fuzzy"
        return None, ""

    return match


def is_typo_of(a, b):
    """`a` có phải là `b` bị gõ sai đúng MỘT ký tự không (thừa / thiếu / gõ nhầm)?"""
    a, b = a.upper(), b.upper()
    if a == b:
        return True
    if abs(len(a) - len(b)) > 1:
        return False
    if len(a) == len(b):                         # gõ nhầm 1 ký tự
        return sum(x != y for x, y in zip(a, b)) == 1
    lo, hi = (a, b) if len(a) < len(b) else (b, a)   # thừa/thiếu 1 ký tự
    return any(hi[:i] + hi[i + 1:] == lo for i in range(len(hi)))


def guess_from_prefix(code, prods, field):
    """Suy `item_group` / `unit` cho mã hàng mới theo các mã cùng tiền tố.

    Chỉ nhận khi tiền tố có ĐỦ NHIỀU mã anh em (>= 5) và chúng gần như THỐNG NHẤT (>= 90%)
    — vd NLT* thì 297/297 mã đều thuộc nhóm 'NL'. Tiền tố mơ hồ (BTP* chia đôi ICARE/NL)
    trả rỗng để người dùng tự phân loại, hơn là đoán bừa rồi xếp nhầm nhóm.
    """
    for n in range(min(6, len(code)), 2, -1):
        sib = [p for p in prods if p.code.startswith(code[:n]) and p.code != code]
        if len(sib) < 5:
            continue
        c = Counter(getattr(p, field) for p in sib if getattr(p, field))
        if c:
            val, k = c.most_common(1)[0]
            if k / len(sib) >= 0.9:
                return val
    return ""


def write_csv(name, header, rows):
    if not REPORT_DIR:
        return
    with open(os.path.join(REPORT_DIR, name), "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)


def main():
    db = SessionLocal()
    try:
        wb = openpyxl.load_workbook(XLSX, data_only=True)
        rows1, st1, bad1 = read_sheet1(wb)
        rows2, st2, bad2 = read_sheet2(wb)
        rows = rows1 + rows2
        bad_qty = bad1 + bad2
        print(f"Đọc file: {XLSX}")
        print(f"  Sheet1 (VTBB)       : {len(rows1)} dòng  (bỏ {st1['kho_noi_bo']} dòng kho "
              f"nội bộ, {st1['khong_gia']} dòng không có giá)")
        print(f"  Sheet2 (Nguyên liệu): {len(rows2)} dòng  (bỏ {st2['khong_gia']} dòng không "
              f"có giá) — quy đổi nghìn đồng -> đồng (x1000): {st2['quy_doi_nghin']} dòng, "
              f"giữ nguyên {len(rows2) - st2['quy_doi_nghin']} dòng đã ghi bằng đồng")
        print(f"  Bỏ do SL >= {QTY_MAX:,}: {len(bad_qty)} dòng (xem ls_sl_bat_thuong.csv)")
        print(f"  TỔNG                : {len(rows)} dòng\n")

        # ---- Ngày ----
        fixed_rows = [r for r in rows if r.get("date_fix")]
        no_date = sum(1 for r in rows if not r["date"])
        print(f"Ngày: sửa {len(fixed_rows)} dòng gõ sai (xem ls_ngay_sua.csv) | "
              f"còn trống: {no_date} dòng")

        # ---- Pháp nhân ----
        co = db.query(Company).filter(Company.id == COMPANY_ID).first()
        company_name = co.name if co else ""
        print(f"Pháp nhân gán cho toàn bộ dòng cũ: [{COMPANY_ID}] {company_name}\n")

        # ---- Khớp NCC (không khớp -> tạo mới, giống import_survey_history) ----
        match = build_supplier_matcher(db)
        distinct = {}
        for r in rows:
            distinct.setdefault(sup_key(r["sup"]), r["sup"])
        sup_code_of, how_of = {}, Counter()
        to_create = []
        for k, disp in sorted(distinct.items()):
            s, how = match(disp)
            if s:
                sup_code_of[k] = s.code
                how_of[how] += 1
            else:
                to_create.append((k, disp))
        print(f"NCC trong file: {len(distinct)} | khớp chính xác: {how_of['exact']} | "
              f"khớp gần đúng: {how_of['fuzzy']} | tạo mới: {len(to_create)}")

        # ---- Đối chiếu mã sản phẩm ----
        # Thứ tự: khớp MÃ -> khớp TÊN (chỉ khi mã LỆCH 1 KÝ TỰ) -> tạo mã mới.
        # Bước khớp tên là để cứu dòng gõ sai mã: NTLT0309 gõ nhầm của NLT0309 — tên hàng
        # trùng khít VÀ mã chỉ thừa một chữ 'T'. KHÔNG được khớp tên suông: danh mục có
        # nhiều mã TRÙNG TÊN nhau (NLT0330 và NLT0323 cùng tên) nên khớp tên suông sẽ gán
        # lịch sử của mã này sang mã khác — đó là hai mặt hàng khác nhau, phải tạo mã mới.
        prods = db.query(Product).all()
        by_code = {p.code: p for p in prods}
        by_name = {}
        for p in prods:
            by_name.setdefault(name_key(p.name), []).append(p)
        code_fix, miss_info = {}, {}
        for r in rows:
            if r["code"] in by_code:
                continue
            same = [p for p in by_name.get(name_key(r["name"]), [])
                    if is_typo_of(r["code"], p.code)]
            if len(same) == 1:
                code_fix[r["code"]] = same[0].code
                continue
            mi = miss_info.setdefault(r["code"], {"name": Counter(), "unit": Counter(),
                                                  "group": Counter(), "n": 0})
            mi["n"] += 1
            if r["name"]:
                mi["name"][r["name"]] += 1
            if r["unit"]:
                mi["unit"][r["unit"]] += 1
            if r["group"]:
                mi["group"][r["group"]] += 1
        print(f"Mã hàng trong file: {len({r['code'] for r in rows})} | sửa mã theo tên: "
              f"{len(code_fix)} | tạo mã mới: {len(miss_info)}\n")

        # ---- Báo cáo ra CSV để rà ----
        write_csv("ls_sl_bat_thuong.csv",
                  ["sheet", "dong_excel", "ncc", "ma_hang", "ten_hang", "so_luong", "don_gia"],
                  bad_qty)
        write_csv("ls_ngay_sua.csv",
                  ["sheet", "dong_excel", "o_goc", "ngay_sau_sua", "cach_suy", "ma_hang"],
                  [(r["sheet"], r["row"], r["date_fix"][0], r["date_fix"][1],
                    r["date_fix"][2], r["code"]) for r in fixed_rows])
        write_csv("ls_ncc_tao_moi.csv", ["ten_ncc_moi", "so_dong"],
                  [(d, sum(1 for r in rows if sup_key(r["sup"]) == k)) for k, d in to_create])
        write_csv("ls_ma_sua_theo_ten.csv", ["ma_trong_file", "ma_trong_danh_muc"],
                  sorted(code_fix.items()))
        write_csv("ls_ma_tao_moi.csv", ["ma", "ten", "dvt", "phan_loai", "so_dong"],
                  [(c, mi["name"].most_common(1)[0][0] if mi["name"] else c,
                    mi["unit"].most_common(1)[0][0] if mi["unit"] else "",
                    mi["group"].most_common(1)[0][0] if mi["group"] else "", mi["n"])
                   for c, mi in sorted(miss_info.items())])
        print("Đã ghi review: ls_sl_bat_thuong.csv, ls_ngay_sua.csv, ls_ncc_tao_moi.csv, "
              "ls_ma_sua_theo_ten.csv, ls_ma_tao_moi.csv")

        if DRY_RUN:
            print("\n(DRY_RUN) Không ghi DB. Bỏ DRY_RUN=1 để nạp thật.")
            return

        # ---- Tạo NCC còn thiếu ----
        for k, disp in to_create:
            code = generate_code(db, Supplier, NEW_SUPPLIER_PREFIX)
            db.add(Supplier(code=code, name=disp, tax_code="", supplier_type="goods",
                            created_by=0, updated_by=0))
            db.flush()
            sup_code_of[k] = code
        db.commit()
        print(f"\nĐã tạo {len(to_create)} NCC mới (MST để trống, chờ bổ sung).")

        # ---- Tạo mã hàng còn thiếu ----
        for c, mi in sorted(miss_info.items()):
            db.add(Product(
                code=c, name=mi["name"].most_common(1)[0][0] if mi["name"] else c,
                item_group=(mi["group"].most_common(1)[0][0] if mi["group"]
                            else guess_from_prefix(c, prods, "item_group")),
                unit=(mi["unit"].most_common(1)[0][0] if mi["unit"]
                      else guess_from_prefix(c, prods, "unit")),
                invoice_name="", legal_name="", hh_code="", hh_name="",
                is_active=True, created_by=0, updated_by=0))
        db.commit()
        print(f"Đã tạo {len(miss_info)} mã hàng mới.")

        # ---- Nạp ----
        old = db.execute(text("SELECT COUNT(*) FROM tab_purchase_history "
                              "WHERE source='legacy'")).scalar()
        db.execute(text("DELETE FROM tab_purchase_history WHERE source='legacy'"))
        db.commit()
        print(f"Đã xóa {old} dòng lịch sử cũ (source='legacy') trước khi nạp lại.")

        n = 0
        for r in rows:
            extra = {
                "nguon": "khaosatsanpham.xlsx", "sheet": r["sheet"], "dong_excel": r["row"],
                "item_group": r["group"], "moq": r["moq"],
                # Giá gốc trong file (Sheet2 ghi bằng nghìn đồng) — giữ lại để truy ngược
                "gia_goc_trong_file": r.get("raw_price"),
            }
            if r["code"] in code_fix:
                extra["ma_trong_file"] = r["code"]
            if r.get("date_fix"):
                extra["ngay_goc_trong_file"] = r["date_fix"][0]
                extra["cach_suy_ngay"] = r["date_fix"][2]
            db.add(PurchaseHistory(
                po_item_id=None, po_id=0, po_code="", source="legacy",
                legacy_key=f"KSLS:{r['sheet']}:{r['row']}",
                product_code=code_fix.get(r["code"], r["code"]), product_name=r["name"],
                supplier_code=sup_code_of.get(sup_key(r["sup"]), ""), supplier_name=r["sup"],
                company_id=COMPANY_ID, company_name=company_name,
                order_date=r["date"], unit=r["unit"], qty_order=r["qty"], price=r["price"],
                vat=r["vat"], amount=r["amount"], completed_at=r["date"],
                extra=json.dumps(extra, ensure_ascii=False),
            ))
            n += 1
            if n % 500 == 0:
                db.commit()
        db.commit()
        tot = db.execute(text("SELECT COUNT(*) FROM tab_purchase_history")).scalar()
        leg = db.execute(text("SELECT COUNT(*) FROM tab_purchase_history "
                              "WHERE source='legacy'")).scalar()
        print(f"\nĐã nạp {n} dòng lịch sử cũ. Bảng hiện có: {tot} dòng "
              f"(cũ {leg} / hệ thống {tot - leg}).")
    finally:
        db.close()


if __name__ == "__main__":
    main()
