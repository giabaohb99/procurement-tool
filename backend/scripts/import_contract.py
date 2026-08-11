"""Đồng bộ Hợp đồng NCC từ file Excel (hợp đồng.xlsx) vào tab_contract.

Nguồn: /app/_hopdong_tmp.xlsx (bind-mount local) hoặc đường dẫn qua biến môi trường HD_XLSX.
Chỉ nạp dòng 'Hợp đồng = Có', KHỚP được NCC (exact/fuzzy) — dòng không khớp thì bỏ + liệt kê.

Ánh xạ cột:
  Tên NCC (col1)      -> khớp Supplier (party_code/party_name)
  Công ty ký (col4)   -> company_id (Ida->IDA, Dego->DEGO, Aba->ABA, Icare->ICARE, BAMBOO->BAMBOO,
                          Dr Xanh->NPP DR.XANH, Dr xanh 2->HỘ KD DR.XANH,
                          SX&NK...Nông nghiệp DEGO->NN DEGO, TRỐNG/không rõ->DEGO)
  Tên hàng hóa (col5) -> note (ghi chú)
  Ngày ký (col6)      -> start_date (YYYY-MM-DD)
  Dạng hợp đồng (col7)-> contract_type
  signed=True, status='Hiệu lực', end_date='' (Excel không có)

Đồng bộ sạch: xoá các HĐ mã 'HDX%' (do script tạo) trước rồi nạp lại — re-run an toàn,
KHÔNG đụng HĐ nhập tay (mã HD#####).

Chạy: docker exec -w /app <api> python -m scripts.import_contract
"""
import os
import re
import unicodedata
from datetime import datetime
from difflib import SequenceMatcher

from openpyxl import load_workbook

import app.core.all_models  # noqa: F401  (đăng ký mapper)
from app.core.database import SessionLocal
from app.modules.contract.model import Contract
from app.modules.supplier.model import Supplier
from app.modules.company.model import Company
from app.modules.user.model import User

FUZZY_THRESHOLD = 0.9
CODE_PREFIX = "HDX"


def _norm(s) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.replace("đ", "d")
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _xlsx_path() -> str:
    p = os.environ.get("HD_XLSX", "/app/_hopdong_tmp.xlsx")
    if not os.path.exists(p):
        alt = os.path.join(os.path.dirname(__file__), "..", "_hopdong_tmp.xlsx")
        if os.path.exists(alt):
            return alt
    return p


def _parse_date(v) -> str:
    """Trả về YYYY-MM-DD hoặc '' nếu không parse được."""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return ""


def _match_supplier(name, sups, by_code, by_name):
    """Khớp NCC: (1) exact tên/mã; (2) TẤT CẢ từ trong MÃ NCC xuất hiện trong tên Excel
    (tín hiệu mạnh, chính xác — vd mã 'Thu Loan' nằm trong 'CÔNG TY ... THU LOAN');
    (3) chứa nhau / difflib theo tên. Trả (supplier, kind) — kind in EXACT/CODE/FUZZY."""
    n = _norm(name)
    if not n:
        return None, None
    if n in by_code:
        return by_code[n], "EXACT"
    if n in by_name:
        return by_name[n], "EXACT"
    ntok = n.split()
    # (2) mã NCC xuất hiện LIÊN TIẾP (cụm từ) trong tên Excel — tránh khớp nhầm qua
    # các từ khuôn "công ty / thành viên..." (vd 'Thành Công' KHÔNG khớp 'MỘT THÀNH VIÊN').
    code_cands = []
    for s in sups:
        ct = _norm(s.code).split()
        L = len(ct)
        if not L:
            continue
        if any(ntok[i:i + L] == ct for i in range(len(ntok) - L + 1)):
            code_cands.append((s, L))
    if code_cands:
        code_cands.sort(key=lambda x: -x[1])
        top = code_cands[0][1]
        best = [s for s, l in code_cands if l == top]
        if len(best) == 1:
            return best[0], "CODE"
    # (3) chứa nhau theo tên
    contain = list({s.id: s for k, s in by_name.items() if k and (k in n or n in k)}.values())
    if len(contain) == 1:
        return contain[0], "FUZZY"
    # (4) difflib theo tên
    best_s, best_score = None, 0.0
    for k, s in by_name.items():
        r = SequenceMatcher(None, n, k).ratio()
        if r > best_score:
            best_s, best_score = s, r
    if best_score >= 0.93:
        return best_s, "FUZZY"
    return None, None


def _match_company(val, coms_by_code, dego, npp_drxanh, hokd_drxanh, nn_dego):
    n = _norm(val)
    if not n:
        return dego
    if "dr xanh" in n or "drxanh" in n:
        return hokd_drxanh if "2" in n else npp_drxanh
    if "nong nghiep dego" in n:  # SX&NK Hóa chất Nông nghiệp DEGO
        return nn_dego
    if n in coms_by_code:
        return coms_by_code[n]
    return dego  # không rõ -> DEGO


def run():
    path = _xlsx_path()
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))[1:]

    db = SessionLocal()
    try:
        sys_user = db.query(User).order_by(User.id).first()
        uid = sys_user.id if sys_user else 1

        sups = db.query(Supplier).all()
        by_code = {_norm(s.code): s for s in sups}
        by_name = {_norm(s.name): s for s in sups}

        coms = db.query(Company).all()
        coms_by_code = {_norm(c.code): c.id for c in coms}

        def _cid(code):
            c = next((x for x in coms if x.code == code), None)
            return c.id if c else 0
        dego, npp, hokd, nndego = _cid("DEGO"), _cid("NPP DR.XANH"), _cid("HỘ KD DR.XANH"), _cid("NN DEGO")

        # forward-fill NCC theo nhóm STT
        cur = None
        recs = []
        for r in rows:
            if len(r) > 1 and r[1] and str(r[1]).strip():
                cur = str(r[1]).strip()
            recs.append((cur, r))

        # Xoá HĐ import cũ (chỉ HDX%)
        old = db.query(Contract).filter(Contract.code.like(CODE_PREFIX + "%")).all()
        for c in old:
            db.delete(c)
        db.flush()
        print(f"Đã xoá {len(old)} HĐ import cũ (mã {CODE_PREFIX}...)")

        created = 0
        review = {"CODE": {}, "FUZZY": {}}   # kind -> {ncc: (supplier_code, supplier_name)}
        skip_ncc = {}      # tên NCC không khớp -> số dòng
        bad_date = 0
        by_type = {}
        seq = 0
        for cur, r in recs:
            if len(r) <= 7:
                continue
            if _norm(r[2]) not in ("co", "c"):   # chỉ 'Có'
                continue
            sup, kind = _match_supplier(cur, sups, by_code, by_name)
            if not sup:
                skip_ncc[cur] = skip_ncc.get(cur, 0) + 1
                continue
            company = _match_company(r[4], coms_by_code, dego, npp, hokd, nndego)
            ctype = str(r[7]).strip() if r[7] else ""
            sdate = _parse_date(r[6])
            if r[6] and not sdate:
                bad_date += 1
            goods = str(r[5]).strip() if r[5] else ""

            seq += 1
            c = Contract(
                code=f"{CODE_PREFIX}{seq:04d}",
                party_type="Nhà cung cấp",
                party_code=sup.code,
                party_name=sup.name,
                company_id=company,
                title=f"{sup.code} - {ctype}".strip(" -"),
                contract_type=ctype,
                start_date=sdate,
                end_date="",
                signed=True,
                status="Hiệu lực",
                note=goods,
                created_by=uid,
                updated_by=uid,
            )
            db.add(c)
            created += 1
            by_type[ctype] = by_type.get(ctype, 0) + 1
            if kind in review:
                review[kind][cur] = (sup.code, sup.name)

        db.commit()
        print("-" * 70)
        print(f"Đã tạo {created} hợp đồng (mã {CODE_PREFIX}0001..).")
        print("Theo dạng HĐ:")
        for k, v in sorted(by_type.items(), key=lambda x: -x[1]):
            print(f"  [{v:3d}] {k}")
        if bad_date:
            print(f"CẢNH BÁO: {bad_date} dòng có ngày ký không parse được -> để trống start_date.")
        for kind, label in (("CODE", "KHỚP THEO MÃ NCC"), ("FUZZY", "KHỚP MỜ (difflib)")):
            d = review[kind]
            if d:
                print(f"\n{label} ({len(d)} NCC) — kiểm tra lại đúng NCC chưa:")
                for name, (code, sname) in d.items():
                    print(f"  - {name!r}  ->  [{code}] {sname}")
        if skip_ncc:
            total_skip = sum(skip_ncc.values())
            print(f"\nBỎ QUA {total_skip} dòng 'Có' do NCC chưa có trong hệ thống ({len(skip_ncc)} NCC):")
            for name, cnt in sorted(skip_ncc.items(), key=lambda x: -x[1]):
                print(f"  [{cnt}] {name}")
    finally:
        db.close()


if __name__ == "__main__":
    run()
