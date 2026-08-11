"""import_survey_history — Đổ lịch sử vật tư mua thực tế (khaosatsanpham.xlsx) thành
KHẢO SÁT GIÁ ĐÃ DUYỆT + PHƯƠNG ÁN cho người dùng chọn.

MÔ HÌNH (chốt với user v4 — CHỈ PHIẾU KHẢO SÁT, BỎ YÊU CẦU BÁO GIÁ):
  1 PHIẾU KHẢO SÁT (Survey) = 1 (PHÂN LOẠI + MÃ VTBB)  -> số phiếu nhiều cũng OK
    └ Survey (approved): has_product_code=True, item_code=mã (điền sẵn ô header),
      ĐỨNG ĐỘC LẬP: survey_request_id=0, sr_code="" (KHÔNG gắn Yêu cầu báo giá).
    └ N DÒNG SP (SurveyProductLine) = MỖI LẦN MUA mã đó = 1 dòng (KHÔNG gộp),
      line_approve="Đã duyệt", hiện thẳng NCC + giá + MOQ + ngày trong phiếu.
  Cùng NCC nhưng khác MOQ/giá/ngày -> giữ tách dòng (user: "bỏ vào chung hết").
  Ngày khảo sát ở màn tìm kết quả lấy từ SurveyProductLine.result_date (theo DÒNG).
  (v3 cũ có tạo SurveyRequest/YCBG + option -> ĐÃ BỎ theo yêu cầu user.)

TỰ MAPPING:
  - Mã VTBB thiếu trong tab_product  -> TẠO mới (phân loại + tên + ĐVT có sẵn).
  - NCC thiếu trong tab_supplier      -> TẠO mới (MST rỗng), trừ kho nội bộ trong SKIP_SUPPLIERS.

IDEMPOTENT: chạy lại sẽ XÓA sạch dữ liệu seed cũ (theo marker) rồi tạo lại.
  - Survey.import_key LIKE 'KSSEED:%'
  - SurveyProductLine.import_line_key LIKE 'KSSEED:%'
  - SurveyRequest.note == SEED_TAG  (dọn nốt YCBG do v3 cũ tạo)
  (KHÔNG xóa NCC/Product đã tạo — chỉ tạo nếu thiếu.)

Chạy trong container:
  DRY_RUN=1 docker exec -w /app <api> python -m scripts.import_survey_history   # chỉ in số liệu
  docker exec -w /app <api> python -m scripts.import_survey_history             # chạy thật

ENV: XLSX (mặc định /app/_khaosat_tmp.xlsx), SHEET (Sheet1), COMPANY_ID (1=DEGO),
     LINES_PER_PHIEU (200), DRY_RUN (0/1), REPORT_DIR (/app).
"""
import os
import re
import unicodedata
from collections import defaultdict, Counter
from datetime import datetime

import openpyxl
from sqlalchemy import text
from app.core import all_models  # noqa: F401  (đăng ký mọi mapper)
from app.core.database import SessionLocal
from app.core.utils import generate_code
from app.modules.supplier.model import Supplier
from app.modules.product.model import Product
from app.modules.survey.model import Survey, SurveyProductLine
from app.modules.survey_request.model import (SurveyRequest, SurveyRequestLine,
                                              SurveyRequestOption, SurveyRequestPr)

XLSX = os.getenv("XLSX", "/app/_khaosat_tmp.xlsx")
SHEET = os.getenv("SHEET", "Sheet1")
COMPANY_ID = int(os.getenv("COMPANY_ID", "1"))
LINES_PER_PHIEU = int(os.getenv("LINES_PER_PHIEU", "200"))
DRY_RUN = os.getenv("DRY_RUN", "0") == "1"
REPORT_DIR = os.getenv("REPORT_DIR", "/app")
SEED_TAG = "SEED-KHAOSAT-LICHSU"

# Kho/VP nội bộ — KHÔNG phải NCC, bỏ mọi dòng của chúng.
SKIP_SUPPLIERS = {"kho vtbb sx", "vp b19"}

# Cột (0-based) theo header đã xác nhận.
C_DATE, C_SUP, C_GROUP, C_CODE, C_NAME, C_UOM = 1, 2, 3, 4, 5, 6
C_QTY, C_MOQ, C_PRICE, C_VAT, C_PRICE_VAT = 7, 8, 9, 10, 11


_BAD = {"#N/A", "N/A", "NA", "-", "NULL", "#REF!"}


def nfc(s):
    return unicodedata.normalize("NFC", (str(s) if s is not None else "").strip())


def clean_text(s):
    s = nfc(s)
    return "" if s.upper() in _BAD else s


def sup_key(s):
    """Khóa so khớp NCC: bỏ dấu công ty, ký tự đặc biệt, gộp khoảng trắng, lower."""
    s = nfc(s).lower()
    s = s.replace("công ty tnhh", "").replace("công ty", "").replace("cty", "")
    s = re.sub(r"[.,()\-/]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def parse_money(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = nfc(v).replace("đ", "").replace("₫", "").replace(" ", "")
    if not s:
        return 0.0
    if "," in s and "." in s:          # 1.234.567,89 -> . là ngăn nghìn
        s = s.replace(".", "").replace(",", ".")
    elif "." in s:                     # 3.500 -> ngăn nghìn (VND nguyên)
        s = s.replace(".", "")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_date(v):
    if not v:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = nfc(v)
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return ""


def load_rows():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or all(c is None for c in r):
            continue
        code = nfc(r[C_CODE])
        group = nfc(r[C_GROUP])
        sup = nfc(r[C_SUP])
        if not code or not group or not sup:
            continue
        price = parse_money(r[C_PRICE])
        if price <= 0:                     # bỏ dòng giá rỗng ('-' trên sheet)
            continue
        vat_amt = parse_money(r[C_VAT])
        vat_pct = round(vat_amt / price * 100) if price else 8
        if vat_pct <= 0 or vat_pct > 30:
            vat_pct = 8
        rows.append({
            "date": parse_date(r[C_DATE]),
            "sup": sup,
            "group": group,
            "code": code,
            "name": clean_text(r[C_NAME]),
            "uom": clean_text(r[C_UOM]),
            "qty": parse_money(r[C_QTY]),
            "moq": parse_money(r[C_MOQ]),
            "price": price,
            "vat": vat_pct,
        })
    return rows


def main():
    db = SessionLocal()
    rows = load_rows()
    total_rows = len(rows)

    # ---- 1. Map NCC ----
    suppliers = db.query(Supplier).all()
    by_key = {}
    exact_keys = {}
    for s in suppliers:
        by_key.setdefault(sup_key(s.name), s)
        by_key.setdefault(sup_key(s.code), s)
        exact_keys.setdefault(sup_key(s.name), s)
    name_keys = [(sup_key(s.name), s) for s in suppliers if sup_key(s.name)]

    def match_supplier(name):
        k = sup_key(name)
        if not k:
            return None, ""
        if k in exact_keys:
            return exact_keys[k], "exact"
        if k in by_key:
            return by_key[k], "exact-code"
        for nk, s in name_keys:
            if len(k) >= 4 and (k in nk or nk in k):
                return s, "fuzzy"
        return None, ""

    distinct_sups = {}
    for r in rows:
        distinct_sups.setdefault(sup_key(r["sup"]), r["sup"])

    sup_code_of = {}      # sup_key -> code trong hệ thống
    matched_report, fuzzy_report, created_sup, skipped_sup = [], [], [], []
    for k, disp in sorted(distinct_sups.items()):
        if k in SKIP_SUPPLIERS:
            skipped_sup.append(disp)
            continue
        s, how = match_supplier(disp)
        if s:
            sup_code_of[k] = s.code
            if how == "fuzzy":
                fuzzy_report.append((disp, s.code, s.name))
            else:
                matched_report.append((disp, s.code, s.name))
        else:
            created_sup.append(disp)   # tạo sau

    # ---- 2. Map/đếm mã thiếu ----
    prod_by_code = {p.code: p for p in db.query(Product).all()}
    # gom info cho mã thiếu (phân loại + tên + đvt phổ biến nhất)
    missing_info = defaultdict(lambda: {"group": Counter(), "name": Counter(), "uom": Counter()})
    all_codes = set()
    for r in rows:
        all_codes.add(r["code"])
        if r["code"] not in prod_by_code:
            mi = missing_info[r["code"]]
            mi["group"][r["group"]] += 1
            if r["name"]:
                mi["name"][r["name"]] += 1
            if r["uom"]:
                mi["uom"][r["uom"]] += 1
    missing_codes = sorted(missing_info.keys())

    # ---- 3. Gom dòng theo (phân loại, mã) ----
    # line_key -> [MỌI dòng mua] (KHÔNG dedup theo NCC nữa: user muốn "bỏ vào chung hết",
    # cùng NCC cùng mã nhưng khác MOQ/giá/ngày -> mỗi lần mua = 1 dòng SP riêng).
    lines = defaultdict(list)
    line_meta = {}
    for r in rows:
        if sup_key(r["sup"]) in SKIP_SUPPLIERS:
            continue
        lk = (r["group"], r["code"])
        lines[lk].append(r)
        m = line_meta.setdefault(lk, {"name": Counter(), "uom": Counter(), "date": ""})
        if r["name"]:
            m["name"][r["name"]] += 1
        if r["uom"]:
            m["uom"][r["uom"]] += 1
        if (r["date"] or "") >= (m["date"] or ""):
            m["date"] = r["date"]

    total_lines = sum(len(v) for v in lines.values())      # tổng dòng SP sẽ tạo

    # 1 PHIẾU = 1 (phân loại, mã). Mỗi line_key là 1 phiếu.
    phieu_plan = sorted(lines.keys(), key=lambda x: (x[0], x[1]))   # [(group, code), ...]

    # ---- IN KẾ HOẠCH ----
    print("=" * 60)
    print(f"Tong dong Excel hop le : {total_rows}")
    print(f"NCC distinct           : {len(distinct_sups)}  "
          f"(khop {len(matched_report)+len(fuzzy_report)}, tao moi {len(created_sup)}, bo qua noi bo {len(skipped_sup)})")
    print(f"  - khop mo (fuzzy)    : {len(fuzzy_report)}  (xem review file)")
    print(f"Ma VTBB distinct       : {len(all_codes)}  (thieu -> tao moi: {len(missing_codes)})")
    print(f"So PHIEU KHAO SAT (Survey): {len(phieu_plan)}  (1 phieu = 1 phan loai + ma)")
    print(f"So DONG SP (moi lan mua): {total_lines}  (KHONG gop, moi lan mua = 1 dong)")
    print(f"Bo qua noi bo          : {skipped_sup}")
    print("=" * 60)

    # ---- Ghi report review ----
    def w(path, header, rows_):
        with open(os.path.join(REPORT_DIR, path), "w", encoding="utf-8-sig") as f:
            f.write(header + "\n")
            for r in rows_:
                f.write(",".join('"' + str(x).replace('"', "'") + '"' for x in r) + "\n")

    w("ks_ncc_fuzzy.csv", "ten_khaosat,ma_master,ten_master", fuzzy_report)
    w("ks_ncc_tao_moi.csv", "ten_ncc_moi", [(x,) for x in created_sup])
    w("ks_ma_tao_moi.csv", "ma,phan_loai,ten,dvt",
      [(c, missing_info[c]["group"].most_common(1)[0][0],
        (missing_info[c]["name"].most_common(1)[0][0] if missing_info[c]["name"] else ""),
        (missing_info[c]["uom"].most_common(1)[0][0] if missing_info[c]["uom"] else "")) for c in missing_codes])
    print(f"Da ghi review: ks_ncc_fuzzy.csv, ks_ncc_tao_moi.csv, ks_ma_tao_moi.csv (tai {REPORT_DIR})")

    if DRY_RUN:
        print("DRY_RUN=1 -> KHONG ghi DB. Xem so lieu + file review o tren.")
        db.close()
        return

    # ================== GHI DB ==================
    # 0. Xóa seed cũ
    print("Xoa du lieu seed cu (neu co)...")
    old_srs = db.query(SurveyRequest.id).filter(SurveyRequest.note == SEED_TAG).all()
    old_ids = [x[0] for x in old_srs]
    if old_ids:
        db.query(SurveyRequestPr).filter(
            SurveyRequestPr.survey_request_id.in_(old_ids)).delete(synchronize_session=False)
        old_lines = [x[0] for x in db.query(SurveyRequestLine.id)
                     .filter(SurveyRequestLine.survey_request_id.in_(old_ids)).all()]
        if old_lines:
            db.query(SurveyRequestOption).filter(
                SurveyRequestOption.survey_request_line_id.in_(old_lines)).delete(synchronize_session=False)
            db.query(SurveyRequestLine).filter(
                SurveyRequestLine.id.in_(old_lines)).delete(synchronize_session=False)
        db.query(SurveyRequest).filter(SurveyRequest.id.in_(old_ids)).delete(synchronize_session=False)
    db.query(SurveyProductLine).filter(SurveyProductLine.import_line_key.like("KSSEED:%")).delete(synchronize_session=False)
    db.query(Survey).filter(Survey.import_key.like("KSSEED:%")).delete(synchronize_session=False)
    db.commit()

    # Reset AUTO_INCREMENT để ID chạy lại từ 1, TRÙNG KHỚP với mã KS##### (user dễ quan sát).
    # CHỈ reset khi bảng đã sạch (không còn phiếu/dòng thật nào) -> an toàn.
    if db.query(Survey).count() == 0:
        db.execute(text("ALTER TABLE tab_survey AUTO_INCREMENT = 1"))
        print("  reset AUTO_INCREMENT tab_survey -> 1 (ID se trung ma KS#####)")
    if db.query(SurveyProductLine).count() == 0:
        db.execute(text("ALTER TABLE tab_survey_product_line AUTO_INCREMENT = 1"))
    db.commit()

    # 1. Tạo NCC thiếu
    name_to_disp = {sup_key(d): d for d in created_sup}
    for disp in created_sup:
        code = generate_code(db, Supplier, "NCCKS")
        s = Supplier(code=code, name=disp, tax_code="", supplier_type="goods",
                     created_by=0, updated_by=0)
        db.add(s)
        db.flush()
        sup_code_of[sup_key(disp)] = code
    db.commit()
    print(f"Da tao {len(created_sup)} NCC moi (MST rong).")

    # 2. Tạo mã thiếu
    for c in missing_codes:
        mi = missing_info[c]
        p = Product(code=c, name=(mi["name"].most_common(1)[0][0] if mi["name"] else c),
                    item_group=mi["group"].most_common(1)[0][0],
                    unit=(mi["uom"].most_common(1)[0][0] if mi["uom"] else ""),
                    invoice_name="", legal_name="", hh_code="", hh_name="",
                    is_active=True, created_by=0, updated_by=0)
        db.add(p)
    db.commit()
    print(f"Da tao {len(missing_codes)} ma VTBB moi.")

    # 3. Tạo PHIẾU KHẢO SÁT (Survey) + dòng SP (mỗi NCC 1 dòng). KHÔNG tạo YCBG.
    # Sinh mã theo BATCH (đếm 1 lần rồi tăng cục bộ) — tránh quét DB mỗi phiếu.
    sv_seq = _max_seq(db, Survey.code, "KS")

    n_phieu = n_sp = 0
    total = len(phieu_plan)
    for (group, code) in phieu_plan:
        meta = line_meta[(group, code)]
        name = meta["name"].most_common(1)[0][0] if meta["name"] else code
        uom = meta["uom"].most_common(1)[0][0] if meta["uom"] else ""
        ldate = meta["date"] or ""
        sv_seq += 1
        # 1 phiếu = 1 mã -> tick "Đã có mã sản phẩm sẵn", điền mã header.
        # ĐỨNG ĐỘC LẬP: survey_request_id=0, sr_code="" (không gắn Yêu cầu báo giá).
        sv = Survey(
            code=f"KS{sv_seq:05d}", survey_type="combined", status="approved",
            approve_status="Duyệt", survey_request_id=0, sr_code="",
            item_group=group, main_content=f"Khao sat gia lich su - {code}",
            requirement_detail=name,
            received_date=ldate, result_due_date=ldate,
            nspt="", has_product_code=True, item_code=code, item_name=name, uom=uom,
            import_key=f"KSSEED:{code}", created_by=0, updated_by=0)
        db.add(sv)
        db.flush()
        n_phieu += 1

        # dòng SP: MỌI lần mua (mới nhất trước), mỗi lần mua = 1 dòng đã duyệt.
        # Cùng NCC nhưng khác MOQ/giá/ngày -> giữ tách dòng để tra cứu đúng từng lần.
        hist = sorted(lines[(group, code)], key=lambda r: (r["date"] or ""), reverse=True)
        for i, r in enumerate(hist):
            scode = sup_code_of.get(sup_key(r["sup"]), "")
            if not scode:
                continue
            price = r["price"]
            vat = r["vat"]
            qty = r["qty"] or 0
            d = r["date"] or ""
            # SL NHẬP lịch sử (nguồn đã fix dấu . -> , nên số chuẩn). Giá là số VND.
            # Ngày liên hệ = ngày phản hồi = ngày trả KQ = ngày giao lịch sử của lần mua đó.
            psl = SurveyProductLine(
                survey_id=sv.id, contact_date=d, reply_date=d, result_date=d,
                supplier_code=scode, internal_code=code, product_name=r["name"] or code,
                spec="", origin="", quote_unit=r["uom"] or "", moq=r["moq"] or 0,
                price_by_volume=price, volume_range="", vat=vat, request_qty=qty,
                amount=qty * price * (1 + vat / 100.0), delivery_time="",
                line_approve="Đã duyệt", import_line_key=f"KSSEED:{code}:{i}",
                created_by=0, updated_by=0)
            db.add(psl)
            n_sp += 1

        if n_phieu % 200 == 0:
            db.commit()
            print(f"  ... {n_phieu}/{total} phieu")
    db.commit()

    db.close()
    print("=" * 60)
    print(f"XONG: {n_phieu} phieu khao sat, {n_sp} dong SP (NCC).")


def _max_seq(db, col, prefix):
    """Số thứ tự lớn nhất của các mã dạng <prefix><số> đang có (để sinh batch)."""
    mx = 0
    for (c,) in db.query(col).filter(col.like(prefix + "%")).all():
        suf = (c or "")[len(prefix):]
        if suf.isdigit():
            mx = max(mx, int(suf))
    return mx


if __name__ == "__main__":
    main()
