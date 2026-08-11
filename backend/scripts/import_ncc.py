"""Đồng bộ NCC (master data) từ file Excel (ncc.xlsx) vào tab_supplier.

Nguồn: /app/_ncc_tmp.xlsx (bind-mount local) hoặc đường dẫn qua biến môi trường NCC_XLSX.
Khối trái cột 0..6 là dữ liệu NCC (khối phải 7..9 là danh sách ngân hàng phụ -> BỎ QUA).

Ánh xạ cột:
  col0  Tên NCC (pháp lý)      -> name
  col1  Tên viết tắt          -> code   (KHOÁ upsert, chuẩn hoá bỏ dấu)
  col2  Địa chỉ               -> address
  col3  MST                   -> tax_code (BỎ ký tự ' ở đầu)
  col4  Tên TK thụ hưởng      -> bank_account_name  (chỉ set khi dòng CÓ cụm NH)
  col5  Số TK                 -> bank_account
  col6  Ngân hàng/Chi nhánh   -> bank_name

Upsert theo MÃ (col1). Có sẵn -> cập nhật name/address/tax_code (+bank nếu có cụm NH).
Chưa có -> tạo mới (supplier_type='goods', is_active=True).
XOÁ các NCC 'KHO ...' (kho nội bộ, không phải NCC thật): Kho Dr. Xanh, Kho B18,
Kho C1-2, Kho F49 - Icare, Kho Lab Dego.

Chạy: docker exec -w /app <api> python -m scripts.import_ncc
"""
import os
import re
import unicodedata

from openpyxl import load_workbook

import app.core.all_models  # noqa: F401  (đăng ký mapper)
from app.core.database import SessionLocal
from app.modules.supplier.model import Supplier


def _norm(s) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.replace("đ", "d")
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _txt(v) -> str:
    return "" if v is None else str(v).strip()


def _tax(v) -> str:
    """MST: bỏ ký tự ' ở đầu (Excel dùng để ép text)."""
    return _txt(v).lstrip("'").strip()


def _xlsx_path() -> str:
    p = os.environ.get("NCC_XLSX", "/app/_ncc_tmp.xlsx")
    if not os.path.exists(p):
        alt = os.path.join(os.path.dirname(__file__), "..", "_ncc_tmp.xlsx")
        if os.path.exists(alt):
            return alt
    return p


def run():
    path = _xlsx_path()
    wb = load_workbook(path, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    rows = [r for r in ws.iter_rows(values_only=True)][1:]
    rows = [r for r in rows if r and r[0] and str(r[0]).strip()]  # có tên NCC ở col0

    db = SessionLocal()
    try:
        sups = db.query(Supplier).all()
        by_code = {_norm(s.code): s for s in sups}

        created, updated = [], []
        name_changed = []       # (code, cũ, mới)
        bank_set = 0
        no_code = []
        seen_codes = set()

        for r in rows:
            name = _txt(r[0])
            code = _txt(r[1])
            if not code:
                no_code.append(name[:45])
                continue
            key = _norm(code)
            if key in seen_codes:      # trùng mã trong file -> bỏ dòng sau
                continue
            seen_codes.add(key)

            address = _txt(r[2]) if len(r) > 2 else ""
            tax = _tax(r[3]) if len(r) > 3 else ""
            has_bank = len(r) > 4 and r[4] and str(r[4]).strip()
            bacc_name = _txt(r[4]) if len(r) > 4 else ""
            bacc = _txt(r[5]) if len(r) > 5 else ""
            bname = _txt(r[6]) if len(r) > 6 else ""

            s = by_code.get(key)
            if s:
                if _norm(s.name) != _norm(name):
                    name_changed.append((code, s.name, name))
                s.name = name
                s.address = address
                s.tax_code = tax
                if has_bank:
                    s.bank_account_name = bacc_name
                    s.bank_account = bacc
                    s.bank_name = bname
                    bank_set += 1
                updated.append(code)
            else:
                s = Supplier(
                    code=code, name=name, address=address, tax_code=tax,
                    supplier_type="goods", legal_type="", is_active=True,
                    bank_account_name=bacc_name if has_bank else "",
                    bank_account=bacc if has_bank else "",
                    bank_name=bname if has_bank else "",
                )
                db.add(s)
                created.append(code)
                if has_bank:
                    bank_set += 1

        # Xoá các NCC 'KHO ...' (kho nội bộ)
        deleted = []
        for s in sups:
            first = _norm(s.name).split()
            if first and first[0] == "kho":
                deleted.append((s.code, s.name))
                db.delete(s)

        db.commit()

        print("=" * 72)
        print(f"File: {len(rows)} dòng NCC (col0). Bỏ qua thiếu mã: {len(no_code)}")
        print(f"TẠO MỚI : {len(created)}  -> {created}")
        print(f"CẬP NHẬT: {len(updated)}")
        print(f"Set cụm ngân hàng (Tên TK/Số TK/NH): {bank_set} NCC")
        if name_changed:
            print(f"\nĐỔI TÊN ({len(name_changed)}) — kiểm tra lại:")
            for code, old, new in name_changed:
                print(f"  [{code}] {old!r}\n        -> {new!r}")
        if no_code:
            print(f"\nDÒNG THIẾU MÃ ({len(no_code)}): {no_code}")
        print(f"\nĐÃ XOÁ NCC 'KHO' ({len(deleted)}):")
        for code, nm in deleted:
            print(f"  - [{code}] {nm}")
        print("=" * 72)
    finally:
        db.close()


if __name__ == "__main__":
    run()
