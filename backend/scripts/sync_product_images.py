"""Đồng bộ HÌNH sản phẩm (Chai và nắp + Thùng) từ link Google Drive trong Excel
lên hệ thống QUA API — KHÔNG đụng DB / R2 trực tiếp.

Chạy TẠI MÁY (treo máy). Chỉ cần cài:
    pip install requests openpyxl pillow

Cấu hình qua BIẾN MÔI TRƯỜNG (hoặc sửa hằng số DEFAULT_* ở dưới):
    API_BASE   URL hệ thống    vd https://thumua.degoholding.vn   (dev: https://devthumua.degoholding.vn)
    API_USER   tài khoản đăng nhập
    API_PASS   mật khẩu
    SP_XLSX    đường dẫn tới sanpham.xlsx
    (tùy chọn) SHEETS="Chai và nắp,Thùng"   ONLY_CODE=CPT0001   LIMIT=50   DRY_RUN=1

Bố cục sheet ("Chai và nắp" / "Thùng"), header ở dòng 2, dữ liệu từ dòng 3:
    Mã VTBB(1)  Tên(2)  Phân loại(3)  [ô ảnh có HYPERLINK tới Drive](4)

Với mỗi mã:
  1) GET /api/products?search=<mã>   -> khớp mã chính xác, lấy product id + thumbnail_url
     - không thấy mã            -> ghi lỗi 'product-not-found'
     - đã có thumbnail_url       -> BỎ QUA (đã có ảnh, re-run an toàn)
  2) tải ảnh từ Drive
     - ô không có hyperlink      -> ghi lỗi 'no-link'
     - Drive trả HTML (mất quyền)-> ghi lỗi 'drive-denied'
     - lỗi tải khác             -> ghi lỗi 'download-failed'
  3) Pillow nén (<=1600px cạnh dài, JPEG q<=85, ép < 5MB)
  4) POST /api/attachments (entity=product, entity_id=id, files=...)
     - lỗi upload               -> ghi lỗi 'upload-failed'

KẾT QUẢ (ghi cạnh file script):
  - in ra màn hình:  Hoàn thành X / Y  (kèm số bỏ qua / lỗi)
  - image_errors.csv : mã, phan_loai, ly_do, chi_tiet   -> DANH SÁCH mã cần bù ảnh tay
  - image_done.txt   : cache các mã đã xong (tăng tốc re-run, có thể xóa để chạy lại từ đầu)
"""
import csv
import io
import os
import re
import sys
import time

import requests
from openpyxl import load_workbook

try:
    from PIL import Image
except ImportError:
    sys.exit("Thiếu Pillow. Cài: pip install pillow")

# ----------------- Cấu hình (env > default) -----------------
DEFAULT_API_BASE = "https://thumua.degoholding.vn"
DEFAULT_SP_XLSX = "sanpham.xlsx"

API_BASE = os.environ.get("API_BASE", DEFAULT_API_BASE).rstrip("/")
API_USER = os.environ.get("API_USER", "")
API_PASS = os.environ.get("API_PASS", "")
SP_XLSX = os.environ.get("SP_XLSX", DEFAULT_SP_XLSX)
SHEETS = [s.strip() for s in os.environ.get("SHEETS", "Chai và nắp,Thùng").split(",") if s.strip()]
ONLY_CODE = os.environ.get("ONLY_CODE", "").strip()
LIMIT = int(os.environ.get("LIMIT", "0") or 0)         # 0 = không giới hạn
DRY_RUN = os.environ.get("DRY_RUN", "") not in ("", "0", "false", "False")

MAX_SIDE = 1600
MAX_BYTES = 5 * 1024 * 1024
IMG_EXTS = (".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp")

_HERE = os.path.dirname(os.path.abspath(__file__))
ERR_PATH = os.path.join(_HERE, "image_errors.csv")
DONE_PATH = os.path.join(_HERE, "image_done.txt")


def _txt(v) -> str:
    return "" if v is None else str(v).strip()


# ----------------- Google Drive -----------------
def drive_id(url: str):
    if not url:
        return None
    m = re.search(r"/d/([A-Za-z0-9_-]{20,})", url) or re.search(r"[?&]id=([A-Za-z0-9_-]{20,})", url)
    return m.group(1) if m else None


def drive_download(fid: str, sess: requests.Session):
    """Trả (bytes, None) nếu OK; (None, reason) nếu lỗi ('drive-denied'/'download-failed')."""
    try:
        r = sess.get("https://drive.usercontent.google.com/download",
                     params={"id": fid, "export": "download", "confirm": "t"},
                     timeout=90, stream=True)
    except requests.RequestException as e:
        return None, f"download-failed: {e}"
    ctype = (r.headers.get("Content-Type") or "").lower()
    data = r.content
    if r.status_code != 200:
        return None, f"download-failed: HTTP {r.status_code}"
    # Drive trả trang HTML khi file bị hạn chế quyền / cần đăng nhập
    if "text/html" in ctype or data[:15].lstrip().lower().startswith((b"<!doctype", b"<html")):
        return None, "drive-denied"
    if not data:
        return None, "download-failed: rỗng"
    return data, None


# ----------------- Nén ảnh -----------------
def compress(raw: bytes):
    """Trả (bytes_jpeg, None) hoặc (None, reason)."""
    try:
        im = Image.open(io.BytesIO(raw))
        im.load()
    except Exception as e:
        return None, f"download-failed: không đọc được ảnh ({e})"
    if im.mode not in ("RGB", "L"):
        im = im.convert("RGB")
    w, h = im.size
    if max(w, h) > MAX_SIDE:
        ratio = MAX_SIDE / float(max(w, h))
        im = im.resize((max(1, int(w * ratio)), max(1, int(h * ratio))))
    out = b""
    for q in (85, 75, 65, 55, 45):
        buf = io.BytesIO()
        im.save(buf, "JPEG", quality=q, optimize=True)
        out = buf.getvalue()
        if len(out) <= MAX_BYTES:
            return out, None
    return out, None   # best-effort (vẫn thử upload; API sẽ chặn nếu > 5MB)


# ----------------- API -----------------
def api_login(sess: requests.Session) -> str:
    """Đăng nhập bằng API_USER/API_PASS, gán Bearer cho session, trả token."""
    r = requests.post(f"{API_BASE}/api/auth/login",
                      json={"username": API_USER, "password": API_PASS}, timeout=30)
    r.raise_for_status()
    tok = (r.json().get("data") or {}).get("access_token")
    if not tok:
        raise SystemExit("Đăng nhập thất bại: không nhận được access_token")
    sess.headers["Authorization"] = f"Bearer {tok}"
    return tok


def _with_retry(sess, fn):
    """Chạy fn(); nếu 401 (token hết hạn) -> đăng nhập lại rồi thử lại 1 lần."""
    r = fn()
    if r.status_code == 401:
        api_login(sess)
        r = fn()
    return r


def find_product(sess: requests.Session, code: str):
    """Trả (id, has_image) khớp MÃ chính xác; (None, False) nếu không thấy."""
    r = _with_retry(sess, lambda: sess.get(
        f"{API_BASE}/api/products",
        params={"search": code, "page": 1, "page_size": 50}, timeout=30))
    r.raise_for_status()
    items = ((r.json().get("data") or {}).get("items")) or []
    for it in items:
        if _txt(it.get("code")) == code:
            return it.get("id"), bool(_txt(it.get("thumbnail_url")))
    return None, False


def upload_image(sess: requests.Session, pid: int, code: str, jpeg: bytes):
    def _post():
        files = {"files": (f"{code}.jpg", jpeg, "image/jpeg")}
        data = {"entity": "product", "entity_id": str(pid), "doc_type": ""}
        return sess.post(f"{API_BASE}/api/attachments", data=data, files=files, timeout=120)
    r = _with_retry(sess, _post)
    if r.status_code not in (200, 201):
        msg = r.text[:200]
        try:
            msg = (r.json().get("error") or {}).get("message") or msg
        except Exception:
            pass
        raise RuntimeError(f"HTTP {r.status_code}: {msg}")


# ----------------- Đọc Excel -----------------
def collect_rows():
    """Trả list (code, phan_loai, drive_url, has_filename_text)."""
    wb = load_workbook(SP_XLSX, data_only=True)
    rows = []
    for sheet in SHEETS:
        if sheet not in wb.sheetnames:
            print(f"  [!] Bỏ qua: không có sheet {sheet!r}")
            continue
        ws = wb[sheet]
        maxcol = min(ws.max_column, 12)
        for r in range(3, ws.max_row + 1):
            code = _txt(ws.cell(r, 1).value)
            if not code:
                continue
            grp = _txt(ws.cell(r, 3).value)
            url, has_name = None, False
            for c in range(1, maxcol + 1):
                cell = ws.cell(r, c)
                h = cell.hyperlink
                if h and h.target and "google.com" in h.target:
                    url = h.target
                    break
            if not url:
                for c in range(1, maxcol + 1):
                    v = _txt(ws.cell(r, c).value).lower()
                    if v.endswith(IMG_EXTS):
                        has_name = True
                        break
            rows.append((code, grp, url, has_name))
    return rows


def main():
    if not DRY_RUN and not (API_USER and API_PASS):
        sys.exit("Thiếu API_USER / API_PASS. Đặt biến môi trường rồi chạy lại (hoặc DRY_RUN=1 để xem thử).")
    if not os.path.exists(SP_XLSX):
        sys.exit(f"Không thấy file Excel: {SP_XLSX} (đặt SP_XLSX=... )")

    print(f"API_BASE = {API_BASE}")
    print(f"Excel    = {SP_XLSX}  |  sheets = {SHEETS}")
    print(f"DRY_RUN  = {DRY_RUN}  |  LIMIT = {LIMIT or 'không'}  |  ONLY_CODE = {ONLY_CODE or '-'}")

    rows = collect_rows()
    if ONLY_CODE:
        rows = [x for x in rows if x[0] == ONLY_CODE]
    print(f"Tổng dòng có mã: {len(rows)}")

    done = set()
    if os.path.exists(DONE_PATH):
        with open(DONE_PATH, encoding="utf-8") as f:
            done = {ln.strip() for ln in f if ln.strip()}
        print(f"Đã xong trước đó (cache): {len(done)}")

    sess = requests.Session()
    if not DRY_RUN:
        api_login(sess)

    ok = skip = 0
    errors = []   # (code, grp, reason, detail)
    total = len(rows)
    for i, (code, grp, url, has_name) in enumerate(rows, 1):
        if LIMIT and ok >= LIMIT:
            print(f"Đạt LIMIT={LIMIT}, dừng.")
            break
        if code in done:
            skip += 1
            continue
        tag = f"[{i}/{total}] {code}"

        if not url:
            reason = "no-link" if has_name else "no-image"
            errors.append((code, grp, reason, ""))
            print(f"{tag}  -> {reason}")
            continue
        fid = drive_id(url)
        if not fid:
            errors.append((code, grp, "no-link", url))
            print(f"{tag}  -> no-link (không tách được id: {url})")
            continue

        if DRY_RUN:
            print(f"{tag}  -> (dry) drive={fid}")
            continue

        try:
            pid, has_img = find_product(sess, code)
        except Exception as e:
            errors.append((code, grp, "api-error", str(e)[:150]))
            print(f"{tag}  -> api-error {e}")
            continue
        if not pid:
            errors.append((code, grp, "product-not-found", ""))
            print(f"{tag}  -> product-not-found")
            continue
        if has_img:
            skip += 1
            done.add(code)
            print(f"{tag}  -> đã có ảnh, bỏ qua")
            continue

        raw, err = drive_download(fid, sess)
        if err:
            errors.append((code, grp, err.split(":")[0], err))
            print(f"{tag}  -> {err}")
            continue
        jpeg, err = compress(raw)
        if err:
            errors.append((code, grp, err.split(":")[0], err))
            print(f"{tag}  -> {err}")
            continue
        try:
            upload_image(sess, pid, code, jpeg)
        except Exception as e:
            errors.append((code, grp, "upload-failed", str(e)[:150]))
            print(f"{tag}  -> upload-failed {e}")
            continue

        ok += 1
        done.add(code)
        print(f"{tag}  -> OK ({len(jpeg)//1024} KB)")
        # ghi cache liên tục để re-run không mất tiến độ
        with open(DONE_PATH, "a", encoding="utf-8") as f:
            f.write(code + "\n")
        time.sleep(0.15)   # nhẹ tay với Drive/API

    # ---- Ghi file lỗi ----
    if errors:
        with open(ERR_PATH, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(["ma", "phan_loai", "ly_do", "chi_tiet"])
            w.writerows(errors)

    print("=" * 64)
    print(f"HOÀN THÀNH: {ok} / {total}    (bỏ qua {skip}, lỗi/thiếu {len(errors)})")
    if errors:
        from collections import Counter
        for reason, n in Counter(e[2] for e in errors).most_common():
            print(f"  - {reason}: {n}")
        print(f"Danh sách mã cần bù ảnh tay -> {ERR_PATH}")
    print(f"Cache đã xong -> {DONE_PATH}")
    print("=" * 64)


if __name__ == "__main__":
    main()
