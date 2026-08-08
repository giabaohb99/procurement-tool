"""Nạp nội dung Trung tâm Hướng dẫn từ FILE EXCEL biên soạn tay (HELP-CENTER-*.xlsx).

Khác `import_help_content.py` ở NGUỒN: bản kia đọc JSON xuất từ chính hệ thống (mang nội dung
giữa các môi trường), bản này đọc file Excel do người soạn viết tay. Cách ghi vào DB giống nhau:
khớp theo TIÊU ĐỀ, có thì cập nhật, chưa có thì tạo, KHÔNG xóa gì.

Sheet đọc: "Nội dung bài viết" — cột: STT | Phân cấp | Tên bài viết | Mô tả ngắn | Nội dung (HTML).
  - Cây thư mục dựng theo cột "Phân cấp" dạng 1 / 1.1 / 1.1.1 (cha của 1.1.1 là 1.1).
  - Tên bài viết trong file có thụt lề bằng khoảng trắng ideographic (U+3000) → cắt bỏ khi khớp,
    nếu không sẽ tạo trùng bài đã có.
  - sort_order = số cuối của "Phân cấp" (thứ tự trong cùng 1 thư mục).
Sheet "Cây thư mục & liên kết" chỉ để người soạn đối chiếu, script không cần đọc.

File nguồn để sẵn trong image (app/seed_data/) nên dev/prod chạy thẳng, không cần copy tay:
    docker compose exec -T api python -m scripts.import_help_content_xlsx           # chạy thử
    docker compose exec -T api python -m scripts.import_help_content_xlsx --nap     # ghi thật

Dùng bản Excel khác (chưa commit) thì copy vào container rồi trỏ --file:
    docker compose cp <file>.xlsx api:/tmp/hc.xlsx
    docker compose exec -T api python -m scripts.import_help_content_xlsx --file /tmp/hc.xlsx --nap

Mặc định CHẠY THỬ (in ra, không ghi gì); chỉ ghi khi có --nap.
"""
import argparse

from openpyxl import load_workbook

import app.core.all_models  # noqa: F401  (đăng ký mapper)
from app.core.database import SessionLocal
from app.modules.help_center.model import HelpArticle
from app.modules.role.model import Role
from app.modules.user.model import UserRole

SHEET = "Nội dung bài viết"
HEADER_ROW = 4          # dòng tiêu đề cột; dữ liệu bắt đầu từ dòng 5
XLSX_PATH = "/app/app/seed_data/HELP-CENTER-CONG-CU-THU-MUA-v2.xlsx"


def actor_id(db) -> int:
    """Người thực hiện = một tài khoản quản trị bất kỳ (cột created_by/updated_by)."""
    role = db.query(Role).filter(Role.code == "admin").first()
    ur = db.query(UserRole).filter(UserRole.role_id == role.id).first() if role else None
    return ur.user_id if ur else 1


def _clean(v) -> str:
    """Bỏ thụt lề U+3000 + khoảng trắng thừa. Tiêu đề trong file thụt lề để nhìn ra cây."""
    return str(v or "").replace("　", " ").strip()


def read_rows(path: str) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"File không có sheet '{SHEET}'. Đang có: {wb.sheetnames}")
    ws = wb[SHEET]
    rows = []
    for r in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        cells = (list(r) + [None] * 5)[:5]
        level, title, summary, content = _clean(cells[1]), _clean(cells[2]), _clean(cells[3]), cells[4]
        if not level or not title:
            continue                      # dòng trống / ghi chú cuối bảng
        rows.append({
            "level": level, "title": title,
            "summary": summary[:255] or None,          # cột summary chỉ 255 ký tự
            "content": str(content or ""),
            "parent_level": level.rsplit(".", 1)[0] if "." in level else None,
            "sort_order": int(level.rsplit(".", 1)[-1] or 0),
        })
    return rows


def main(path: str, nap: bool) -> None:
    rows = read_rows(path)
    db = SessionLocal()
    try:
        me = actor_id(db)
        print(f"File: {path} — {len(rows)} mục")
        print(f"Chế độ: {'NẠP THẬT' if nap else 'CHẠY THỬ (không ghi gì)'} — người thực hiện: user #{me}")
        print("-" * 78)

        theo_ten = {a.title: a for a in db.query(HelpArticle).all()}
        id_theo_level: dict[str, int] = {}     # 'Phân cấp' -> id thật trong DB
        them = capnhat = 0
        for r in rows:
            cha = id_theo_level.get(r["parent_level"]) if r["parent_level"] else None
            if r["parent_level"] and cha is None:
                print(f"  ! BỎ QUA {r['level']} {r['title']} — không thấy mục cha {r['parent_level']}")
                continue
            obj = theo_ten.get(r["title"])
            if obj:
                capnhat += 1
                print(f"  ~ cập nhật  #{obj.id:<4} {r['level']:<8} {r['title']}")
                if nap:
                    obj.parent_id = cha
                    obj.content = r["content"]
                    obj.summary = r["summary"]
                    obj.sort_order = r["sort_order"]
                    obj.updated_by = me
                    db.flush()
                id_theo_level[r["level"]] = obj.id
            else:
                them += 1
                print(f"  + thêm mới       {r['level']:<8} {r['title']}")
                if nap:
                    obj = HelpArticle(parent_id=cha, title=r["title"], content=r["content"],
                                      summary=r["summary"], sort_order=r["sort_order"],
                                      created_by=me, updated_by=me)
                    db.add(obj)
                    db.flush()
                    id_theo_level[r["level"]] = obj.id
                else:
                    # chạy thử: giữ chỗ bằng id âm để quan hệ cha-con phía dưới không vỡ
                    id_theo_level[r["level"]] = -len(id_theo_level) - 1

        du = sorted(set(theo_ten) - {r["title"] for r in rows})
        if du:
            print(f"\n  Bài viết CÓ trong DB mà file không có ({len(du)}) — GIỮ NGUYÊN, tự quyết:")
            for t in du:
                print(f"    ? {t}")

        if nap:
            db.commit()
        print("-" * 78)
        print(f"Thêm {them}, cập nhật {capnhat}. {'ĐÃ GHI VÀO DB.' if nap else 'Chạy thử — chưa ghi gì.'}")
        print(f"Tổng sau khi chạy: {db.query(HelpArticle).count()} bài viết")
    finally:
        db.close()


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Nạp Help Center từ file Excel")
    ap.add_argument("--file", default=XLSX_PATH, help="Đường dẫn file .xlsx TRONG container")
    ap.add_argument("--nap", action="store_true", help="Ghi vào DB (mặc định chỉ chạy thử)")
    a = ap.parse_args()
    main(a.file, a.nap)
