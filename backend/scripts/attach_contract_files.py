"""Tải file hợp đồng từ Google Drive (cột 'Link HĐ' trong hợp đồng.xlsx) -> R2 -> gắn vào HĐ.

Ghép file <-> hợp đồng bằng cách TÁI TẠO đúng thứ tự khớp NCC như import_contract
(mỗi dòng 'Có' + khớp NCC = 1 HĐ mã HDX{seq:04d} theo đúng thứ tự duyệt).

Idempotent: HĐ đã có đính kèm thì BỎ QUA (chạy lại an toàn).
Biến môi trường: ATTACH_LIMIT=N (chỉ xử lý N HĐ đầu, để test); HD_XLSX (đường dẫn khác).

Chạy: docker exec -w /app <api> python -m scripts.attach_contract_files
"""
import io
import os
import re
import http.cookiejar
import urllib.request

from openpyxl import load_workbook

import app.core.all_models  # noqa: F401
from app.core.database import SessionLocal
from app.core.storage import dated_key, upload_fileobj
from scripts.import_contract import _norm, _match_supplier, _xlsx_path, CODE_PREFIX
from app.modules.attachment.model import FileLink, StoredFile
from app.modules.contract.model import Contract
from app.modules.supplier.model import Supplier
from app.modules.user.model import User

# cột (1-based) trong file
C_NCC, C_HD, C_LINK = 2, 3, 9

_UA = {"User-Agent": "Mozilla/5.0"}
_MAGIC = {b"%PDF-": ("pdf", "application/pdf"),
          b"\xff\xd8\xff": ("jpg", "image/jpeg"),
          b"\x89PNG\r\n\x1a\n": ("png", "image/png")}


def _file_id(url: str):
    m = re.search(r"/file/d/([^/]+)", url) or re.search(r"[?&]id=([^&]+)", url)
    return m.group(1) if m else None


def _detect(data: bytes):
    for magic, (ext, ct) in _MAGIC.items():
        if data[:len(magic)] == magic:
            return ext, ct
    return None, None


def _drive_download(fid: str):
    """Trả (bytes, content_type) hoặc (None, lý do) nếu không tải được (bị hạn chế quyền...)."""
    cj = http.cookiejar.CookieJar()
    op = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cj))
    base = "https://drive.google.com/uc?export=download&id=" + fid
    try:
        data = op.open(urllib.request.Request(base, headers=_UA), timeout=90).read()
    except Exception as e:
        return None, "loi_tai:%s" % type(e).__name__
    if _detect(data)[0]:
        return data, _detect(data)[1]
    # HTML -> trang xác nhận (file lớn) hoặc đăng nhập (bị hạn chế)
    low = data[:4000].lower()
    if b"<html" in low or b"<!doctype" in low:
        token = next((c.value for c in cj if c.name.startswith("download_warning")), None)
        if not token:
            m = re.search(rb"confirm=([0-9A-Za-z_\-]+)", data)
            token = m.group(1).decode() if m else None
        cand = []
        if token:
            cand.append(base + "&confirm=" + token)
        m = re.search(rb'action="(https://[^"]+)"', data)
        if m:
            cand.append(m.group(1).decode().replace("&amp;", "&"))
        for url2 in cand:
            try:
                data2 = op.open(urllib.request.Request(url2, headers=_UA), timeout=120).read()
            except Exception:
                continue
            if _detect(data2)[0]:
                return data2, _detect(data2)[1]
        return None, "bi_han_che_hoac_dang_nhap"
    return None, "dinh_dang_la"


def _plan(path):
    """Tái tạo thứ tự HĐ -> [(code, ncc, link_target, display)]."""
    wb = load_workbook(path)  # giữ hyperlink
    ws = wb.active
    db = SessionLocal()
    try:
        sups = db.query(Supplier).all()
        by_code = {_norm(s.code): s for s in sups}
        by_name = {_norm(s.name): s for s in sups}
    finally:
        db.close()
    cur = None
    seq = 0
    out = []
    for r in range(2, ws.max_row + 1):
        v_ncc = ws.cell(r, C_NCC).value
        if v_ncc and str(v_ncc).strip():
            cur = str(v_ncc).strip()
        if _norm(ws.cell(r, C_HD).value) not in ("co", "c"):
            continue
        sup, _ = _match_supplier(cur, sups, by_code, by_name)
        if not sup:
            continue
        seq += 1
        cell = ws.cell(r, C_LINK)
        target = cell.hyperlink.target if cell.hyperlink else None
        disp = str(cell.value).strip() if cell.value else ""
        out.append(("%s%04d" % (CODE_PREFIX, seq), cur, target, disp))
    return out


def run():
    plan = _plan(_xlsx_path())
    limit = int(os.environ.get("ATTACH_LIMIT", "0") or "0")
    db = SessionLocal()
    try:
        u = db.query(User).order_by(User.id).first()
        uid = u.id if u else 1
        attached = already = 0
        no_link, folders, failed, missing = [], [], [], []
        processed = 0
        for code, ncc, target, disp in plan:
            if not target:
                no_link.append(code)
                continue
            if "/folders/" in target:
                folders.append((code, disp))
                continue
            fid = _file_id(target)
            if not fid:
                failed.append((code, disp, "khong_ro_id"))
                continue
            c = db.query(Contract).filter(Contract.code == code).first()
            if not c:
                missing.append(code)
                continue
            if db.query(FileLink).filter(FileLink.entity == "contract",
                                         FileLink.entity_id == c.id).first():
                already += 1
                continue
            if limit and processed >= limit:
                break
            processed += 1
            data, ct = _drive_download(fid)
            if not data:
                failed.append((code, disp, ct))
                continue
            ext, ctype = _detect(data)
            fname = disp or ("%s.%s" % (code, ext))
            if "." not in fname:
                fname = "%s.%s" % (fname, ext)
            sf = StoredFile(filename=fname, file_key="", url="", content_type=ctype,
                            size=len(data), created_by=uid, updated_by=uid)
            db.add(sf)
            db.flush()
            key = dated_key("attachment", fname, sf.id)
            sf.file_key = key
            sf.url = upload_fileobj(io.BytesIO(data), key, ctype)
            db.add(FileLink(file_id=sf.id, entity="contract", entity_id=c.id,
                            doc_type="", created_by=uid, updated_by=uid))
            db.commit()
            attached += 1
            if attached % 10 == 0:
                print("... da gan %d file" % attached)

        print("=" * 68)
        print("Tong HĐ trong plan:", len(plan))
        print("ĐÃ GẮN file moi   :", attached)
        print("Da co san (bo qua):", already)
        print("HĐ khong co link  :", len(no_link))
        print("Link THU MUC (tay):", len(folders), [d for _, d in folders])
        print("KHONG tai duoc    :", len(failed))
        for code, disp, why in failed[:30]:
            print("   - [%s] %s (%s)" % (code, disp[:36], why))
        if missing:
            print("HĐ khong tim thay :", missing[:20])
        print("=" * 68)
    finally:
        db.close()


if __name__ == "__main__":
    run()
