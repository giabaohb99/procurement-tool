# -*- coding: utf-8 -*-
"""Import dữ liệu thật từ file Excel vào DB (chạy trên máy host khi container đang chạy).

  pip install openpyxl pymysql
  python tools/import_excel.py

Kết nối MySQL qua localhost:3306 (đổi bằng biến môi trường MYSQL_HOST/PORT/USER/PASSWORD/DB nếu cần).
Đọc: data/Copy of 3. THU MUA_MR TIÊN.xlsx  (sheet '1. DATA CHUNG', '2. DATA NL, VTBB').
"""
import os
import sys

import openpyxl
import pymysql

sys.stdout.reconfigure(encoding="utf-8")

HERE = os.path.dirname(os.path.abspath(__file__))
DATA_FILE = os.path.normpath(os.path.join(HERE, "..", "..", "data", "Copy of 3. THU MUA_MR TIÊN.xlsx"))

CONN = dict(
    host=os.environ.get("MYSQL_HOST", "127.0.0.1"),
    port=int(os.environ.get("MYSQL_PORT", "3306")),
    user=os.environ.get("MYSQL_USER", "app"),
    password=os.environ.get("MYSQL_PASSWORD", "app_password"),
    database=os.environ.get("MYSQL_DB", "procurement"),
    charset="utf8mb4",
)


def s(v) -> str:
    if v is None:
        return ""
    t = str(v).strip()
    if t in ("#VALUE!", "None"):
        return ""
    if t.endswith(".0") and t[:-2].isdigit():
        t = t[:-2]
    return t


def fnum(v, d=0.0) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return d


def main():
    if not os.path.exists(DATA_FILE):
        raise SystemExit(f"Không tìm thấy file: {DATA_FILE}")
    print("Đọc Excel...")
    wb = openpyxl.load_workbook(DATA_FILE, data_only=True)
    dc = wb["1. DATA CHUNG"]
    dn = wb["2. DATA NL, VTBB"]

    companies, suppliers, products = {}, {}, {}

    # DATA CHUNG: cột D/E/F/G = công ty; T/U/V/W/AK/AI = NCC; AC = đơn vị vận chuyển
    for r in dc.iter_rows(min_row=4):
        cells = {c.column_letter: c.value for c in r}
        # company
        code = s(cells.get("D"))
        if code and code not in companies:
            companies[code] = (code, s(cells.get("E")) or code, s(cells.get("G")), s(cells.get("F")))
        # supplier (goods)
        scode = s(cells.get("U"))
        if scode and scode not in suppliers:
            suppliers[scode] = (scode, s(cells.get("T")) or scode, s(cells.get("W")), s(cells.get("V")),
                                "goods", s(cells.get("AK")), fnum(cells.get("AI"), 0.08))
        # carrier (transport)
        ccode = s(cells.get("AC"))
        if ccode and ccode not in suppliers:
            suppliers[ccode] = (ccode, ccode, "", "", "transport", "", 0.08)

    # DATA NL, VTBB: B=phân loại, C=mã, D=tên, E=tên hóa đơn
    for r in dn.iter_rows(min_row=3):
        cells = {c.column_letter: c.value for c in r}
        pcode = s(cells.get("C"))
        name = s(cells.get("D"))
        if pcode and name and pcode not in products:
            products[pcode] = (pcode, name, s(cells.get("E")), s(cells.get("B")), "")

    print(f"Công ty: {len(companies)} · NCC: {len(suppliers)} · Sản phẩm: {len(products)}")

    conn = pymysql.connect(**CONN)
    cur = conn.cursor()

    cur.executemany(
        "INSERT INTO tab_company (code,name,tax_code,address,invoice_email,parent,is_active,created_by,updated_by) "
        "VALUES (%s,%s,%s,%s,'',0,1,0,0) ON DUPLICATE KEY UPDATE name=VALUES(name), tax_code=VALUES(tax_code), address=VALUES(address)",
        list(companies.values()),
    )
    cur.executemany(
        "INSERT INTO tab_supplier (code,name,tax_code,address,supplier_type,payment_terms,vat,is_active,created_by,updated_by) "
        "VALUES (%s,%s,%s,%s,%s,%s,%s,1,0,0) ON DUPLICATE KEY UPDATE name=VALUES(name), tax_code=VALUES(tax_code), "
        "address=VALUES(address), supplier_type=VALUES(supplier_type), payment_terms=VALUES(payment_terms), vat=VALUES(vat)",
        list(suppliers.values()),
    )
    # sản phẩm nhiều -> chia lô
    pvals = list(products.values())
    for i in range(0, len(pvals), 1000):
        cur.executemany(
            "INSERT INTO tab_product (code,name,invoice_name,item_group,unit,is_active,created_by,updated_by) "
            "VALUES (%s,%s,%s,%s,%s,1,0,0) ON DUPLICATE KEY UPDATE name=VALUES(name), "
            "invoice_name=VALUES(invoice_name), item_group=VALUES(item_group)",
            pvals[i:i + 1000],
        )
        print(f"  ...sản phẩm {min(i+1000, len(pvals))}/{len(pvals)}")

    conn.commit()
    cur.close()
    conn.close()
    print("Import xong!")


if __name__ == "__main__":
    main()
