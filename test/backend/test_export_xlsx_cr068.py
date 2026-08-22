"""CR-068 — test xuất Excel các màn danh sách (YCMH · YCBG · ĐMH · Tiến độ mua hàng).

Kiểm 3 nhóm:
- Helper dùng chung `app/core/export_xlsx.py`: chọn cột theo bảng, trần dòng, đổi kiểu ô, dựng file.
- `build_rows` từng màn: bung dòng hàng, cụm đầu phiếu lặp lại, tiền/ngày tính giống màn danh sách.
- Che dữ liệu: YCBG bỏ cột lộ NCC + lọc dòng theo NSTM; Tiến độ bỏ cột NCC/vận chuyển.

Không đụng DB thật — dùng fixture SQLite in-memory ở conftest.
"""
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace

import openpyxl
import pytest
from fastapi import HTTPException

from app.core.export_xlsx import (
    MAX_ROWS, Col, cell_value, check_row_limit, parse_ids, pick_columns, xlsx_response,
)


# ── Helper dùng chung ───────────────────────────────────────────────────────────
def test_parse_ids_bo_qua_rac():
    assert parse_ids("12, 15,x, ,7") == [12, 15, 7]
    assert parse_ids("") == []
    assert parse_ids(None) == []


SPEC = [Col("a", "A"), Col("b", "B"), Col("c", "C")]


def test_pick_columns_giu_dung_thu_tu_nguoi_dung_thay():
    assert [c.key for c in pick_columns(SPEC, "c,a")] == ["c", "a"]


def test_pick_columns_bo_key_la_va_rong_thi_lay_tron_bo():
    assert [c.key for c in pick_columns(SPEC, "b,zzz")] == ["b"]
    assert [c.key for c in pick_columns(SPEC, "")] == ["a", "b", "c"]
    assert [c.key for c in pick_columns(SPEC, None)] == ["a", "b", "c"]
    # Toàn key lạ -> không ra file rỗng cột, rơi về trọn bộ
    assert [c.key for c in pick_columns(SPEC, "zzz,yyy")] == ["a", "b", "c"]


def test_pick_columns_always_luon_co_va_khong_lap():
    assert [c.key for c in pick_columns(SPEC, "b", always=["a", "b"])] == ["b", "a"]


def test_check_row_limit():
    check_row_limit(MAX_ROWS)   # đúng trần thì vẫn cho xuất
    with pytest.raises(HTTPException) as e:
        check_row_limit(MAX_ROWS + 1)
    assert e.value.status_code == 400
    assert "lọc bớt" in e.value.detail


def test_cell_value_doi_kieu_theo_cot():
    assert cell_value(Col("x", "X", "money"), {"x": Decimal("1500.5")}) == 1500.5
    assert cell_value(Col("x", "X", "money"), {}) == 0          # ô tiền trống -> 0 để cộng được
    assert cell_value(Col("x", "X", "date"), {"x": "2026-08-13"}) == datetime(2026, 8, 13).date()
    assert cell_value(Col("x", "X", "date"), {"x": ""}) is None
    assert cell_value(Col("x", "X", "date"), {"x": "hôm qua"}) == "hôm qua"   # không nuốt dữ liệu lạ
    assert cell_value(Col("x", "X", "bool"), {"x": True}) == "Có"
    assert cell_value(Col("x", "X", "bool"), {"x": False}) == ""


def test_cell_value_created_at_doi_sang_gio_vn():
    # DB lưu UTC naive; file phải ra giờ VN cho khớp màn hình
    got = cell_value(Col("created_at", "Ngày tạo", "datetime"), {"created_at": datetime(2026, 8, 13, 2, 30)})
    assert got == datetime(2026, 8, 13, 9, 30)


def _load(resp):
    return openpyxl.load_workbook(BytesIO(resp.body))


def test_xlsx_response_dung_header_dinh_dang_va_ten_file():
    cols = [Col("code", "Mã PYC"), Col("total", "Tổng tiền", "money")]
    resp = xlsx_response("yeu-cau-mua-hang", cols, [{"code": "PYC1", "total": 1000}], "Yeu cau mua hang")
    ws = _load(resp).active
    assert ws.title == "Yeu cau mua hang"
    assert [c.value for c in ws[1]] == ["Mã PYC", "Tổng tiền"]
    assert ws["A2"].value == "PYC1"
    assert ws["B2"].value == 1000                    # số giữ kiểu số, không kèm "đ"
    assert ws["B2"].number_format == "#,##0"
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref == "A1:B2"
    cd = resp.headers["content-disposition"]
    assert cd.startswith('attachment; filename="yeu-cau-mua-hang-') and cd.endswith('.xlsx"')


# ── YCMH ────────────────────────────────────────────────────────────────────────
def _pr(db, **kw):
    from app.modules.purchase_request.model import PurchaseRequest
    pr = PurchaseRequest(code=kw.pop("code", "PYC-1"), requester="Người YC", department="Phòng Test",
                         status=kw.pop("status", "approved"), **kw)
    db.add(pr)
    db.flush()
    return pr


def _pr_item(db, pr, **kw):
    from app.modules.purchase_request.model import PurchaseRequestItem
    it = PurchaseRequestItem(pr_id=pr.id, **kw)
    db.add(it)
    db.flush()
    return it


def test_ycmh_moi_dong_hang_mot_dong_excel(db):
    from app.modules.purchase_request import export as ex
    pr = _pr(db, need_date="2026-09-30", is_urgent=True)
    _pr_item(db, pr, product_code="SP1", product_name="Sản phẩm 1", qty=10, price=1000,
             vat_pct=8, amount=10800, required_date="2026-09-10")
    _pr_item(db, pr, product_code="SP2", product_name="Sản phẩm 2", qty=5, price=2000,
             vat_pct=8, amount=10800, required_date="2026-09-05")
    rows = ex.build_rows(db, [pr])

    assert len(rows) == 2
    assert [r["line_no"] for r in rows] == [1, 2]
    assert [r["product_code"] for r in rows] == ["SP1", "SP2"]
    # Cụm đầu phiếu lặp lại y hệt ở mọi dòng
    assert rows[0]["code"] == rows[1]["code"] == "PYC-1"
    assert rows[0]["is_urgent"] is True
    assert rows[0]["status"] == "Đã duyệt"


def test_ycmh_tong_tien_va_ngay_can_tinh_theo_dong(db):
    from app.modules.purchase_request import export as ex
    pr = _pr(db, need_date="2026-12-31")
    _pr_item(db, pr, product_code="SP1", product_name="SP 1", amount=10800, required_date="2026-09-10")
    _pr_item(db, pr, product_code="SP2", product_name="SP 2", amount=5000, required_date="2026-09-05")
    # Dòng đã hủy: vẫn ra file nhưng không kéo "ngày cần hàng" đầu phiếu về sớm hơn
    _pr_item(db, pr, product_code="SP3", product_name="SP 3", amount=0,
             required_date="2026-01-01", line_status="cancelled")   # B-06: cột lưu MÃ
    r = ex.build_rows(db, [pr])[0]

    assert r["total"] == 15800
    assert r["need_date"] == "2026-09-05"


def test_ycmh_phieu_chua_co_dong_van_ra_mot_hang(db):
    from app.modules.purchase_request import export as ex
    pr = _pr(db, code="PYC-RONG", status="draft")
    rows = ex.build_rows(db, [pr])
    assert len(rows) == 1
    assert rows[0]["code"] == "PYC-RONG"
    assert rows[0]["status"] == "Nháp"
    assert "line_no" not in rows[0]        # cụm dòng để trống


def test_ycmh_nstm_hien_ma_kem_ten(db, seed):
    from app.modules.purchase_request import export as ex
    pr = _pr(db)
    _pr_item(db, pr, product_code="SP1", product_name="SP 1", assignee="DEMONV")
    _pr_item(db, pr, product_code="SP2", product_name="SP 2", assignee="KHONGCO")
    rows = ex.build_rows(db, [pr])
    assert rows[0]["assignee_name"] == "DEMONV — NSTM Chính"
    assert rows[1]["assignee_name"] == "KHONGCO"     # mã lạ giữ nguyên, không nuốt


def test_ycmh_cum_dong_hang_luon_xuat_du_bang_an_cot(db):
    """Người dùng ẩn bớt cột đầu phiếu -> file bỏ theo, nhưng cụm dòng hàng vẫn còn nguyên."""
    from app.modules.purchase_request import export as ex
    cols = pick_columns(ex.HEADER_COLS, "code,status") + list(ex.LINE_COLS)
    keys = [c.key for c in cols]
    assert keys[:2] == ["code", "status"]
    assert "product_code" in keys and "amount" in keys


# ── YCBG ────────────────────────────────────────────────────────────────────────
def _sr(db, **kw):
    from app.modules.survey_request.model import SurveyRequest
    sr = SurveyRequest(code=kw.pop("code", "YCKS-1"), requester="Người YC", department="Phòng Test",
                       purpose="Mua mới", status=kw.pop("status", "survey_done"), **kw)
    db.add(sr)
    db.flush()
    return sr


def _sr_line(db, sr, **kw):
    from app.modules.survey_request.model import SurveyRequestLine
    ln = SurveyRequestLine(survey_request_id=sr.id, **kw)
    db.add(ln)
    db.flush()
    return ln


def _sr_opt(db, line, **kw):
    from app.modules.survey_request.model import SurveyRequestOption
    o = SurveyRequestOption(survey_request_line_id=line.id, **kw)
    db.add(o)
    db.flush()
    return o


def test_ycbg_chi_xuat_phuong_an_da_chot(db):
    from app.modules.survey_request import export as ex
    sr = _sr(db)
    ln = _sr_line(db, sr, item_group="Nhãn", request_qty=100, uom="cuộn")
    _sr_opt(db, ln, public_id=1, display_label="Option 1", supplier_code="NX",
            snap_product_name="Phương án bị loại", snap_price_by_volume=9000)
    _sr_opt(db, ln, public_id=2, display_label="Option 2", is_chosen=True, supplier_code="NX",
            snap_product_name="Phương án chốt", snap_price_by_volume=5000, snap_vat=8)
    rows = ex.build_rows(db, [sr])

    assert len(rows) == 1                                   # 1 dòng yêu cầu = 1 hàng, không bung theo option
    assert rows[0]["opt_label"] == "Option 2"
    assert rows[0]["opt_product_name"] == "Phương án chốt"
    assert rows[0]["opt_price"] == 5000


def test_ycbg_dong_chua_chot_thi_cum_phuong_an_de_trong(db):
    from app.modules.survey_request import export as ex
    sr = _sr(db, status="processing")
    ln = _sr_line(db, sr, item_group="Nhãn")
    _sr_opt(db, ln, public_id=1, display_label="Option 1", snap_product_name="Chưa chốt")
    r = ex.build_rows(db, [sr])[0]

    assert r["line_no"] == 1
    assert r["status"] == "Đang xử lý"
    assert "opt_product_name" not in r


def test_ycbg_bo_cot_lo_ncc_khi_khong_co_quyen_supplier_read():
    from app.modules.survey_request import export as ex
    full = {c.key for c in ex.columns_for(True)}
    masked = {c.key for c in ex.columns_for(False)}
    assert ex.SUPPLIER_ONLY <= full
    assert not (ex.SUPPLIER_ONLY & masked)
    assert masked == full - ex.SUPPLIER_ONLY          # chỉ bỏ đúng cột nhạy cảm, không bỏ lố


def _profile(scope, employee_id=0, emp_code="", write=False, approve=False):
    return {"employee_id": employee_id, "emp_code": emp_code,
            "grants": [{"perms": {"survey_request": {
                "read": True, "write": write, "approve": approve, "scope": scope}}}]}


def test_ycbg_nstm_chi_xuat_dong_cua_minh(db, seed):
    """NSTM xuất file chỉ thấy dòng được giao hoặc thuộc phân loại mình phụ trách."""
    from app.modules.survey_request import export as ex
    sr = _sr(db)
    ln_giao = _sr_line(db, sr, item_group="Sắt", assignee=seed.emp_nstm_code)
    ln_phan_loai = _sr_line(db, sr, item_group="Nhãn")          # NSTM phụ trách phân loại Nhãn
    ln_nguoi_khac = _sr_line(db, sr, item_group="Sắt", assignee="AICOKHAC")
    db.commit()

    user = SimpleNamespace(id=999, employee_id=seed.emp_nstm_id)
    prof = _profile("proc", seed.emp_nstm_id, seed.emp_nstm_code, write=True)
    rows = ex.build_rows(db, [sr], visible=ex.line_visible_fn(db, user, prof))

    assert [r["item_group"] for r in rows] == ["Sắt", "Nhãn"]
    assert [r["line_no"] for r in rows] == [1, 2]              # đánh lại STT theo dòng thấy được
    assert ln_nguoi_khac.id and ln_giao.id and ln_phan_loai.id


def test_ycbg_quan_ly_xuat_het_dong(db, seed):
    from app.modules.survey_request import export as ex
    sr = _sr(db)
    for g in ("Sắt", "Nhãn", "Thùng"):
        _sr_line(db, sr, item_group=g, assignee="AICOKHAC")
    db.commit()

    user = SimpleNamespace(id=999, employee_id=seed.emp_tp_id)
    rows = ex.build_rows(db, [sr], visible=ex.line_visible_fn(db, user, _profile("all")))
    assert len(rows) == 3


# ── ĐMH ─────────────────────────────────────────────────────────────────────────
def _po(db, **kw):
    from app.modules.purchase_order.model import PurchaseOrder
    po = PurchaseOrder(code=kw.pop("code", "PO-1"), status=kw.pop("status", "approved"), **kw)
    db.add(po)
    db.flush()
    return po


def _po_item(db, po, **kw):
    from app.modules.purchase_order.model import POItem
    it = POItem(po_id=po.id, **kw)
    db.add(it)
    db.flush()
    return it


def _po_delivery(db, it, **kw):
    from app.modules.purchase_order.model import PODelivery
    dl = PODelivery(po_id=it.po_id, po_item_id=it.id, **kw)
    db.add(dl)
    db.flush()
    return dl


def test_dmh_dung_bo_cot_cua_man_tien_do(db):
    """Yêu cầu khách: cột dòng hàng của file ĐMH = cột màn Tiến độ mua hàng."""
    from app.modules.purchase_order import export as ex
    from app.modules.purchase_progress import export as pex
    line_keys = [c.key for c in ex.LINE_COLS]
    head_keys = {c.key for c in ex.HEADER_COLS}

    # Mọi cột Tiến độ đều có mặt, hoặc giữ nguyên key, hoặc đã nằm sẵn ở cụm đầu đơn
    for c in pex.COLS:
        assert c.key in line_keys or c.key in head_keys or c.key in ("stt", "po_code", "amount")
    # `amount` của Tiến độ ("Thành tiền nhận") phải đổi key, kẻo đè "Tiền hàng" của đầu đơn
    assert "recv_amount" in line_keys and "amount" not in line_keys
    assert dict((c.key, c.label) for c in ex.LINE_COLS)["recv_amount"] == "Thành tiền nhận"
    # Thứ tự cột giữ đúng như trên màn Tiến độ
    keep = [c.key for c in pex.COLS if c.key in line_keys]
    assert keep == [k for k in line_keys if k in set(keep)]
    # Không được có hai cột trùng NHÃN, người đọc file không biết cột nào là cột nào
    labels = [c.label for c in ex.HEADER_COLS] + [c.label for c in ex.LINE_COLS]
    assert len(labels) == len(set(labels))


def test_dmh_bung_theo_lan_giao(db):
    from app.modules.purchase_order import export as ex
    po = _po(db, supplier_code="NX", pr_code="PYC-1", document_status="full")  # B-06: cột lưu MÃ
    # Đặt 10 (10.800 gồm VAT), giao 2 lần: nhận 6 rồi nhận 4
    it = _po_item(db, po, product_code="SP1", qty_order=10, price=1000, vat=8,
                  qty_received=10, qty_remaining=0, amount=10800)
    _po_delivery(db, it, delivery_no=1, received_qty=6)
    _po_delivery(db, it, delivery_no=2, received_qty=4)
    rows = ex.build_rows(db, [po])

    assert len(rows) == 2                    # một hàng cho mỗi lần giao
    assert [r["line_no"] for r in rows] == [1, 2]
    assert [r["delivery_no"] for r in rows] == [1, 2]
    assert rows[0]["order_amount"] == rows[1]["order_amount"] == 10800   # giá trị ĐẶT, lặp lại
    assert rows[0]["recv_amount"] == 6480 and rows[1]["recv_amount"] == 4320
    assert rows[0]["document_status"] == "Đã đủ chứng từ"                # nhãn của đầu đơn thắng
    assert rows[0]["status"] == "Đã duyệt"


def test_dmh_dong_chua_co_lan_giao_van_ra_mot_hang(db):
    from app.modules.purchase_order import export as ex
    po = _po(db)
    _po_item(db, po, product_code="SP1", qty_order=1, price=1, vat=0)
    r = ex.build_rows(db, [po])[0]
    assert r["line_no"] == 1
    assert r["delivery_no"] in (None, "")
    assert r["progress_status"] == "Chưa đặt hàng"


def test_dmh_tien_hang_dau_don_khong_nhan_theo_lan_giao(db):
    """Lọc "STT dòng = 1" rồi cộng "Tiền hàng" phải ra đúng tổng đơn, dù dòng có nhiều lần giao."""
    from app.modules.purchase_order import export as ex
    po = _po(db)
    it1 = _po_item(db, po, product_code="SP1", qty_order=10, price=1000, vat=8, amount=0)
    _po_item(db, po, product_code="SP2", qty_order=2, price=500, vat=0, amount=0)
    _po_delivery(db, it1, delivery_no=1, received_qty=6)
    _po_delivery(db, it1, delivery_no=2, received_qty=4)
    rows = ex.build_rows(db, [po])
    assert len(rows) == 3                                  # 2 lần giao + 1 dòng chưa giao
    assert {r["amount"] for r in rows} == {11800}          # 10.800 + 1.000, không nhân đôi
    assert [r["line_no"] for r in rows] == [1, 2, 3]


def test_dmh_thieu_ma_ncc_thi_hien_ten(db):
    from app.modules.purchase_order import export as ex
    po = _po(db, supplier_name="Nhà Xuất NX")
    r = ex.build_rows(db, [po])[0]
    assert r["supplier_code"] == "Nhà Xuất NX"
    assert "line_no" not in r                # đơn chưa có dòng vẫn ra một hàng


def test_dmh_bo_cot_ncc_van_chuyen_khi_khong_du_quyen(db):
    from app.modules.purchase_order import export as ex
    po = _po(db, supplier_code="NX", supplier_name="Nhà Xuất NX")
    it = _po_item(db, po, product_code="SP1", qty_order=1, price=1000, vat=0)
    _po_delivery(db, it, delivery_no=1, received_qty=1, carrier_code="VC1",
                 carrier_name="Giao Nhanh", shipping_amount=50000)

    full = {c.key for c in ex.line_columns(True)}
    masked = {c.key for c in ex.line_columns(False)}
    assert masked == full - ex.SUPPLIER_ONLY
    assert "supplier_code" not in ex.SUPPLIER_ONLY        # cột đầu đơn của bảng ĐMH, vẫn giữ

    r = ex.build_rows(db, [po], show_supplier=False)[0]
    for k in ex.SUPPLIER_ONLY:
        assert k not in r                                 # dữ liệu bị gỡ, không chỉ ẩn cột
    assert r["supplier_code"] == "NX"


# ── Tiến độ mua hàng ────────────────────────────────────────────────────────────
def test_tien_do_bo_cot_ncc_va_van_chuyen_khi_khong_du_quyen():
    from app.modules.purchase_progress import export as ex
    full = {c.key for c in ex.columns_for(True)}
    masked = {c.key for c in ex.columns_for(False)}
    assert ex.SUPPLIER_ONLY <= full
    assert masked == full - ex.SUPPLIER_ONLY


def test_tien_do_bo_cot_khop_voi_luat_an_cua_controller():
    """Mọi key bị controller gỡ khỏi row đều phải nằm trong danh sách cột bị bỏ khi xuất,
    kẻo file Excel ra cột trống (hoặc tệ hơn: lộ dữ liệu nếu sau này controller thôi gỡ)."""
    from app.modules.purchase_progress import controller, export as ex
    keys = {c.key for c in ex.COLS}
    assert ex.SUPPLIER_ONLY == set(controller._SUPPLIER_HIDDEN) & keys
    assert ex.SUPPLIER_ONLY <= set(controller._SUPPLIER_HIDDEN)


@pytest.mark.parametrize("module", [
    "app.modules.purchase_request.export",
    "app.modules.survey_request.export",
    "app.modules.purchase_order.export",
    "app.modules.purchase_progress.export",
])
def test_khong_co_key_cot_trung_nhau(module):
    """Key trùng làm `pick_columns` lấy nhầm cột và Excel ra 2 cột giống hệt."""
    import importlib
    m = importlib.import_module(module)
    cols = getattr(m, "COLS", None) or (list(m.HEADER_COLS) + list(getattr(m, "LINE_COLS", []))
                                        + list(getattr(m, "OPTION_COLS", [])))
    keys = [c.key for c in cols]
    assert len(keys) == len(set(keys))
