"""Cụm 04 — Tài chính · Kho · Báo cáo & Trang chủ.

Ba cụm trước lọc **DÒNG**. Cụm này lọc **SỐ TỔNG**, và đó là chỗ khác nhau căn
bản: một bảng lọc sai thì người dùng thấy một dòng lạ và đi hỏi; một ô "Tổng chi
tiêu" cộng thêm pháp nhân khác thì trên màn hình chỉ có **một con số to hơn sự
thật**, không có dòng nào để ai phát hiện ra.

Bởi vậy mọi ca ở đây dựng dữ liệu **hai pháp nhân** rồi khẳng định bằng **giá
trị số**, không khẳng định "gọi được" hay "khác rỗng". Số không đổi khi phạm vi
đổi = số không lọc.

Bốn nhóm:

* **A1–A14** — Báo cáo. 11 route, cả 11 gác bằng đúng một khóa
  `require("report", ...)`, và tầng service **không gọi `apply_scope` lần nào**
  (`report/service.py:366` nói thẳng: *"Báo cáo = TOÀN CÔNG TY (không áp scope
  user)"*). Nhóm này ghim hiện trạng ấy thành con số.
* **B1–B9** — Trang chủ. `/overview` gác TỪNG KHỐI bằng `can(entity)` rồi **bỏ
  hẳn khóa** khi thiếu quyền; `/stats` thì không gác gì cả.
* **C1–C9** — Công nợ + Yêu cầu thanh toán.
* **D1–D5** — Tồn kho + Kho + Luân chuyển kho.

⚠️ Test này KHÔNG sửa `app/core/`. Ca nào hành vi hiện tại còn phải hỏi người
dùng thì khẳng định ĐÚNG HÀNH VI HÔM NAY kèm `# QUYẾT ĐỊNH CHỜ:` — đỏ lên khi có
ai đổi, nhưng không tự nhận là đúng.

**Đã VÁ ngày 05/09/2026 (P0 #3 #4 #5 #6 + P1 #14)** — năm ca dưới đây đổi từ *ghim
hiện trạng hỏng* sang *canh không tái phát*, mỗi ca kèm vế đối chứng "không chặn nhầm
người trong phạm vi": **C5** bản in YCTT · **C6/C6b/C6c** tám đường ghi YCTT ·
**C9** cấn trừ tiền treo công nợ · **D3** luân chuyển kho · **D4** điều chỉnh tồn kho.
Nhóm A (Báo cáo) và B9 (`/dashboard/stats`) vẫn là ghim hiện trạng — chưa vá.
"""
import json
from datetime import datetime
from types import SimpleNamespace

import pytest
from fastapi import BackgroundTasks, HTTPException

from scope_factory import build_world  # noqa: F401 — fixture `world` dùng nó

YEAR = str(datetime.now().year)   # dữ liệu bám năm hiện tại: báo cáo lọc theo năm


# ══════════════════════════════════════════════════════════════════════════════
#  Khung gọi controller
# ══════════════════════════════════════════════════════════════════════════════


def make_request(**params):
    """`Request` giả — mọi controller ở cụm này chỉ đụng `request.query_params`.

    Dùng `QueryParams` thật chứ không dùng dict: `apply_filters` gọi `.get()` và
    `.items()`, còn `filter_operators.collect_conditions_map` gọi `.multi_items()`
    (một field lọc được nhiều lần). Dict thiếu hàm cuối → AttributeError giữa test.
    """
    from starlette.datastructures import QueryParams

    return SimpleNamespace(query_params=QueryParams(
        [(k, str(v)) for k, v in params.items()]))


def unwrap(resp):
    """`success()` trả `JSONResponse`, không trả dict — phải bóc thân phản hồi."""
    return json.loads(resp.body)["data"]


def read_stream_body(resp) -> bytes:
    """Gom trọn thân một `StreamingResponse` (route xuất Excel) thành bytes.

    `body_iterator` của Starlette là iterator BẤT ĐỒNG BỘ kể cả khi nội dung là
    một `BytesIO` sẵn trong bộ nhớ, nên phải chạy qua vòng lặp sự kiện.
    """
    import asyncio

    async def drain():
        return b"".join([chunk async for chunk in resp.body_iterator])

    return asyncio.run(drain())


# ══════════════════════════════════════════════════════════════════════════════
#  Dữ liệu nền: HAI pháp nhân, mỗi thứ một cặp A/B để mọi con số so được hiệu
# ══════════════════════════════════════════════════════════════════════════════


def create_purchase_order(db, **kw):
    from app.modules.purchase_order.model import PurchaseOrder

    kw.setdefault("order_date", f"{YEAR}-03-10")
    row = PurchaseOrder(status="approved", **kw)
    db.add(row)
    db.flush()
    return row


def create_po_item(db, po_id, **kw):
    from app.modules.purchase_order.model import POItem

    row = POItem(po_id=po_id, vat=0, **kw)
    db.add(row)
    db.flush()
    return row


def create_delivery(db, po_id, po_item_id, **kw):
    from app.modules.purchase_order.model import PODelivery

    row = PODelivery(po_id=po_id, po_item_id=po_item_id, **kw)
    db.add(row)
    db.flush()
    return row


def create_payable(db, **kw):
    from app.modules.payable.model import Payable

    kw.setdefault("period", YEAR)
    kw.setdefault("source_type", "goods")
    kw.setdefault("paid_amount", 0)
    row = Payable(status="unpaid", due_date="", **kw)
    db.add(row)
    db.flush()
    return row


def create_purchase_request(db, **kw):
    from app.modules.purchase_request.model import PurchaseRequest

    row = PurchaseRequest(requester_id=0, status="submitted", **kw)
    db.add(row)
    db.flush()
    return row


def create_inventory(db, **kw):
    from app.modules.inventory.model import Inventory

    row = Inventory(**kw)
    db.add(row)
    db.flush()
    return row


@pytest.fixture()
def two_company_data(world):
    """Một bộ chứng từ **đối xứng** giữa pháp nhân A và pháp nhân B.

    Đối xứng là điều kiện để mọi khẳng định đọc được thành câu: A đúng 1000, B
    đúng 2000, nên bất kỳ con số nào ra 3000 là đã gộp cả hai. Ba phiếu ĐMH chứ
    không phải hai: `po_b_kt` nằm ở pháp nhân B nhưng mang **đúng tên phòng** của
    a1 ("Phòng Kế toán"), để bắt đường lọc-theo-TÊN của `report_dept_scope`.
    """
    db = world.db
    co_a, co_b = world.co["A"], world.co["B"]

    po_a = create_purchase_order(
        db, code="PO_A", company_id=co_a, department_id=world.dept["A.kt"],
        department="Phòng Kế toán", supplier_code="NCCA", supplier_name="NCC A",
        nspt="NSPT A", is_urgent=True, created_by=world.actor("a1").user.id)
    po_b = create_purchase_order(
        db, code="PO_B", company_id=co_b, department_id=world.dept["B.hc"],
        department="Phòng Hành chính", supplier_code="NCCB", supplier_name="NCC B",
        nspt="NSPT B", order_date=f"{YEAR}-03-11", created_by=world.actor("b1").user.id)
    po_b_kt = create_purchase_order(
        db, code="PO_B_KT", company_id=co_b, department_id=world.dept["B.kt"],
        department="Phòng Kế toán", supplier_code="NCCB", supplier_name="NCC B",
        nspt="NSPT B", order_date=f"{YEAR}-03-12", created_by=world.actor("b1").user.id)

    it_a = create_po_item(db, po_a.id, product_code="SP_A", product_name="Hàng A",
                          item_group="Nhãn", qty_order=10, price=100, qty_received=10)
    it_b = create_po_item(db, po_b.id, product_code="SP_B", product_name="Hàng B",
                          item_group="Thùng", qty_order=20, price=100, qty_received=20)

    create_delivery(db, po_a.id, it_a.id, received_qty=10, received_date=f"{YEAR}-03-15",
                    diff_regulated=-1, carrier_code="VCA", carrier_name="Vận chuyển A",
                    shipping_amount=50, invoice_no="HD_A")
    create_delivery(db, po_b.id, it_b.id, received_qty=20, received_date=f"{YEAR}-03-16",
                    diff_regulated=0, carrier_code="VCB", carrier_name="Vận chuyển B",
                    shipping_amount=80, invoice_no="HD_B")

    pay_a = create_payable(db, company_id=co_a, supplier_code="NCCA", supplier_name="NCC A",
                           po_id=po_a.id, po_code="PO_A", incur_date=f"{YEAR}-03-15",
                           total=1000, remaining=1000, created_by=world.actor("a1").user.id)
    pay_b = create_payable(db, company_id=co_b, supplier_code="NCCB", supplier_name="NCC B",
                           po_id=po_b.id, po_code="PO_B", incur_date=f"{YEAR}-03-16",
                           total=2000, remaining=2000, created_by=world.actor("b1").user.id)

    pr_a = create_purchase_request(db, code="YC_A", company_id=co_a,
                                   department_id=world.dept["A.kt"], department="Phòng Kế toán",
                                   request_date=f"{YEAR}-03-05",
                                   created_by=world.actor("a1").user.id)
    pr_b = create_purchase_request(db, code="YC_B", company_id=co_b,
                                   department_id=world.dept["B.hc"], department="Phòng Hành chính",
                                   request_date=f"{YEAR}-03-06",
                                   created_by=world.actor("b1").user.id)
    pr_b_kt = create_purchase_request(db, code="YC_B_KT", company_id=co_b,
                                      department_id=world.dept["B.kt"], department="Phòng Kế toán",
                                      request_date=f"{YEAR}-03-07",
                                      created_by=world.actor("b1").user.id)

    inv_a = create_inventory(db, company_id=co_a, warehouse_code="K_A", product_code="SP_A",
                             product_name="Hàng A", qty=5, avg_cost=100, value=500)
    inv_b = create_inventory(db, company_id=co_b, warehouse_code="K_B", product_code="SP_B",
                             product_name="Hàng B", qty=7, avg_cost=100, value=700)
    db.commit()

    return SimpleNamespace(po_a=po_a, po_b=po_b, po_b_kt=po_b_kt, it_a=it_a, it_b=it_b,
                           pay_a=pay_a, pay_b=pay_b, pr_a=pr_a, pr_b=pr_b, pr_b_kt=pr_b_kt,
                           inv_a=inv_a, inv_b=inv_b)


def grant_report_reader(world, key="a1", scope="company", **kw):
    """Người xem báo cáo bó trong pháp nhân của chính mình (a1 ⇒ pháp nhân A)."""
    return world.grant(key, "report", scope=scope, actions=("read", "export"), **kw)


# ══════════════════════════════════════════════════════════════════════════════
#  A. Báo cáo — 11 route, 11 `require`, 0 lần `apply_scope`
# ══════════════════════════════════════════════════════════════════════════════


def test_a1_bao_cao_ngay_cong_ca_chi_phi_cua_phap_nhan_khac(world, two_company_data):
    """🔴 `/reports/daily` — người phạm vi pháp nhân A đọc ra tổng của CẢ A LẪN B.

    `report/controller.py:35-39`: câu truy vấn là `db.query(Payable)` trần, chỗ
    duy nhất thu hẹp theo pháp nhân là tham số `company_id` **do chính người xem
    tự truyền**. Bỏ trống — mà giao diện mặc định bỏ trống — thì con số là toàn
    tập đoàn.

    Hỏng theo kiểu không ai kêu: màn Công nợ của người này chỉ có phiếu pháp nhân
    A (đường kia CÓ `apply_scope`), nên hai màn cạnh nhau nói hai con số khác
    nhau và không màn nào tự nhận là sai.

    # QUYẾT ĐỊNH CHỜ: `report` đang khai `PUBLIC` trong `SCOPE_FIELDS` (một quyền
    # HÀNH ĐỘNG, không có bảng để lọc). Bịt lỗ này là gắn `apply_scope` theo
    # entity NGUỒN (`payable`/`purchase_order`) chứ không theo `report` — tức là
    # đổi ý nghĩa của khóa `report.read`. Cần chốt trước khi sửa.
    """
    from app.modules.report.controller import daily

    a1 = grant_report_reader(world)
    data = unwrap(daily(request=make_request(month=f"{YEAR}-03"), db=world.db, user=a1.user))

    assert data["total"] == 3000.0, "1000 (A) + 2000 (B) — phạm vi pháp nhân không cắt gì"
    assert sum(d["amount"] for d in data["days"]) == 3000.0

    # Vế đối chứng: chỗ DUY NHẤT lọc được là tham số người dùng tự chọn.
    only_a = unwrap(daily(request=make_request(month=f"{YEAR}-03", company_id=world.co["A"]),
                          db=world.db, user=a1.user))
    assert only_a["total"] == 1000.0, "lọc đúng — nhưng do người xem tự gõ, không do phạm vi"


def test_a2_bao_cao_ma_tran_hien_du_phong_ban_cua_hai_phap_nhan(world, two_company_data):
    """🔴 `/reports/matrix` — ma trận đối tượng × tháng không hề biết tới pháp nhân.

    `report_dept_scope` (`service.py:340-362`) trả `None` ngay khi phạm vi vai trò
    là `company` hoặc `all`, và `None` ở đây nghĩa là *thấy mọi phòng ban*. Người
    khai quyền đặt phạm vi «công ty» tưởng là đã bó vào pháp nhân mình; thực ra
    nhánh đó là nhánh **bỏ lọc**.

    Vế `supplier` rỗng không phải nhờ phạm vi mà nhờ `_can_see_ncc` — một cổng
    QUYỀN (`purchase_order.read`), xem A11.
    """
    from app.modules.report.controller import matrix

    a1 = grant_report_reader(world)
    snap = unwrap(matrix(request=make_request(year=YEAR, refresh=1),
                         background=BackgroundTasks(), db=world.db, user=a1.user))

    assert {r["key"] for r in snap["department"]} == {"Phòng Kế toán", "Phòng Hành chính"}
    ke_toan = next(r for r in snap["department"] if r["key"] == "Phòng Kế toán")
    assert ke_toan["orders"] == 2, "gộp ĐMH của pháp nhân A với ĐMH trùng tên phòng của B"
    assert snap["supplier"] == [], "NCC bị ẩn vì thiếu `purchase_order.read`, KHÔNG vì phạm vi"


def test_a3_pham_vi_phong_ban_cua_bao_cao_khop_theo_TEN_nen_lan_sang_phap_nhan_khac(
        world, two_company_data):
    """🔴 Nhánh DUY NHẤT có lọc của báo cáo lại lọc bằng **tên phòng**.

    `report_dept_scope` gom `prof["dept_name"]` — một CHUỖI — rồi
    `compute_request_matrix:417` giữ dòng nào `r["key"] in allow`. Khóa `key`
    cũng là tên phòng. Hai pháp nhân đặt tên phòng theo cùng một khuôn (hệ thật
    có 11 pháp nhân) thì trưởng phòng Kế toán công ty A đọc luôn số của phòng Kế
    toán công ty B, gộp chung một dòng nên không cách nào tách ra mà nhìn.

    Ca này là bản thu nhỏ của rủi ro cả nhóm A: chỗ nào báo cáo *có* lọc thì lọc
    theo tên, chỗ nào *không* lọc thì không lọc gì.

    # QUYẾT ĐỊNH CHỜ: báo cáo gom dòng theo TÊN phòng (N-008 chưa dời sang id).
    # Bó theo pháp nhân đòi đổi khóa gom sang `department_id` — việc của lần dựng
    # lại màn Báo cáo, hay vá tạm bằng cách AND thêm `company_id` của người xem?
    """
    from app.modules.report.controller import request_matrix

    a1 = grant_report_reader(world, scope="dept")
    data = unwrap(request_matrix(request=make_request(kind="pyc", year=YEAR),
                                 db=world.db, user=a1.user))

    assert [r["key"] for r in data["rows"]] == ["Phòng Kế toán"], "đã bó về phòng của mình"
    assert data["rows"][0]["total"] == 2, (
        "nhưng 2 phiếu: YC_A (pháp nhân A) + YC_B_KT (pháp nhân B, TRÙNG TÊN phòng)")


def test_a4_bao_cao_ycmh_pham_vi_cong_ty_khong_cat_gi(world, two_company_data):
    """🔴 `/reports/request-matrix` với phạm vi `company` đếm cả ba phiếu YCMH.

    Cùng gốc với A2 nhưng trên tập chứng từ khác, và đây là chỗ đau hơn: bảng
    danh sách YCMH của chính người này (`purchase_request` CÓ `apply_scope`) chỉ
    ra 1 phiếu, còn ô tổng trên Báo cáo ghi 3.
    """
    from app.modules.report.controller import request_matrix

    a1 = grant_report_reader(world)
    a1.grant("purchase_request", scope="company")

    data = unwrap(request_matrix(request=make_request(kind="pyc", year=YEAR),
                                 db=world.db, user=a1.user))
    assert sum(r["total"] for r in data["rows"]) == 3

    from app.modules.purchase_request.model import PurchaseRequest
    assert a1.sees(PurchaseRequest, "purchase_request") == {two_company_data.pr_a.id}, (
        "cùng một người, cùng một tập phiếu: danh sách ra 1, báo cáo ra 3")


def test_a5_bao_cao_ycmh_theo_khoang_ngay_cung_khong_cat_theo_phap_nhan(world, two_company_data):
    """🔴 `/reports/request-range` — bản phẳng của A4, đi qua đúng
    `_req_scoped_rows` (`service.py:365`), hàm mà chính docstring của nó ghi
    *"Báo cáo = TOÀN CÔNG TY (không áp scope user)"*.

    Tách thành ca riêng vì màn Báo cáo có hai bộ lọc (theo năm / theo khoảng
    ngày) chạy hai hàm khác nhau — vá một hàm mà quên hàm kia là vá nửa lỗ.
    """
    from app.modules.report.controller import request_range

    a1 = grant_report_reader(world)
    rows = unwrap(request_range(request=make_request(kind="pyc", date_from=f"{YEAR}-01-01",
                                                     date_to=f"{YEAR}-12-31"),
                                db=world.db, user=a1.user))
    assert {r["key"]: r["total"] for r in rows} == {"Phòng Kế toán": 2, "Phòng Hành chính": 1}


def test_a6_bao_cao_nspt_theo_khoang_ngay_lo_nspt_cua_phap_nhan_khac(world, two_company_data):
    """🔴 `/reports/nspt-range` — lộ cả DANH SÁCH TÊN người phụ trách bên kia.

    `compute_nspt_range` (`service.py:191-228`) chỉ lọc theo `company_id` truyền
    vào; không truyền thì `po_by` gom đơn của mọi pháp nhân. Ở đây thứ rò ra
    không chỉ là số: khóa `key` chính là tên NSPT, tức là sơ đồ nhân sự thu mua
    của công ty khác.
    """
    from app.modules.report.controller import nspt_range

    a1 = grant_report_reader(world)
    rows = unwrap(nspt_range(request=make_request(date_from=f"{YEAR}-01-01",
                                                  date_to=f"{YEAR}-12-31"),
                             db=world.db, user=a1.user))
    assert {r["key"]: r["orders"] for r in rows} == {"NSPT A": 1, "NSPT B": 1}


def test_a7_bao_cao_phan_loai_theo_khoang_ngay_cong_chi_phi_hai_phap_nhan(world, two_company_data):
    """🔴 `/reports/ig-range` — chi phí theo loại VTBB/NL gộp cả hai pháp nhân.

    Khẳng định bằng TIỀN chứ không bằng số dòng: hai pháp nhân mua hai loại khác
    nhau nên đếm dòng vẫn "có vẻ" hợp lý, chỉ khi cộng tiền mới thấy 2000 đồng
    của pháp nhân B nằm trong bảng của người chỉ được xem pháp nhân A.
    """
    from app.modules.report.controller import ig_range

    a1 = grant_report_reader(world)
    rows = unwrap(ig_range(request=make_request(date_from=f"{YEAR}-01-01",
                                                date_to=f"{YEAR}-12-31"),
                           db=world.db, user=a1.user))
    assert {r["key"]: r["cost"] for r in rows} == {"Nhãn": 1000.0, "Thùng": 2000.0}


def test_a8_bao_cao_ncc_theo_khoang_ngay_gac_bang_QUYEN_chu_khong_bang_pham_vi(
        world, two_company_data):
    """🔴 `/reports/sup-range` — cổng duy nhất là `_can_see_ncc`, một cổng QUYỀN.

    `controller.py:166` trả `[]` khi thiếu `purchase_order.read`. Có khóa đó rồi
    thì **phạm vi của nó không được đọc tới**: người xem ĐMH bó trong pháp nhân A
    vẫn đọc trọn bảng đánh giá giao hàng của NCC pháp nhân B.

    Hai vế trong một ca là cố ý: tách ra thì vế sau dễ bị đọc nhầm thành "thiếu
    quyền nên rỗng", mà cái phải ghim là *có quyền thì phạm vi vô nghĩa*.
    """
    from app.modules.report.controller import sup_range

    a1 = grant_report_reader(world)
    req = make_request(date_from=f"{YEAR}-01-01", date_to=f"{YEAR}-12-31")
    assert unwrap(sup_range(request=req, db=world.db, user=a1.user)) == [], "chưa có khóa ĐMH"

    a1.grant("purchase_order", scope="company")
    rows = unwrap(sup_range(request=req, db=world.db, user=a1.user))
    assert {r["key"]: r["trans"] for r in rows} == {"NCC A": 1, "NCC B": 1}

    from app.modules.purchase_order.model import PurchaseOrder
    assert a1.sees(PurchaseOrder, "purchase_order") == {two_company_data.po_a.id}, (
        "phạm vi ĐMH của người này chỉ có 1 đơn, báo cáo NCC vẫn dựng trên 2")


def test_a9_bao_cao_bo_phan_theo_khoang_ngay_dem_ca_don_phap_nhan_khac(world, two_company_data):
    """🔴 `/reports/dept-range` — đặt hàng & đơn gấp theo bộ phận, không lọc pháp nhân.

    `compute_dept_range` (`service.py:297-319`) gom theo `p.department` (chuỗi),
    nên hai phòng trùng tên khác pháp nhân **nhập làm một dòng**: cột "đơn gấp"
    của phòng Kế toán A bị pha loãng bởi đơn thường của phòng Kế toán B, tỷ lệ
    gấp tụt từ 100% xuống 50% mà không ai biết vì sao.
    """
    from app.modules.report.controller import dept_range

    a1 = grant_report_reader(world)
    rows = unwrap(dept_range(request=make_request(date_from=f"{YEAR}-01-01",
                                                  date_to=f"{YEAR}-12-31"),
                             db=world.db, user=a1.user))
    by_key = {r["key"]: r for r in rows}
    assert set(by_key) == {"Phòng Kế toán", "Phòng Hành chính"}
    assert by_key["Phòng Kế toán"]["orders"] == 2
    assert by_key["Phòng Kế toán"]["rate"] == 50.0, "1 đơn gấp / 2 đơn — mẫu số có đơn của B"


def test_a10_chi_tiet_van_chuyen_lo_ten_don_vi_van_chuyen_cua_phap_nhan_khac(
        world, two_company_data):
    """🔴 `/reports/shipping-detail` — bảng CHI TIẾT, không phải số tổng.

    Đây là route báo cáo duy nhất trả ra từng dòng chứng từ (số hóa đơn, mã hàng,
    ngày nhận). Không lọc phạm vi ở đây nghĩa là lộ dữ liệu ở mức thô nhất, và
    dropdown `carriers` còn liệt kê sẵn tên đơn vị vận chuyển của pháp nhân kia.
    """
    from app.modules.report.controller import shipping_detail

    a1 = grant_report_reader(world)
    data = unwrap(shipping_detail(request=make_request(year=YEAR), db=world.db, user=a1.user))

    assert data["total"] == 2
    assert sorted(data["carriers"]) == ["Vận chuyển A", "Vận chuyển B"]
    assert sorted(r["invoice_no"] for r in data["items"]) == ["HD_A", "HD_B"]


def test_a11_bao_cao_mua_hang_cong_moi_the_so_cua_hai_phap_nhan(world, two_company_data):
    """🔴 `/reports/procurement` — route nặng nhất, sáu khối số cùng lệch một lượt.

    `controller.py:223-229` dựng ba câu truy vấn trần (ĐMH · công nợ · tồn kho)
    và chỉ thu hẹp khi có `company_id` trên URL. Ca này khẳng định **từng thẻ
    số** chứ không khẳng định tổng thể: giá trị đặt hàng, nợ hàng hóa, giá trị
    tồn kho là ba nguồn dữ liệu khác nhau, vá một chỗ không kéo theo hai chỗ kia.
    """
    from app.modules.report.controller import procurement

    a1 = grant_report_reader(world)
    data = unwrap(procurement(request=make_request(year=YEAR), db=world.db, user=a1.user))

    assert data["po_count"] == 3
    assert data["order_value"] == 3000.0, "1000 (A) + 2000 (B)"
    assert data["payable_goods"]["total"] == 3000.0
    assert data["inventory_value"] == 1200.0, "500 (kho A) + 700 (kho B)"
    assert {r["key"] for r in data["by_supplier"]} == {"NCC A", "NCC B"}
    assert {r["key"] for r in data["by_nspt"]} == {"NSPT A", "NSPT B"}


def test_a12_xuat_excel_ghi_ra_tep_dung_pham_vi_da_lo_tren_man_hinh(world, two_company_data):
    """🔴 `/reports/export` — nội dung tệp KHÔNG hẹp hơn màn hình, và đó là vấn đề.

    Kiểm bằng cách mở workbook thật rồi soi ô, không kiểm "có trả về tệp": lỗ ở
    đây là dữ liệu bên trong. Tệp xlsx rời khỏi hệ thống (gửi mail, up drive) nên
    mọi thứ lọt vào nó không còn cách nào thu hồi — nặng hơn hẳn một con số hiện
    trên màn.
    """
    from io import BytesIO

    from openpyxl import load_workbook

    from app.modules.report.controller import export_excel

    a1 = grant_report_reader(world)
    resp = export_excel(request=make_request(sheet="supplier", year=YEAR),
                        db=world.db, user=a1.user)
    wb = load_workbook(BytesIO(read_stream_body(resp)))
    cells = {str(c.value) for ws in wb.worksheets for row in ws.iter_rows() for c in row}

    assert "NCC A" in cells
    assert "NCC B" in cells, "NCC của pháp nhân B nằm trong tệp người pháp nhân A tải về"


def test_a13_o_loai_tru_phong_ban_khong_an_vao_bao_cao(world, two_company_data):
    """🔴 Ô «Loại trừ phòng ban» của hộp thoại phạm vi là ô CÂM trên Báo cáo.

    `report_dept_scope` chỉ đọc `scope["inc"]["department_name"]`
    (`service.py:360`); nhánh `exc` không được nhắc tới một lần nào. Ô loại trừ
    sinh ra để giấu phòng Kế toán / Nhân sự khỏi người xem rộng — nó im lặng mất
    tác dụng đúng ở màn tổng hợp mọi con số tiền.

    # QUYẾT ĐỊNH CHỜ: `report_dept_scope` có nên trừ `exc.department_name` khỏi
    # tập `allow`, và làm gì với nhánh `None` (scope company/all) — hôm nay nhánh
    # đó bỏ qua cả include lẫn exclude?
    """
    from app.modules.report.controller import request_range

    a1 = grant_report_reader(world, scope="dept", exc_dept=["A.kt"])
    rows = unwrap(request_range(request=make_request(kind="pyc", date_from=f"{YEAR}-01-01",
                                                     date_to=f"{YEAR}-12-31"),
                                db=world.db, user=a1.user))
    assert [r["key"] for r in rows] == ["Phòng Kế toán"], "phòng bị loại trừ vẫn còn nguyên"


def test_a14_khoa_report_khong_ke_thua_khoa_nghiep_vu(world, two_company_data):
    """Khóa `report.read` là khóa RIÊNG — không có `purchase_request.read` vẫn
    đọc được số liệu YCMH của toàn hệ.

    Ghim để phân biệt rạch ròi hai chuyện: các ca trên nói *phạm vi* không được
    áp; ca này nói ngay cả *quyền entity nguồn* cũng không được hỏi tới. Ai định
    vá bằng "kiểm thêm quyền nghiệp vụ" thì đây là mốc so trước/sau.
    """
    from app.modules.report.controller import request_matrix

    a1 = grant_report_reader(world)
    assert a1.profile()["perms_union"].get("purchase_request") is None

    data = unwrap(request_matrix(request=make_request(kind="pyc", year=YEAR),
                                 db=world.db, user=a1.user))
    assert sum(r["total"] for r in data["rows"]) == 3


# ══════════════════════════════════════════════════════════════════════════════
#  B. Trang chủ — gác từng khối bằng `can(entity)`, bỏ hẳn khóa khi thiếu quyền
# ══════════════════════════════════════════════════════════════════════════════


def read_overview(db, user) -> dict:
    from app.modules.dashboard.controller import overview

    return unwrap(overview(db=db, user=user))


def test_b1_thieu_quyen_thi_khoa_VANG_MAT_chu_khong_phai_bang_khong(world, two_company_data):
    """Phân biệt "không được xem" với "xem được, bằng 0" — hai thứ khác hẳn nhau.

    `/overview` chỉ đòi đăng nhập rồi gác từng khối bằng `can(entity)`; thiếu
    quyền thì khóa **không được đặt vào `kpi`**. FE đọc `can` để chọn ẩn khối hay
    hiện số 0. Nếu ai đó "cho gọn" bằng cách gán 0 mặc định thì người dùng thấy
    một con số sai (0 đồng nợ) thay vì một khối vắng mặt — và không ai ăn 403 để
    biết mình thiếu quyền.
    """
    a1 = world.grant("a1", "payable", scope="all")
    data = read_overview(world.db, a1.user)

    assert set(data["kpi"]) == {"due_soon", "overdue"}, "chỉ khối Công nợ có mặt"
    assert "inv_value" not in data["kpi"] and "pr_pending" not in data["kpi"]
    assert data["can"] == {"purchase_request": False, "purchase_order": False, "survey": False,
                           "survey_request": False, "payable": True, "contract": False,
                           "inventory": False, "report": False}


def test_b2_moi_khoi_chi_hien_khi_co_dung_khoa_cua_no(world, two_company_data):
    """Cấp đủ ba khóa thì đủ ba nhóm khóa `kpi` — vế đối chứng của B1.

    Không có ca này thì B1 xanh cả khi `/overview` hỏng hoàn toàn và không dựng
    nổi khóa nào.
    """
    a1 = world.grant("a1", "payable", scope="all")
    a1.grant("inventory", scope="all")
    a1.grant("purchase_request", scope="all")

    kpi = read_overview(world.db, a1.user)["kpi"]
    assert set(kpi) == {"due_soon", "overdue", "inv_value", "out_of_stock", "pr_pending"}
    assert kpi["inv_value"] == 1200.0 and kpi["pr_pending"] == 3


def test_b3_top_ncc_chi_tieu_va_top_ncc_no_nam_o_hai_khoi_quyen_khac_nhau(world, two_company_data):
    """`top_suppliers` = CHI TIÊU (khối `purchase_order`) · `top_debt_suppliers` =
    NỢ CÒN LẠI (khối `payable`). Hai khóa cạnh nhau trong cùng một phản hồi, tên
    gần giống nhau, nguồn dữ liệu và khóa quyền thì khác hẳn.

    Đọc nhầm không ai ăn 403 — chỉ thấy bảng rỗng vĩnh viễn (đúng lỗi đã xảy ra ở
    Đ-11). Ca này khẳng định cả GIÁ TRỊ để hai bên không thể tình cờ bằng nhau:
    chi tiêu của NCC A là 1000 (dòng đơn), còn nợ là 1000 nhưng của một khóa khác.
    """
    ke_toan = world.grant("a1", "payable", scope="all")
    thu_mua = world.grant("a3", "purchase_order", scope="all")

    d_kt = read_overview(world.db, ke_toan.user)
    assert d_kt["top_debt_suppliers"] == [{"name": "NCC B", "value": 2000},
                                          {"name": "NCC A", "value": 1000}]
    assert d_kt["top_suppliers"] == [], "kế toán không có khóa ĐMH → khối chi tiêu vắng"

    d_tm = read_overview(world.db, thu_mua.user)
    assert d_tm["top_suppliers"] == [{"name": "NCC B", "value": 2000},
                                     {"name": "NCC A", "value": 1000}]
    assert d_tm["top_debt_suppliers"] == [], "thu mua không có khóa Công nợ → khối nợ vắng"


def test_b4_pham_vi_cong_ty_lam_doi_moi_con_so_tren_trang_chu(world, two_company_data):
    """Điều kiện "xong" của cả cụm: mỗi con số phải ĐỔI khi phạm vi đổi.

    a1 (pháp nhân A) và b1 (pháp nhân B) cùng bộ khóa, cùng bậc `company`. Mọi ô
    phải lệch nhau đúng theo dữ liệu của từng pháp nhân — trùng nhau nghĩa là ô
    đó không đi qua `apply_scope`.
    """
    a1 = world.grant("a1", "payable", scope="company")
    a1.grant("inventory", scope="company")
    a1.grant("purchase_order", scope="company")
    b1 = world.grant("b1", "payable", scope="company")
    b1.grant("inventory", scope="company")
    b1.grant("purchase_order", scope="company")

    da, db_ = read_overview(world.db, a1.user), read_overview(world.db, b1.user)

    assert da["kpi"]["inv_value"] == 500.0 and db_["kpi"]["inv_value"] == 700.0
    assert da["ap_aging"] == [{"label": "Chưa đến hạn", "value": 1000},
                              {"label": "1–30 ngày", "value": 0},
                              {"label": "31–60 ngày", "value": 0},
                              {"label": "> 60 ngày", "value": 0}]
    assert db_["ap_aging"][0] == {"label": "Chưa đến hạn", "value": 2000}
    assert da["top_debt_suppliers"] == [{"name": "NCC A", "value": 1000}]
    assert db_["top_debt_suppliers"] == [{"name": "NCC B", "value": 2000}]
    assert [p["code"] for p in da["recent_pos"]] == ["PO_A"]
    assert sorted(p["code"] for p in db_["recent_pos"]) == ["PO_B", "PO_B_KT"]


def test_b5_loai_tru_phong_ban_tru_dung_khoi_chi_tieu_theo_bo_phan(world, two_company_data):
    """Khối *Chi tiêu theo bộ phận* dựng trên `apply_scope(PurchaseOrder)` nên ô
    loại trừ phòng ban ăn thật — khác hẳn Báo cáo ở A13.

    Đặt cạnh A13 là có chủ ý: cùng một ô trên cùng một hộp thoại, ăn ở Trang chủ
    và câm ở Báo cáo. Người khai quyền không có cách nào biết điều đó.
    """
    a1 = world.grant("a1", "purchase_order", scope="all", exc_dept=["B.hc"])
    data = read_overview(world.db, a1.user)

    assert {r["name"] for r in data["dept_spend"]} == {"Phòng Kế toán"}
    assert data["dept_spend"] == [{"name": "Phòng Kế toán", "value": 1000}], (
        "chỉ còn ĐMH phòng Kế toán; PO_B_KT không có dòng hàng nên không cộng tiền")


def test_b6_duyet_ycmh_ngoai_pham_vi_bi_chan_du_co_khoa_duyet(world, two_company_data):
    """Thao tác nhanh *Duyệt* trên Trang chủ bấm thẳng vào `/purchase-requests/{id}/approve`.

    Danh sách `pending_prs_list` đã lọc phạm vi, nhưng id thì nằm trong HTML —
    chốt thật phải ở endpoint. `_in_approve_scope` (`purchase_request/controller.py:48-53`)
    là chốt đó; gỡ nó thì mọi phiếu chờ duyệt của toàn hệ duyệt được bằng một
    lệnh curl.
    """
    from app.modules.purchase_request.controller import approve_pr
    from app.modules.purchase_request.schema import ApproveIn

    a1 = world.grant("a1", "purchase_request", scope="company", actions=("read", "approve"))
    data = read_overview(world.db, a1.user)
    assert [p["code"] for p in data["pending_prs_list"]] == ["YC_A"], "danh sách đã lọc"

    with pytest.raises(HTTPException) as err:
        approve_pr(pid=two_company_data.pr_b.id, data=ApproveIn(), background_tasks=BackgroundTasks(),
                   db=world.db, user=a1.user)
    assert err.value.status_code == 403


def test_b7_tra_lai_ycmh_ngoai_pham_vi_cung_bi_chan(world, two_company_data):
    """Nút *Trả lại* đi đường khác nút *Duyệt* (`_ensure_can_return_or_reject`),
    nên phải có ca riêng — bịt một đường mà quên đường kia là bịt nửa lỗ.

    ⚠️ Chốt này chỉ đứng vững khi người thao tác **không** có quyền `cancel`: ai
    có `cancel` thì hàm trả về ngay ở dòng đầu, không hỏi phạm vi. Đó là thiết kế
    ("Quản lý làm được mọi giai đoạn"), và cũng là lý do ca này cấp đúng
    `read + approve`.
    """
    from app.modules.purchase_request.controller import return_pr
    from app.modules.purchase_request.schema import ReasonIn

    a1 = world.grant("a1", "purchase_request", scope="company", actions=("read", "approve"))
    with pytest.raises(HTTPException) as err:
        return_pr(pid=two_company_data.pr_b.id, data=ReasonIn(reason="x"),
                  background_tasks=BackgroundTasks(), db=world.db, user=a1.user)
    assert err.value.status_code == 403


def test_b8_viec_can_lam_chi_ra_chung_tu_trong_pham_vi(world, two_company_data):
    """`/dashboard/tasks` không có `require` nào — chốt duy nhất là `can(entity)`
    + `apply_scope` bên trong `build_my_tasks`.

    Tab này là nơi người dùng bấm vào nhiều nhất mỗi sáng, và mỗi việc mang kèm
    `link` dẫn thẳng tới chứng từ. Lọt một dòng ở đây là vừa lộ mã phiếu vừa mời
    người ta bấm vào.
    """
    from app.modules.dashboard.controller import build_my_tasks

    a1 = world.grant("a1", "purchase_request", scope="company")
    tasks = build_my_tasks(world.db, a1.user, a1.profile())

    assert [t["code"] for t in tasks] == ["YC_A"]
    assert all(t["type"] == "pr" for t in tasks), "chỉ khối YCMH — các khóa khác chưa cấp"


def test_b9_thong_ke_trang_chu_cu_gac_tung_khoi_va_bo_khoa_khi_thieu_quyen(world, two_company_data):
    """Bài giữ của lỗ B9 — **đã vá 05/09/2026**.

    Trước bản vá, `/dashboard/stats` (`dashboard/controller.py:12-13`) chỉ đòi
    `get_current_user`; bên trong là `db.query(...).count()` trần. Người **không
    có một khóa nào** đếm được trọn nhà cung cấp, sản phẩm, nhân sự, YCMH và ĐMH
    của cả tập đoàn. Route này phục vụ Trang chủ bản `frontend/` cũ (đang đóng
    băng) nên dễ bị bỏ quên khi rà soát — nhưng nó vẫn đăng ký trong `main.py`.

    Vá theo khuôn route anh em `/overview` **nằm cùng tệp**: gác từng khối bằng
    `can(entity)` rồi `apply_scope` theo entity nguồn, và khi thiếu quyền thì
    **BỎ HẲN KHÓA** thay vì trả `0`.

    ⚠️ Bỏ khóa chứ không trả `0` là điểm chính, không phải chi tiết trình bày:
    `0` nói dối — nó bảo "không có dữ liệu" trong khi sự thật là "bạn không được
    xem". Khóa vắng mặt buộc nơi gọi phải phân biệt hai chuyện đó. Map `can` trả
    kèm chính là để giao diện nói ra được sự khác nhau ấy.
    """
    from app.modules.dashboard.controller import stats

    khong_quyen = world.actor("a1")
    assert khong_quyen.profile()["perms_union"] == {}, "tài khoản trắng, không một khóa nào"

    data = unwrap(stats(days="all", db=world.db, user=khong_quyen.user))
    for khoa in ("suppliers", "products", "employees", "pr_total", "po_ordered"):
        assert khoa not in data, (
            f"khóa {khoa!r} phải VẮNG MẶT khi thiếu quyền, không được trả 0 — "
            "0 nói dối là 'không có dữ liệu'")
    assert data["can"] == {k: False for k in data["can"]}, "map can phải nói rõ là không có quyền"


def test_b9b_co_quyen_thi_van_dem_duoc_nhung_chi_trong_pham_vi(world, two_company_data):
    """VẾ ĐỐI CHỨNG của B9 — thiếu nó thì bản vá khóa sạch cũng xanh.

    Hai điều phải đúng cùng lúc: người có khóa vẫn nhận được số, **và** số đó
    chỉ gộp phần trong phạm vi của họ.
    """
    from app.modules.dashboard.controller import stats

    a1 = world.grant("a1", "purchase_request", scope="company", actions=("read",))
    data = unwrap(stats(days="all", db=world.db, user=a1.user))

    assert data["can"]["purchase_request"] is True
    assert "pr_total" in data, "có khóa thì phải có số"
    assert data["pr_total"] < 3, (
        "chỉ đếm YCMH của pháp nhân mình — trước bản vá chỗ này đếm cả 3 phiếu "
        "của mọi pháp nhân")
    assert "employees" not in data, "khóa khác vẫn phải vắng mặt"


# ══════════════════════════════════════════════════════════════════════════════
#  C. Công nợ (`payable`) + Yêu cầu thanh toán (`payment_request`)
# ══════════════════════════════════════════════════════════════════════════════


def test_c1_cong_no_danh_sach_va_o_tong_cung_bo_theo_phap_nhan(world, two_company_data):
    """Danh sách và ô tổng của màn Công nợ phải nói CÙNG một con số.

    Cả hai đi qua `_filtered` (`payable/controller.py:39-42`) nên dùng chung một
    lần `apply_scope` — ca này khóa ràng buộc đó lại. Tách hai đường (ví dụ tối
    ưu ô tổng bằng câu SQL riêng) là kiểu sửa làm hỏng đúng chỗ này.
    """
    from app.modules.payable.controller import list_payables, summary

    a1 = world.grant("a1", "payable", scope="company")
    pg = {"offset": 0, "limit": 50}

    rows = unwrap(list_payables(request=make_request(year=YEAR), pg=pg,
                                db=world.db, user=a1.user))
    assert [r["id"] for r in rows["items"]] == [two_company_data.pay_a.id]
    assert rows["total"] == 1

    tong = unwrap(summary(request=make_request(year=YEAR), db=world.db, user=a1.user))
    assert tong == {"total": 1000.0, "paid": 0.0, "remaining": 1000.0, "overdue": 0.0}


def test_c2_go_thang_id_khoan_no_ngoai_pham_vi_vao_URL_khong_ra(world, two_company_data):
    """Tham số `ids` của màn Công nợ là danh sách id gõ thẳng trên URL (CR-025,
    dùng cho màn "Tạo yêu cầu thanh toán" mở lại sau F5).

    `apply_scope` chạy TRƯỚC khi lọc `ids` nên id ngoài phạm vi rơi ra ngoài. Đảo
    thứ tự hai bước ấy là mở cửa hậu: lọc `ids` xong mới lọc phạm vi thì vẫn
    đúng, nhưng bỏ `apply_scope` đi vì "đã lọc theo ids rồi" thì hỏng — ca này
    canh đúng chuyện đó.
    """
    from app.modules.payable.controller import list_payables

    a1 = world.grant("a1", "payable", scope="company")
    ids = f"{two_company_data.pay_a.id},{two_company_data.pay_b.id}"
    rows = unwrap(list_payables(request=make_request(ids=ids), pg={"offset": 0, "limit": 50},
                                db=world.db, user=a1.user))
    assert [r["id"] for r in rows["items"]] == [two_company_data.pay_a.id]


def test_c3_tao_yctt_tu_khoan_no_ngoai_pham_vi_bi_chan(world, two_company_data):
    """Tick chọn dòng ngoài phạm vi rồi *Tạo đề nghị thanh toán* → 403.

    Chốt nằm ở `payment_request/controller.py:156-163` (bao-CR-274), KHÔNG nằm ở
    service: `service.create_requests` lấy khoản nợ bằng `db.get` nên tự nó
    không biết phạm vi. Gỡ đoạn đối chiếu ở controller là kéo được nợ pháp nhân
    khác vào phiếu của mình.
    """
    from app.modules.payment_request.controller import create_
    from app.modules.payment_request.schema import LineIn, PRequestCreate

    a1 = world.grant("a1", "payable", scope="company")
    a1.grant("payment_request", scope="company", actions=("read", "create"))

    payload = PRequestCreate(request_date=f"{YEAR}-04-01",
                             lines=[LineIn(payable_id=two_company_data.pay_b.id, amount=2000)])
    with pytest.raises(HTTPException) as err:
        create_(data=payload, db=world.db, user=a1.user)
    assert err.value.status_code == 403


def create_payment_request(db, **kw):
    from app.modules.payment_request.model import PaymentRequest

    kw.setdefault("status", "draft")
    kw.setdefault("request_date", f"{YEAR}-04-01")
    row = PaymentRequest(**kw)
    db.add(row)
    db.flush()
    db.commit()
    return row


def test_c4_chi_tiet_yctt_ngoai_pham_vi_bi_chan(world, two_company_data):
    """Vế ĐÚNG của cụm YCTT — mốc so cho C5/C6 ngay dưới.

    `get_(rid)` đi qua `apply_scope` (`controller.py:115-118`). Không có ca này
    thì hai ca dưới có thể bị đọc thành "phạm vi YCTT không hoạt động", trong khi
    sự thật hẹp hơn nhiều: nó hoạt động ở đường XEM và biến mất ở đường IN + mọi
    đường GHI.
    """
    from app.modules.payment_request.controller import get_

    req_b = create_payment_request(world.db, code="YCTT_B", company_id=world.co["B"],
                                   supplier_code="NCCB", supplier_name="NCC B", total=2000)
    a1 = world.grant("a1", "payment_request", scope="company")

    with pytest.raises(HTTPException) as err:
        get_(rid=req_b.id, db=world.db, user=a1.user)
    assert err.value.status_code == 403


def test_c5_ban_in_yctt_ngoai_pham_vi_bi_chan(world, two_company_data):
    """`/payment-requests/{rid}/print` phải hỏi phạm vi như đường xem — vá 05/09/2026.

    Trước bản vá route đọc bằng `service.get_request` (`db.get` trần), nên cùng một
    phiếu: `GET /{rid}` trả 403 (C4) còn `GET /{rid}/print` trả **đủ dữ liệu** — mà bản
    in còn RỘNG hơn màn chi tiết: tên pháp nhân, mã số thuế, **số tài khoản ngân hàng
    NCC**, chức vụ + trưởng phòng người lập.

    Khóa `print` là một action riêng trong ma trận phân quyền, nhưng nó chỉ trả lời
    "được in hay không", không trả lời "được in phiếu NÀO" — câu sau là việc của phạm vi.

    404 chứ không 403 theo đúng quy ước `get_scoped`: người ngoài phạm vi không cần biết
    phiếu đó có thật hay không.
    """
    from app.modules.payment_request.controller import print_

    req_a = create_payment_request(world.db, code="YCTT_A", company_id=world.co["A"],
                                   supplier_code="NCCA", supplier_name="NCC A", total=1000)
    req_b = create_payment_request(world.db, code="YCTT_B", company_id=world.co["B"],
                                   supplier_code="NCCB", supplier_name="NCC B", total=2000)
    a1 = world.grant("a1", "payment_request", scope="company", actions=("read", "print"))

    with pytest.raises(HTTPException) as err:
        print_(rid=req_b.id, db=world.db, user=a1.user)
    assert err.value.status_code == 404

    #  Vế đối chứng — bản vá không được khóa nhầm người TRONG phạm vi. Thiếu vế này
    #  thì một bản vá "chặn sạch" cũng xanh, rồi bị gỡ ra ở lần đầu có người kêu.
    data = unwrap(print_(rid=req_a.id, db=world.db, user=a1.user))
    assert data["code"] == "YCTT_A" and data["company"]["name"] == "Công ty A"


def test_c6_moi_duong_GHI_cua_yctt_deu_hoi_pham_vi(world, two_company_data):
    """Tám đường GHI của YCTT đều phải nạp phiếu qua `get_scoped` — vá 05/09/2026.

    Trước bản vá `payment_request/service.py` không có một lần `apply_scope` nào: cả bốn
    hàm ghi đi qua `db.get(PaymentRequest, rid)`, còn `require(...)` thì chỉ hỏi "có khóa
    `approve` không", không hỏi "trên phiếu nào". Kế toán pháp nhân A **duyệt chi** rồi
    **xóa** trọn chứng từ của pháp nhân B, trong khi chỉ ĐỌC phiếu đó thì bị 403.

    Ca này đi hết tám cửa (kể cả xóa hàng loạt) vì chúng là tám route riêng: bịt bảy cái
    rồi quên một cái là lỗ vẫn còn nguyên, chỉ khó thấy hơn.
    """
    from app.modules.payment_request.controller import (
        approve_, bulk_delete_requests, delete_, pay_, refund_, reject_, submit_, update_)
    from app.modules.payment_request.schema import PRequestUpdate

    from app.modules.payment_request.model import PaymentRequest

    #  Hai phiếu vì luật NGHIỆP VỤ khóa sửa sau khi duyệt — luật đó có thật và
    #  vẫn chạy; thứ ca này canh là luật PHẠM VI đứng TRƯỚC nó.
    nhap_b = create_payment_request(world.db, code="YCTT_B1", company_id=world.co["B"],
                                    supplier_code="NCCB", supplier_name="NCC B", total=2000)
    cho_duyet_b = create_payment_request(world.db, code="YCTT_B2", company_id=world.co["B"],
                                         supplier_code="NCCB", supplier_name="NCC B",
                                         total=2000, status="submitted")
    a1 = world.grant("a1", "payment_request", scope="company",
                     actions=("read", "write", "approve", "delete"))
    assert a1.sees(PaymentRequest) == set(), "cả hai phiếu đều NGOÀI phạm vi của a1"

    bt = BackgroundTasks()
    cac_cua = [
        ("PATCH", lambda: update_(rid=nhap_b.id, data=PRequestUpdate(note="sửa trộm"),
                                  db=world.db, user=a1.user)),
        ("submit", lambda: submit_(rid=nhap_b.id, background_tasks=bt, db=world.db, user=a1.user)),
        ("approve", lambda: approve_(rid=cho_duyet_b.id, background_tasks=bt,
                                     db=world.db, user=a1.user)),
        ("reject", lambda: reject_(rid=cho_duyet_b.id, data={"reason": "x"},
                                   background_tasks=bt, db=world.db, user=a1.user)),
        ("pay", lambda: pay_(rid=cho_duyet_b.id, background_tasks=bt, db=world.db, user=a1.user)),
        ("refund", lambda: refund_(rid=cho_duyet_b.id, data={"amount": 0},
                                   db=world.db, user=a1.user)),
        ("DELETE", lambda: delete_(rid=nhap_b.id, db=world.db, user=a1.user)),
    ]
    for ten, goi in cac_cua:
        with pytest.raises(HTTPException) as err:
            goi()
        assert err.value.status_code == 404, f"cửa {ten} vẫn mở cho phiếu ngoài phạm vi"

    #  Xóa hàng loạt lọc phạm vi TRƯỚC vòng lặp (khuôn `contract`) — không id nào
    #  trong phạm vi thì 403, và không phiếu nào biến mất.
    with pytest.raises(HTTPException) as err:
        bulk_delete_requests(ids=f"{nhap_b.id},{cho_duyet_b.id}", db=world.db, user=a1.user)
    assert err.value.status_code == 403

    world.db.expire_all()
    assert world.db.get(PaymentRequest, nhap_b.id).status == "draft", "không gì đổi"
    assert world.db.get(PaymentRequest, cho_duyet_b.id).status == "submitted"


def test_c6b_nguoi_TRONG_pham_vi_van_lam_du_moi_thao_tac_yctt(world, two_company_data):
    """Vế đối chứng của C6 — bản vá không được khóa nhầm người trong phạm vi.

    Không có ca này thì một bản vá "chặn sạch mọi đường ghi" cũng xanh, rồi bị gỡ ra
    tuần sau vì kế toán không duyệt chi được phiếu của chính công ty mình.

    Đi đúng chuỗi vòng đời một phiếu trả trước: sửa nháp → gửi duyệt → duyệt → ghi
    nhận chi → NCC hoàn tiền, cộng thêm một phiếu nữa để kiểm đường xóa.
    """
    from app.modules.payment_request.controller import (
        approve_, bulk_delete_requests, delete_, pay_, refund_, submit_, update_)
    from app.modules.payment_request.model import PaymentRequest, PaymentRequestLine
    from app.modules.payment_request.schema import PRequestUpdate

    req_a = create_payment_request(world.db, code="YCTT_A1", company_id=world.co["A"],
                                   supplier_code="NCCA", supplier_name="NCC A",
                                   prepay=1, total=300)
    world.db.add(PaymentRequestLine(request_id=req_a.id, amount=300,
                                    allocated_amount=0, refunded_amount=0))
    xoa_a = create_payment_request(world.db, code="YCTT_A2", company_id=world.co["A"],
                                   supplier_code="NCCA", supplier_name="NCC A", total=100)
    hang_loat_a = create_payment_request(world.db, code="YCTT_A3", company_id=world.co["A"],
                                         supplier_code="NCCA", supplier_name="NCC A", total=100)
    world.db.commit()

    a1 = world.grant("a1", "payment_request", scope="company",
                     actions=("read", "write", "approve", "delete"))
    assert a1.sees(PaymentRequest) == {req_a.id, xoa_a.id, hang_loat_a.id}

    bt = BackgroundTasks()
    assert unwrap(update_(rid=req_a.id, data=PRequestUpdate(note="ghi chú"),
                          db=world.db, user=a1.user))["note"] == "ghi chú"
    assert unwrap(submit_(rid=req_a.id, background_tasks=bt,
                          db=world.db, user=a1.user))["status"] == "submitted"
    assert unwrap(approve_(rid=req_a.id, background_tasks=bt,
                           db=world.db, user=a1.user))["status"] == "approved"
    assert unwrap(pay_(rid=req_a.id, background_tasks=bt,
                       db=world.db, user=a1.user))["status"] == "paid"
    refund_(rid=req_a.id, data={"amount": 100}, db=world.db, user=a1.user)

    delete_(rid=xoa_a.id, db=world.db, user=a1.user)
    bulk_delete_requests(ids=str(hang_loat_a.id), db=world.db, user=a1.user)
    world.db.expire_all()
    assert world.db.get(PaymentRequest, xoa_a.id) is None
    assert world.db.get(PaymentRequest, hang_loat_a.id) is None


def test_c6c_moi_cua_yctt_soi_DUNG_action_cua_no(world, two_company_data):
    """Chốt phạm vi phải hỏi ĐÚNG action của route, không mượn tạm `read`.

    Mượn `read` sai cả hai chiều (đúng lỗ #27 ở cụm khác):

    * **nở ra** — người "xem toàn hệ, chỉ sửa pháp nhân mình" sửa được phiếu bên kia;
    * **đóng sạch** — người chỉ có `approve` (không có `read`) thì không duyệt nổi
      phiếu của chính mình, dù đó là cấu hình bình thường của người duyệt chi.

    Hai vế trong một ca vì cùng một dòng mã quyết định cả hai.
    """
    from app.modules.payment_request.controller import approve_, update_
    from app.modules.payment_request.schema import PRequestUpdate

    req_a = create_payment_request(world.db, code="YCTT_A9", company_id=world.co["A"],
                                   supplier_code="NCCA", supplier_name="NCC A",
                                   total=1000, status="submitted")
    req_b = create_payment_request(world.db, code="YCTT_B9", company_id=world.co["B"],
                                   supplier_code="NCCB", supplier_name="NCC B", total=2000)

    rong_doc = world.grant("a1", "payment_request", scope="all", actions=("read",))
    rong_doc.grant("payment_request", scope="company", actions=("write",))
    with pytest.raises(HTTPException) as err:
        update_(rid=req_b.id, data=PRequestUpdate(note="sửa trộm"), db=world.db, user=rong_doc.user)
    assert err.value.status_code == 404, "phạm vi SỬA hẹp hơn phạm vi XEM — phải theo cái hẹp"

    chi_duyet = world.grant("a2", "payment_request", scope="company", actions=("approve",))
    assert unwrap(approve_(rid=req_a.id, background_tasks=BackgroundTasks(),
                           db=world.db, user=chi_duyet.user))["status"] == "approved", (
        "người chỉ có khóa DUYỆT vẫn phải duyệt được phiếu trong phạm vi của mình")


def test_c9_can_tru_tien_treo_di_theo_pham_vi_GHI(world, two_company_data):
    """`POST /payables/{pid}/offset-prepay` là thao tác **đổi tiền**, phải soi phạm vi
    `write` — vá 05/09/2026 (trước đó `apply_scope` bỏ trống `action` nên mượn `read`).

    Cấu hình dựng ở đây là cấu hình hộp thoại «Phạm vi» cho phép và người khai quyền đọc
    là "xem cả tập đoàn, sửa trong công ty mình": vai trò 1 = XEM toàn hệ, vai trò 2 =
    SỬA trong pháp nhân A.
    """
    from app.modules.payable.controller import offset_prepay_
    from app.modules.payment_request.model import PaymentRequestLine

    #  Tiền treo cấp NCC: một phiếu TRẢ TRƯỚC đã chi, chưa đối trừ đồng nào.
    treo = create_payment_request(world.db, code="YCTT_TT", company_id=world.co["A"],
                                  supplier_code="NCCA", supplier_name="NCC A",
                                  prepay=1, status="paid", total=300)
    world.db.add(PaymentRequestLine(request_id=treo.id, amount=300,
                                    allocated_amount=0, refunded_amount=0))
    world.db.commit()

    a1 = world.grant("a1", "payable", scope="all", actions=("read",))
    a1.grant("payable", scope="company", actions=("write",))

    with pytest.raises(HTTPException) as err:
        offset_prepay_(pid=two_company_data.pay_b.id, data={"amount": 100},
                       db=world.db, user=a1.user)
    assert err.value.status_code == 403, "khoản nợ pháp nhân B: xem được, cấn trừ thì không"

    #  Vế đối chứng: khoản nợ TRONG phạm vi ghi vẫn cấn trừ được, và tiền chạy thật.
    ok = unwrap(offset_prepay_(pid=two_company_data.pay_a.id, data={"amount": 100},
                               db=world.db, user=a1.user))
    assert ok["id"] == two_company_data.pay_a.id
    assert ok["paid_amount"] == 100.0, "100 đ tiền treo đã cấn vào khoản nợ pháp nhân A"


def test_c7_tien_treo_theo_ncc_khong_bo_theo_phap_nhan(world, two_company_data):
    """🔴 `/payment-requests/hanging` gom theo MÃ NCC, không đọc pháp nhân.

    `summarize_hanging` (`service.py:443`) nhận `supplier_code` rồi cộng dồn mọi
    phiếu trả trước của mã đó. Nhà cung cấp dùng chung giữa các pháp nhân là
    chuyện thường (`supplier` khai `PUBLIC` chính vì vậy), nên số tiền treo trả
    ra là số của cả tập đoàn cộng lại.

    Ca này dựng hai phiếu trả trước cùng mã NCC ở hai pháp nhân rồi khẳng định
    **giá trị cộng dồn**, không khẳng định "gọi được".
    """
    from app.modules.payment_request.controller import get_hanging_
    from app.modules.payment_request.model import PaymentRequestLine

    for co, code, amount in ((world.co["A"], "YCTT_TT_A", 300), (world.co["B"], "YCTT_TT_B", 700)):
        req = create_payment_request(world.db, code=code, company_id=co, supplier_code="NCCA",
                                     supplier_name="NCC A", prepay=1, status="paid", total=amount)
        world.db.add(PaymentRequestLine(request_id=req.id, amount=amount, allocated_amount=0,
                                        refunded_amount=0))
    world.db.commit()

    a1 = world.grant("a1", "payment_request", scope="company")
    data = unwrap(get_hanging_(supplier_code="NCCA", unlinked=1, db=world.db, user=a1.user))
    assert data["total"] == 1000.0, "300 (pháp nhân A) + 700 (pháp nhân B)"


def test_c8_hai_o_phong_ban_vo_hieu_tren_cong_no_va_yctt(world, two_company_data):
    """`payable` và `payment_request` khai mỗi `company` + `owner`
    (`scoping.py:45,51`) — **không có chiều phòng ban**.

    Nên hai ô phòng ban của hộp thoại «Phạm vi» là ô câm trên hai màn này:
    `_explicit_cond` gọi `_dept_match` với `f` không có `dept_id`/`dept_name` nên
    nhận `None` rồi bỏ qua, không log, không lỗi. Người khai quyền tick, bấm Lưu,
    nhận thông báo "Đã lưu phạm vi" — và không có gì đổi (nối tiếp nghi ngờ 3 của
    cụm 01).

    # QUYẾT ĐỊNH CHỜ: hộp thoại có nên ẩn/khóa hai ô phòng ban cho entity không
    # khai chiều đó, kèm câu giải thích? (Cùng câu hỏi với B6/B7 của cụm 01.)
    """
    from app.modules.payable.model import Payable

    a1 = world.grant("a1", "payable", scope="all", inc_dept=["A.kt"], exc_dept=["B.kt"])
    assert a1.sees(Payable) == {two_company_data.pay_a.id, two_company_data.pay_b.id}, (
        "hai ô phòng ban không cắt gì trên Công nợ")

    a2 = world.grant("a2", "payable", scope="company", exc_dept=["A.kt"])
    assert a2.sees(Payable) == {two_company_data.pay_a.id}, (
        "bậc `company` vẫn cắt đúng — vô hiệu là RIÊNG chiều phòng ban")


# ══════════════════════════════════════════════════════════════════════════════
#  D. Tồn kho (`inventory`) + Kho (`warehouse`) + Luân chuyển kho
# ══════════════════════════════════════════════════════════════════════════════


def test_d1_ton_kho_chi_co_chieu_phap_nhan_nen_own_va_dept_roi_xuong_company(
        world, two_company_data):
    """`inventory` khai đúng một chiều `company` (`scoping.py:52`).

    Hệ quả cần ghim: bậc `own` rơi xuống `company` (`scoping.py:363` — không có
    `owner` lẫn `self`), bậc `dept` cũng ra đúng điều kiện pháp nhân (nhánh
    `dept` chỉ gom được vế công ty). Cả hai **không** trở thành "thấy tất" —
    đó là điều quan trọng nhất ở đây, vì rơi xuống `None` là lỗi B-07 cũ.

    Ba bậc kiểm trên ba tài khoản khác nhau: `scope_condition` HỢP các grant nên
    dồn vào một người thì bậc rộng đè bậc hẹp và ca mất nghĩa.
    """
    from app.modules.inventory.model import Inventory

    chi_a = {two_company_data.inv_a.id}
    assert world.grant("a1", "inventory", scope="own").sees(Inventory) == chi_a
    assert world.grant("a2", "inventory", scope="dept").sees(Inventory) == chi_a
    assert world.grant("a3", "inventory", scope="company").sees(Inventory) == chi_a
    assert world.grant("b1", "inventory", scope="company").sees(Inventory) == {
        two_company_data.inv_b.id}


def test_d2_nhan_su_chua_gan_phap_nhan_thi_ton_kho_chan_het(world, two_company_data):
    """Người chưa gắn pháp nhân (`company_id = 0`) không thấy dòng tồn kho nào.

    `_chan(...)` ở `scoping.py:390` chặn + ghi WARNING thay vì trả `None`. Nếu
    nhánh này lỏng ra thì mọi tài khoản chưa khai nhân sự — đúng nhóm tài khoản
    dịch vụ / mới tạo — đọc trọn tồn kho toàn hệ.
    """
    from app.modules.inventory.model import Inventory

    khongcty = world.grant("khongcty", "inventory", scope="company")
    assert khongcty.sees(Inventory) == set()


def test_d3_luan_chuyen_kho_di_qua_pham_vi_nhu_man_ton_kho(world, two_company_data):
    """`/api/inventory/moves` phải lọc như màn Tồn kho ngay cạnh — vá 05/09/2026.

    Trước bản vá câu truy vấn là `db.query(InventoryMove)` trần, chỗ duy nhất lọc pháp
    nhân là tham số `company_id` **người dùng tự truyền**. Cùng một khóa `inventory.read`,
    cùng một controller, hai route sát nhau: một có `apply_scope`, một không — kiểu lỗ chỉ
    lộ khi đọc mã, vì đếm route "có `require`" thì cả hai đều đủ.

    Nhật ký nhập/xuất còn chi tiết hơn bảng tồn kho: mã hàng, **đơn giá nhập** và tên
    người thao tác.

    Khẳng định hai chiều bằng hai tài khoản đối xứng: a1 chỉ ra K_A, b1 chỉ ra K_B. Một
    bản vá chặn sạch sẽ đỏ ở vế b1.
    """
    from app.modules.inventory.controller import list_inventory, list_moves
    from app.modules.inventory.model import InventoryMove

    world.db.add_all([
        InventoryMove(company_id=world.co["A"], warehouse_code="K_A", product_code="SP_A",
                      qty=5, unit_price=100, ref_type="adjust"),
        InventoryMove(company_id=world.co["B"], warehouse_code="K_B", product_code="SP_B",
                      qty=7, unit_price=100, ref_type="adjust"),
    ])
    world.db.commit()

    a1 = world.grant("a1", "inventory", scope="company")
    b1 = world.grant("b1", "inventory", scope="company")
    pg = {"offset": 0, "limit": 50}

    ton = unwrap(list_inventory(request=make_request(), pg=pg, db=world.db, user=a1.user))
    assert [r["warehouse_code"] for r in ton["items"]] == ["K_A"], "màn Tồn kho lọc đúng"

    moves = unwrap(list_moves(request=make_request(), pg=pg, db=world.db, user=a1.user))
    assert moves["total"] == 1
    assert [m["warehouse_code"] for m in moves["items"]] == ["K_A"], "…và màn Luân chuyển kho nay nói cùng con số"

    moves_b = unwrap(list_moves(request=make_request(), pg=pg, db=world.db, user=b1.user))
    assert [m["warehouse_code"] for m in moves_b["items"]] == ["K_B"], (
        "vế đối chứng: pháp nhân B vẫn đọc được nhật ký của chính mình")


def test_d4_dieu_chinh_ton_kho_khong_tin_company_id_trong_body(world, two_company_data):
    """`/api/inventory/adjust` lấy pháp nhân từ HỒ SƠ QUYỀN, không từ body — vá 05/09/2026.

    Đây là đường GHI: trước bản vá, người có `inventory.write` bó trong pháp nhân A chỉ
    cần gõ `company_id` của pháp nhân B vào body là sửa thẳng số tồn bên đó, lại có
    `record(...)` đàng hoàng nên nhật ký nhìn vẫn hợp lệ.

    Ba vế: chặn pháp nhân khác · bỏ trống thì lấy pháp nhân của mình · người có phạm vi
    **tất cả** (quản trị đa pháp nhân) vẫn ghi được cho pháp nhân bất kỳ.
    """
    from app.modules.inventory.controller import adjust
    from app.modules.inventory.model import Inventory
    from app.modules.inventory.schema import AdjustIn

    a1 = world.grant("a1", "inventory", scope="company", actions=("read", "write"))
    assert a1.sees(Inventory) == {two_company_data.inv_a.id}, "chỉ thấy kho pháp nhân A"

    with pytest.raises(HTTPException) as err:
        adjust(data=AdjustIn(company_id=world.co["B"], warehouse_code="K_B",
                             product_code="SP_B", product_name="Hàng B", unit="cái",
                             qty=99, note="ghi đè", unit_price=100),
               db=world.db, user=a1.user)
    assert err.value.status_code == 403

    world.db.expire_all()
    assert float(world.db.get(Inventory, two_company_data.inv_b.id).qty) == 7.0, "tồn kho B nguyên vẹn"

    #  Vế đối chứng 1 — kho của chính mình vẫn điều chỉnh được, và bỏ trống `company_id`
    #  thì lấy pháp nhân trong hồ sơ chứ không ghi vào "pháp nhân số 0".
    ra = unwrap(adjust(data=AdjustIn(warehouse_code="K_A", product_code="SP_A",
                                     product_name="Hàng A", unit="cái", qty=3,
                                     note="kiểm kê", unit_price=100),
                       db=world.db, user=a1.user))
    assert ra["company_id"] == world.co["A"] and ra["qty"] == 3.0

    #  Vế đối chứng 2 — quản trị đa pháp nhân (phạm vi `tất cả`) không bị bản vá khóa lại.
    quan_tri = world.grant("a2", "inventory", scope="all", actions=("read", "write"))
    unwrap(adjust(data=AdjustIn(company_id=world.co["B"], warehouse_code="K_B",
                                product_code="SP_B", product_name="Hàng B", unit="cái",
                                qty=2, note="kiểm kê", unit_price=100),
                  db=world.db, user=quan_tri.user))
    world.db.expire_all()
    assert float(world.db.get(Inventory, two_company_data.inv_b.id).qty) == 2.0


def test_d5_danh_muc_kho_la_public_nen_thay_kho_moi_phap_nhan(world, two_company_data):
    """`warehouse` khai `PUBLIC` — đúng thiết kế, nhưng phải nói ra hệ quả.

    `tab_warehouse` **không có cột `company_id`** (`catalog/model.py:7-12`), nên
    đây không phải lựa chọn của `scoping.py` mà là khoảng trống của mô hình dữ
    liệu. Hệ quả kéo theo: phiếu luân chuyển kho giữa hai pháp nhân không có gì
    trong tầng phạm vi chặn được — mã kho là chuỗi tự do, ai cũng chọn được kho
    của bên kia.
    """
    from app.core.scoping import SCOPE_FIELDS, PUBLIC
    from app.modules.catalog.model import Warehouse

    assert SCOPE_FIELDS["warehouse"] is PUBLIC
    assert not hasattr(Warehouse, "company_id"), "bảng kho không có chiều pháp nhân để lọc"

    world.db.add_all([Warehouse(code="K_A", name="Kho A"), Warehouse(code="K_B", name="Kho B")])
    world.db.commit()
    ids = {w.id for w in world.db.query(Warehouse).all()}
    assert len(ids) == 2

    a1 = world.grant("a1", "warehouse", scope="company", inc_company=["A"])
    assert a1.sees(Warehouse) == ids, "bậc company + ô Công ty được xem đều không cắt gì"
