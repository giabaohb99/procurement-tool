"""Test CR-175 — import chứng từ nhiều dòng (Yêu cầu báo giá / Yêu cầu mua hàng).

Kiểm: gộp dòng theo mã phiếu, dry-run không ghi, dòng thiếu trường bắt buộc bị bỏ,
phiếu không còn dòng hợp lệ thì không tạo, mã trùng thì bỏ qua, và revert xoá cả phiếu.
"""
import openpyxl
import pytest

from app.modules.import_tool import catalog_import, doc_import, service
from app.modules.import_tool.model import (ImportBatch, ImportMode, ImportModule,
                                           ImportStatus)
from app.modules.purchase_request.model import PurchaseRequest, PurchaseRequestItem
from app.modules.survey_request.model import SurveyRequest


def _wb(module: int, rows: list[dict]):
    a = doc_import.DOC_ADAPTERS[module]
    cols = [a["code"], *a["header_fields"], *a["line_fields"]]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = a["sheet"]
    for i, f in enumerate(cols, start=1):
        ws.cell(row=1, column=i, value=f["header"])
    for ri, row in enumerate(rows, start=2):
        for i, f in enumerate(cols, start=1):
            if f["attr"] in row:
                ws.cell(row=ri, column=i, value=row[f["attr"]])
    return wb


def _batch(db, module, mode):
    b = ImportBatch(module=module, mode=mode, filename="t.xlsx", file_id=0,
                    status=ImportStatus.QUEUED, created_by=1, updated_by=1)
    db.add(b); db.commit(); db.refresh(b)
    return b


_PR = ImportModule.PURCHASE_REQUEST


def test_gom_dong_theo_ma_phieu(db):
    wb = _wb(_PR, [
        {"code": "PYC1", "requester": "Ng A", "purpose": "MĐ", "product_name": "SP A", "qty": 10},
        {"code": "PYC1", "product_name": "SP B", "qty": 5},   # cùng phiếu -> 2 dòng
        {"code": "PYC2", "requester": "Ng B", "product_name": "SP C", "qty": 3},
    ])
    b = _batch(db, _PR, ImportMode.APPLY)
    doc_import.run(db, b, wb, apply=True)

    assert b.created_count == 2   # 2 phiếu
    p1 = db.query(PurchaseRequest).filter(PurchaseRequest.code == "PYC1").first()
    assert p1 is not None and p1.requester == "Ng A"   # header lấy từ dòng đầu
    assert db.query(PurchaseRequestItem).filter(PurchaseRequestItem.pr_id == p1.id).count() == 2


def test_dry_run_khong_ghi(db):
    wb = _wb(_PR, [{"code": "PYC9", "product_name": "SP", "qty": 1}])
    b = _batch(db, _PR, ImportMode.DRY_RUN)
    doc_import.run(db, b, wb, apply=False)
    assert b.created_count == 1
    assert db.query(PurchaseRequest).filter(PurchaseRequest.code == "PYC9").count() == 0


def test_dong_thieu_ten_sp_bi_bo_phieu_rong_khong_tao(db):
    # Phiếu chỉ có 1 dòng, dòng đó thiếu Tên sản phẩm -> bỏ dòng -> phiếu 0 dòng -> KHÔNG tạo.
    wb = _wb(_PR, [{"code": "PYCX", "qty": 9}])
    b = _batch(db, _PR, ImportMode.APPLY)
    doc_import.run(db, b, wb, apply=True)
    assert b.created_count == 0
    assert db.query(PurchaseRequest).filter(PurchaseRequest.code == "PYCX").count() == 0


def test_ma_da_ton_tai_thi_bo_qua(db):
    db.add(PurchaseRequest(code="PYCDUP", requester="Cũ", created_by=1, updated_by=1))
    db.commit()
    wb = _wb(_PR, [{"code": "PYCDUP", "requester": "Mới", "product_name": "SP", "qty": 1}])
    b = _batch(db, _PR, ImportMode.APPLY)
    doc_import.run(db, b, wb, apply=True)
    assert b.created_count == 0
    # Phiếu cũ giữ nguyên, không bị ghi đè.
    assert db.query(PurchaseRequest).filter(PurchaseRequest.code == "PYCDUP").first().requester == "Cũ"


def test_revert_xoa_ca_phieu_va_dong(db):
    wb = _wb(_PR, [
        {"code": "PYCR", "product_name": "SP A", "qty": 1},
        {"code": "PYCR", "product_name": "SP B", "qty": 2},
    ])
    b = _batch(db, _PR, ImportMode.APPLY)
    doc_import.run(db, b, wb, apply=True)
    p = db.query(PurchaseRequest).filter(PurchaseRequest.code == "PYCR").first()
    assert db.query(PurchaseRequestItem).filter(PurchaseRequestItem.pr_id == p.id).count() == 2

    res = service.revert_batch(db, b, user_id=1)
    assert res["ok"] is True
    assert db.query(PurchaseRequest).filter(PurchaseRequest.code == "PYCR").count() == 0
    assert db.query(PurchaseRequestItem).filter(PurchaseRequestItem.pr_id == p.id).count() == 0


def test_ycbg_gom_dong(db):
    m = ImportModule.SURVEY_REQUEST
    wb = _wb(m, [
        {"code": "YCBG1", "requester": "A", "requirement_detail": "TSKT 1", "request_qty": 100},
        {"code": "YCBG1", "requirement_detail": "TSKT 2", "request_qty": 50},
    ])
    b = _batch(db, m, ImportMode.APPLY)
    doc_import.run(db, b, wb, apply=True)
    assert b.created_count == 1
    sr = db.query(SurveyRequest).filter(SurveyRequest.code == "YCBG1").first()
    assert sr is not None and sr.requester == "A"


def test_build_template_la_xlsx_hop_le(db):
    for m in (ImportModule.SURVEY_REQUEST, ImportModule.PURCHASE_REQUEST):
        data = doc_import.build_template(m)
        assert data[:2] == b"PK"


def test_is_doc_module(db):
    # CR-176: Khảo sát + ĐMH cũng chuyển sang mẫu chuẩn (doc_import).
    assert doc_import.is_doc_module(ImportModule.SURVEY_REQUEST)
    assert doc_import.is_doc_module(ImportModule.PURCHASE_REQUEST)
    assert doc_import.is_doc_module(ImportModule.SURVEY)
    assert doc_import.is_doc_module(ImportModule.PURCHASE_ORDER)
    assert not doc_import.is_doc_module(ImportModule.COMPANY)


def test_import_khao_sat_mau_chuan(db):
    # 1 phiếu khảo sát + N dòng NCC; survey_type mặc định 'supplier'.
    m = ImportModule.SURVEY
    wb = _wb(m, [
        {"code": "KS001", "item_group": "Thùng", "supplier_name": "NCC A", "tax_code": "111"},
        {"code": "KS001", "supplier_name": "NCC B", "tax_code": "222"},
    ])
    b = _batch(db, m, ImportMode.APPLY)
    doc_import.run(db, b, wb, apply=True)
    from app.modules.survey.model import Survey, SurveySupplierLine
    s = db.query(Survey).filter(Survey.code == "KS001").first()
    assert s is not None and s.survey_type == "supplier"
    assert db.query(SurveySupplierLine).filter(SurveySupplierLine.survey_id == s.id).count() == 2


def test_import_dmh_mau_chuan(db):
    m = ImportModule.PURCHASE_ORDER
    wb = _wb(m, [
        {"code": "PO900", "supplier_name": "NCC X", "product_name": "SP 1", "qty_order": 10, "price": 1000},
        {"code": "PO900", "product_name": "SP 2", "qty_order": 5},
    ])
    b = _batch(db, m, ImportMode.APPLY)
    doc_import.run(db, b, wb, apply=True)
    from app.modules.purchase_order.model import POItem, PurchaseOrder
    p = db.query(PurchaseOrder).filter(PurchaseOrder.code == "PO900").first()
    assert p is not None and p.supplier_name == "NCC X"
    assert db.query(POItem).filter(POItem.po_id == p.id).count() == 2


def test_chon_nham_bang_chung_tu_thi_chan(db):
    # File không có cột bắt buộc của YCMH («Mã phiếu», «Tên sản phẩm») -> chặn ngay.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Cột lạ 1", "Cột lạ 2"])
    ws.append(["x", "y"])
    b = _batch(db, _PR, ImportMode.APPLY)
    with pytest.raises(catalog_import.ImportValidationError):
        doc_import.run(db, b, wb, apply=True)
    assert db.query(PurchaseRequest).count() == 0
