"""Ticket #18 (bao-CR-279) — cột Mã đơn Misa trên màn Công nợ phải trả.

Payable KHÔNG lưu misa_code (mã nhập/sửa trên ĐMH sau khi nợ đã sinh) nên phải join
lúc đọc qua `service.misa_code_by_po` — gom MỘT truy vấn cho cả trang, không N+1.
Kiểm 3 chỗ ăn nó: serializer danh sách `_out`, `build_rows` xuất Excel, và endpoint
xuất khi người dùng bật cột (cols có misa_code).

Không đụng DB thật — fixture SQLite in-memory ở conftest.
"""
import json
from datetime import datetime, timedelta
from io import BytesIO
from types import SimpleNamespace

import openpyxl
import pytest
from starlette.datastructures import QueryParams

from app.core.auth import perm_cache_clear
from app.modules.payable import controller as C
from app.modules.payable import service
from app.modules.payable.export import build_rows
from app.modules.payable.model import Payable
from app.modules.purchase_order.model import PurchaseOrder


@pytest.fixture(autouse=True)
def _clear_perm_cache():
    perm_cache_clear()
    yield
    perm_cache_clear()


def _grant(db, user_id: int, entity: str, scope: str):
    from app.modules.role.model import Permission, Role
    from app.modules.user.model import UserRole
    role = Role(code=f"R{user_id}{scope}{entity[:4]}", name="Vai trò test")
    db.add(role)
    db.flush()
    db.add(Permission(role_id=role.id, entity=entity, scope=scope,
                      can_read=True, can_export=True))
    db.add(UserRole(user_id=user_id, role_id=role.id))
    db.flush()
    perm_cache_clear()


def _d(days: int) -> str:
    return (datetime.now().date() + timedelta(days=days)).strftime("%Y-%m-%d")


def _po(db, code: str, misa: str) -> PurchaseOrder:
    po = PurchaseOrder(code=code, misa_code=misa)
    db.add(po)
    db.flush()
    return po


def _payable(db, company_id, **kw):
    vals = dict(company_id=company_id, supplier_code="NX", supplier_name="NCC Xanh",
                source_type="goods", po_code="PO-1", invoice_no="HD-1",
                incur_date=_d(-5), due_date=_d(10), amount=1000, vat=0,
                total=1000, paid_amount=0, remaining=1000, status="unpaid")
    vals.update(kw)
    p = Payable(**vals)
    db.add(p)
    db.flush()
    return p


def _req(qs: str = ""):
    return SimpleNamespace(query_params=QueryParams(qs))


def _user(db, seed):
    from app.modules.user.model import User
    return db.get(User, seed.u_req_id)


def test_danh_sach_tra_misa_code_theo_po(db, seed):
    """`_out` phải kèm misa_code lấy từ ĐMH; khoản không gắn PO trả chuỗi rỗng."""
    po = _po(db, "PO-M1", "MISA-001")
    co_po = _payable(db, seed.company_id, po_id=po.id, po_code=po.code)
    khong_po = _payable(db, seed.company_id, po_id=0, po_code="", invoice_no="HD-2")
    user = _user(db, seed)
    _grant(db, user.id, "payable", "company")

    resp = C.list_payables(request=_req("year=all"),
                           pg={"offset": 0, "limit": 50}, db=db, user=user)
    data = json.loads(resp.body)["data"]
    items = {it["id"]: it for it in data["items"]}
    assert items[co_po.id]["misa_code"] == "MISA-001"
    assert items[khong_po.id]["misa_code"] == ""


def test_misa_code_by_po_gom_dung_map(db, seed):
    """Map po_id -> misa_code: PO chưa nhập mã ra '', danh sách rỗng không truy vấn."""
    po1 = _po(db, "PO-M1", "MISA-001")
    po2 = _po(db, "PO-M2", "")            # ĐMH chưa nhập mã MISA
    p1 = _payable(db, seed.company_id, po_id=po1.id)
    p2 = _payable(db, seed.company_id, po_id=po2.id)
    p3 = _payable(db, seed.company_id, po_id=0)

    m = service.misa_code_by_po(db, [p1, p2, p3])
    assert m == {po1.id: "MISA-001", po2.id: ""}
    assert service.misa_code_by_po(db, []) == {}


def test_xuat_excel_co_cot_ma_don_misa(db, seed):
    """build_rows kèm misa_code, và endpoint xuất đúng khi người dùng bật cột."""
    po = _po(db, "PO-M1", "MISA-001")
    _payable(db, seed.company_id, po_id=po.id, po_code=po.code)
    rows = build_rows(db, [db.query(Payable).one()])
    assert rows[0]["misa_code"] == "MISA-001"

    user = _user(db, seed)
    _grant(db, user.id, "payable", "company")
    resp = C.export_xlsx(request=_req("year=all"), cols="po_code,misa_code",
                         db=db, user=user)
    ws = openpyxl.load_workbook(BytesIO(resp.body)).active
    assert [c.value for c in ws[1]] == ["PO", "Mã đơn Misa"]
    assert (ws["A2"].value, ws["B2"].value) == ("PO-M1", "MISA-001")
