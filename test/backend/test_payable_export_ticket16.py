"""Ticket #16 (bao-CR-275) — xuất Excel màn Công nợ phải trả.

Kiểm 2 nhóm:
- `build_rows`: nhãn đầy đủ như màn hình (trạng thái / loại nợ / tuổi nợ), tên công ty,
  created_at rỗng rơi về ngày phát sinh, tiền giữ kiểu số.
- Endpoint `/api/payables/export/xlsx` gọi trực tiếp: đi qua `_filtered` nên ăn đúng
  bộ lọc + phạm vi dữ liệu (scope company), `ids` tick chọn, `cols` theo cột đang hiện.

Không đụng DB thật — fixture SQLite in-memory ở conftest.
"""
from datetime import datetime, timedelta
from io import BytesIO
from types import SimpleNamespace

import openpyxl
import pytest
from starlette.datastructures import QueryParams

from app.core.auth import perm_cache_clear
from app.modules.payable import controller as C
from app.modules.payable.export import build_rows
from app.modules.payable.model import Payable


@pytest.fixture(autouse=True)
def _clear_perm_cache():
    perm_cache_clear()
    yield
    perm_cache_clear()


def _grant(db, user_id: int, entity: str, scope: str, **actions):
    """Cấp cho user một vai trò mới có đúng bộ quyền cần cho ca test."""
    from app.modules.role.model import Permission, Role
    from app.modules.user.model import UserRole
    role = Role(code=f"R{user_id}{scope}{entity[:4]}", name="Vai trò test")
    db.add(role)
    db.flush()
    db.add(Permission(role_id=role.id, entity=entity, scope=scope,
                      can_read=actions.get("read", True),
                      can_export=actions.get("export", True)))
    db.add(UserRole(user_id=user_id, role_id=role.id))
    db.flush()
    perm_cache_clear()
    return role.id


def _d(days: int) -> str:
    return (datetime.now().date() + timedelta(days=days)).strftime("%Y-%m-%d")


def _payable(db, company_id, **kw):
    vals = dict(company_id=company_id, supplier_code="NX", supplier_name="NCC Xanh",
                source_type="goods", po_code="PO-1", invoice_no="HD-1",
                incur_date=_d(-5), due_date=_d(10), amount=1000, vat=0,
                total=1000, paid_amount=0, remaining=1000, status="Chờ TT")
    vals.update(kw)
    p = Payable(**vals)
    db.add(p)
    db.flush()
    return p


def _req(qs: str = ""):
    """Request giả — `_filtered`/`apply_filters` chỉ đụng tới .query_params."""
    return SimpleNamespace(query_params=QueryParams(qs))


def _user(db, seed):
    from app.modules.user.model import User
    return db.get(User, seed.u_req_id)


def _sheet(resp):
    return openpyxl.load_workbook(BytesIO(resp.body)).active


# ── build_rows ──────────────────────────────────────────────────────────────────
def test_build_rows_nhan_day_du_nhu_man_hinh(db, seed):
    chua_han = _payable(db, seed.company_id)
    qua_han = _payable(db, seed.company_id, source_type="shipping", status="Đã TT",
                       due_date=_d(-40), paid_amount=1000, remaining=0)
    rows = build_rows(db, [chua_han, qua_han])

    assert rows[0]["status"] == "Chờ thanh toán"
    assert rows[0]["source_type"] == "Hàng hóa"
    assert rows[0]["aging"] == "Chưa đến hạn"          # chưa tới hạn thì không gắn "ngày"
    assert rows[0]["company"] == "Cty Test"
    assert float(rows[0]["total"]) == 1000             # tiền giữ kiểu số cho Excel cộng được

    assert rows[1]["status"] == "Đã thanh toán"
    assert rows[1]["source_type"] == "Vận chuyển"
    assert rows[1]["aging"] == "31-60 ngày"


def test_created_at_rong_roi_ve_ngay_phat_sinh(db, seed):
    # Khoản chưa flush (created_at chưa được gán) — như màn hình, cột Ngày phát sinh
    # phải rơi về incur_date chứ không được để trống
    p = Payable(company_id=seed.company_id, supplier_code="NX", incur_date="2026-01-15")
    rows = build_rows(db, [p])
    assert rows[0]["created_at"] == "2026-01-15"


# ── endpoint ────────────────────────────────────────────────────────────────────
def test_xuat_theo_bo_loc_va_pham_vi_cong_ty(db, seed):
    from app.modules.company.model import Company
    cty2 = Company(name="Cty Khac", code="CT02", is_active=True)
    db.add(cty2)
    db.flush()
    _payable(db, seed.company_id)
    _payable(db, seed.company_id, supplier_code="NY", supplier_name="NCC Vang", po_code="PO-2")
    _payable(db, cty2.id, po_code="PO-NGOAI")          # ngoài phạm vi -> không được ra file
    user = _user(db, seed)
    _grant(db, user.id, "payable", "company")

    ws = _sheet(C.export_xlsx(request=_req("year=all"), cols="", db=db, user=user))
    pos = [r[4].value for r in ws.iter_rows(min_row=2)]   # cột thứ 5 = PO
    assert sorted(pos) == ["PO-1", "PO-2"]

    # Lọc thêm theo NCC -> chỉ còn khoản của NCC đó
    ws = _sheet(C.export_xlsx(request=_req("year=all&supplier_code=NY"), cols="", db=db, user=user))
    assert [r[4].value for r in ws.iter_rows(min_row=2)] == ["PO-2"]


def test_tick_chon_thi_chi_xuat_khoan_da_tick(db, seed):
    p1 = _payable(db, seed.company_id)
    _payable(db, seed.company_id, po_code="PO-2")
    user = _user(db, seed)
    _grant(db, user.id, "payable", "company")

    ws = _sheet(C.export_xlsx(request=_req(f"ids={p1.id}"), cols="", db=db, user=user))
    assert [r[4].value for r in ws.iter_rows(min_row=2)] == ["PO-1"]


def test_cols_xuat_dung_cot_dang_hien_theo_thu_tu(db, seed):
    _payable(db, seed.company_id)
    user = _user(db, seed)
    _grant(db, user.id, "payable", "company")

    # `sel` là cột tick chọn trên màn hình — key lạ với file, phải bị bỏ qua
    ws = _sheet(C.export_xlsx(request=_req("year=all"), cols="sel,total,supplier_name",
                              db=db, user=user))
    assert [c.value for c in ws[1]] == ["Tổng nợ", "Nhà cung cấp"]
    assert ws["A2"].value == 1000
    assert ws["B2"].value == "NCC Xanh"
