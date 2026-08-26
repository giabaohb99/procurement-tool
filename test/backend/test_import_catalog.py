"""Test Đ-13d — nhập hàng loạt danh mục nền (Công ty · Phòng ban · Nhân sự).

Chạy thẳng `catalog_import.run` + `service.revert_batch` trên DB SQLite in-memory
(không qua Celery/HTTP), phủ đúng các nhánh dễ vỡ: dry-run không ghi nhưng giữ log,
apply ghi + snapshot, thiếu trường bắt buộc bị bỏ, dedupe cập nhật, revert (xoá bản
ghi mới / khôi phục bản ghi cũ), tham chiếu theo mã, và file mẫu.
"""
import openpyxl
import pytest

from app.modules.catalog.model import ItemGroup, Warehouse
from app.modules.company.model import Company
from app.modules.department.model import Department
from app.modules.employee.model import Employee
from app.modules.import_tool import catalog_import, service
from app.modules.supplier.model import Supplier
from app.modules.import_tool.model import (ImportBatch, ImportChange, ImportLog,
                                           ImportMode, ImportModule, ImportStatus,
                                           LogLevel)
from app.modules.user.model import User


# ── helper ──────────────────────────────────────────────────────────────────────
def _wb(module: int, rows_by_attr: list[dict]):
    """Dựng workbook đúng bộ cột của adapter; mỗi dòng khai theo TÊN THUỘC TÍNH."""
    adapter = catalog_import.ADAPTERS[module]
    fields = adapter["fields"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = adapter["sheet"]
    for i, f in enumerate(fields, start=1):
        ws.cell(row=1, column=i, value=f["header"])
    for ri, row in enumerate(rows_by_attr, start=2):
        for i, f in enumerate(fields, start=1):
            if f["attr"] in row:
                ws.cell(row=ri, column=i, value=row[f["attr"]])
    return wb


def _batch(db, module: int, mode: int) -> ImportBatch:
    b = ImportBatch(module=module, mode=mode, filename="t.xlsx", file_id=0,
                    status=ImportStatus.QUEUED, created_by=1, updated_by=1)
    db.add(b); db.commit(); db.refresh(b)
    return b


def _logs(db, batch_id, level=None):
    q = db.query(ImportLog).filter(ImportLog.batch_id == batch_id)
    if level is not None:
        q = q.filter(ImportLog.level == level)
    return q.all()


# ── dry-run ──────────────────────────────────────────────────────────────────────
def test_chay_thu_khong_ghi_nhung_van_giu_log(db):
    wb = _wb(ImportModule.COMPANY, [{"code": "C1", "name": "Cty 1", "level": 2}])
    b = _batch(db, ImportModule.COMPANY, ImportMode.DRY_RUN)
    catalog_import.run(db, b, wb, apply=False)

    # Không một dòng nào ghi vào bảng thật.
    assert db.query(Company).filter(Company.code == "C1").count() == 0
    # Nhưng batch vẫn đếm được và log xem-trước còn nguyên.
    assert b.status == ImportStatus.DONE
    assert b.created_count == 1
    assert len(_logs(db, b.id)) >= 1
    # Chạy thử thì KHÔNG lưu snapshot để hoàn tác.
    assert db.query(ImportChange).filter(ImportChange.batch_id == b.id).count() == 0


# ── apply ────────────────────────────────────────────────────────────────────────
def test_ghi_that_tao_ban_ghi_va_snapshot(db):
    wb = _wb(ImportModule.COMPANY,
             [{"code": "C1", "name": "Cty 1", "tax_code": "999", "level": 3, "is_active": 1}])
    b = _batch(db, ImportModule.COMPANY, ImportMode.APPLY)
    catalog_import.run(db, b, wb, apply=True)

    c = db.query(Company).filter(Company.code == "C1").first()
    assert c is not None and c.name == "Cty 1" and c.tax_code == "999" and c.level == 3
    assert b.created_count == 1 and b.status == ImportStatus.DONE
    ch = db.query(ImportChange).filter(ImportChange.batch_id == b.id).all()
    assert len(ch) == 1 and ch[0].was_new == 1


def test_thieu_truong_bat_buoc_bi_bo_qua(db):
    # Thiếu 'name' (bắt buộc) -> bỏ qua dòng, đếm skipped + log ERROR.
    wb = _wb(ImportModule.COMPANY, [{"code": "C1"}])
    b = _batch(db, ImportModule.COMPANY, ImportMode.APPLY)
    catalog_import.run(db, b, wb, apply=True)

    assert db.query(Company).filter(Company.code == "C1").count() == 0
    assert b.skipped_count == 1 and b.error_count == 1
    assert _logs(db, b.id, LogLevel.ERROR)


# ── dedupe + revert (khôi phục bản cũ) ──────────────────────────────────────────
def test_dedupe_cap_nhat_roi_revert_khoi_phuc_gia_tri_cu(db):
    db.add(Company(code="C1", name="Tên CŨ", tax_code="111", level=2,
                   created_by=1, updated_by=1))
    db.commit()

    wb = _wb(ImportModule.COMPANY,
             [{"code": "C1", "name": "Tên MỚI", "tax_code": "222", "level": 2}])
    b = _batch(db, ImportModule.COMPANY, ImportMode.APPLY)
    catalog_import.run(db, b, wb, apply=True)

    c = db.query(Company).filter(Company.code == "C1").first()
    assert b.updated_count == 1 and c.name == "Tên MỚI" and c.tax_code == "222"

    res = service.revert_batch(db, b, user_id=1)
    assert res["ok"] is True
    c = db.query(Company).filter(Company.code == "C1").first()
    assert c is not None and c.name == "Tên CŨ" and c.tax_code == "111"
    assert b.status == ImportStatus.REVERTED


def test_revert_xoa_ban_ghi_moi(db):
    wb = _wb(ImportModule.COMPANY, [{"code": "C2", "name": "Cty 2"}])
    b = _batch(db, ImportModule.COMPANY, ImportMode.APPLY)
    catalog_import.run(db, b, wb, apply=True)
    assert db.query(Company).filter(Company.code == "C2").count() == 1

    res = service.revert_batch(db, b, user_id=1)
    assert res["ok"] is True
    assert db.query(Company).filter(Company.code == "C2").count() == 0
    assert b.status == ImportStatus.REVERTED


def test_khong_revert_duoc_ban_chay_thu(db):
    wb = _wb(ImportModule.COMPANY, [{"code": "C3", "name": "Cty 3"}])
    b = _batch(db, ImportModule.COMPANY, ImportMode.DRY_RUN)
    catalog_import.run(db, b, wb, apply=False)
    res = service.revert_batch(db, b, user_id=1)
    assert res["ok"] is False   # chạy thử không ghi gì để hoàn tác


# ── tham chiếu theo mã ───────────────────────────────────────────────────────────
def test_phong_ban_tham_chieu_cong_ty_theo_ma(db):
    db.add(Company(code="CO1", name="Cty Mẹ", created_by=1, updated_by=1))
    db.commit()
    co_id = db.query(Company).filter(Company.code == "CO1").first().id

    wb = _wb(ImportModule.DEPARTMENT, [
        {"code": "D1", "name": "Phòng A", "company_id": "CO1"},   # khớp mã -> id
        {"code": "D2", "name": "Phòng B", "company_id": "KHONGCO"},  # không khớp -> 0 + REVIEW
    ])
    b = _batch(db, ImportModule.DEPARTMENT, ImportMode.APPLY)
    catalog_import.run(db, b, wb, apply=True)

    d1 = db.query(Department).filter(Department.code == "D1").first()
    d2 = db.query(Department).filter(Department.code == "D2").first()
    assert d1.company_id == co_id
    assert d2.company_id == 0
    assert _logs(db, b.id, LogLevel.REVIEW)   # có log rà soát cho mã không khớp


def test_nhan_su_khong_tao_tai_khoan_dang_nhap(db):
    """Giới hạn v1: import Nhân sự KHÔNG sinh tài khoản User."""
    before = db.query(User).count()
    wb = _wb(ImportModule.EMPLOYEE, [{"code": "NV1", "full_name": "Nguyễn Văn A"}])
    b = _batch(db, ImportModule.EMPLOYEE, ImportMode.APPLY)
    catalog_import.run(db, b, wb, apply=True)

    e = db.query(Employee).filter(Employee.code == "NV1").first()
    assert e is not None and e.full_name == "Nguyễn Văn A"
    assert e.status == "official"                 # mặc định khi để trống
    assert db.query(User).count() == before       # không đẻ tài khoản


# ── file mẫu ─────────────────────────────────────────────────────────────────────
def test_file_mau_sinh_ra_xlsx_hop_le(db):
    for module in (ImportModule.COMPANY, ImportModule.DEPARTMENT, ImportModule.EMPLOYEE):
        data = catalog_import.build_template(module)
        assert isinstance(data, bytes) and len(data) > 0
        assert data[:2] == b"PK"   # .xlsx là zip, bắt đầu bằng 'PK'


def test_is_catalog_module_phan_biet_dung(db):
    assert catalog_import.is_catalog_module(ImportModule.COMPANY)
    assert catalog_import.is_catalog_module(ImportModule.EMPLOYEE)
    assert catalog_import.is_catalog_module(ImportModule.WAREHOUSE)
    assert not catalog_import.is_catalog_module(ImportModule.SURVEY)
    assert not catalog_import.is_catalog_module(ImportModule.PURCHASE_ORDER)


# ── danh mục Sản xuất + Kho (CR-174) ─────────────────────────────────────────────
def test_import_kho_don_gian(db):
    wb = _wb(ImportModule.WAREHOUSE, [{"code": "KHOA", "name": "Kho A", "address": "HN"}])
    b = _batch(db, ImportModule.WAREHOUSE, ImportMode.APPLY)
    catalog_import.run(db, b, wb, apply=True)

    w = db.query(Warehouse).filter(Warehouse.code == "KHOA").first()
    assert w is not None and w.name == "Kho A" and w.address == "HN"
    assert b.created_count == 1


def test_import_ncc_vat_kieu_float(db):
    # VAT lưu TỈ LỆ (0.1), kiểu float — kiểm `_to_float` của adapter.
    wb = _wb(ImportModule.SUPPLIER, [{"code": "NCC1", "name": "NCC Test", "vat": "0,1"}])
    b = _batch(db, ImportModule.SUPPLIER, ImportMode.APPLY)
    catalog_import.run(db, b, wb, apply=True)

    s = db.query(Supplier).filter(Supplier.code == "NCC1").first()
    assert s is not None and abs(s.vat - 0.1) < 1e-9   # nhận cả dấu phẩy thập phân


# ── nhập bằng CSV (không chỉ xlsx) ────────────────────────────────────────────────
def _csv_bytes(module: int, rows_by_attr: list[dict], sep: str = ";") -> bytes:
    """Nội dung CSV đúng bộ cột adapter, kèm BOM — giống hệt Excel VN «Lưu thành CSV»."""
    fields = catalog_import.ADAPTERS[module]["fields"]
    lines = [sep.join(str(f["header"]) for f in fields)]
    for row in rows_by_attr:
        lines.append(sep.join(str(row.get(f["attr"], "")) for f in fields))
    return ("﻿" + "\n".join(lines)).encode("utf-8")


def test_nhap_tu_file_csv_semicolon_va_bom(db):
    """CSV (dấu `;` + BOM) đi qua `_csv_to_workbook` rồi import y như xlsx."""
    from app.modules.import_tool import tasks
    wb = tasks._csv_to_workbook(
        _csv_bytes(ImportModule.COMPANY, [{"code": "CSV1", "name": "Công ty CSV", "level": 2}]))
    b = _batch(db, ImportModule.COMPANY, ImportMode.APPLY)
    catalog_import.run(db, b, wb, apply=True)

    c = db.query(Company).filter(Company.code == "CSV1").first()
    assert c is not None and c.name == "Công ty CSV" and c.level == 2
    assert b.created_count == 1 and b.status == ImportStatus.DONE


def test_csv_dau_phay_thap_phan_giu_nguyen_khi_phan_cach_bang_cham_phay(db):
    # Excel VN xuất CSV bằng `;` chính vì `,` là dấu thập phân — VAT «0,1» phải
    # nằm nguyên một ô rồi `_to_float` đổi thành 0.1, không bị cắt làm hai cột.
    from app.modules.import_tool import tasks
    wb = tasks._csv_to_workbook(
        _csv_bytes(ImportModule.SUPPLIER, [{"code": "NCCCSV", "name": "NCC CSV", "vat": "0,1"}]))
    b = _batch(db, ImportModule.SUPPLIER, ImportMode.APPLY)
    catalog_import.run(db, b, wb, apply=True)

    s = db.query(Supplier).filter(Supplier.code == "NCCCSV").first()
    assert s is not None and abs(s.vat - 0.1) < 1e-9


# ── kiểm tra file có đúng bảng không ──────────────────────────────────────────────
def test_chon_nham_bang_thi_chan_va_khong_ghi_gi(db):
    # File mang tiêu đề của Công ty nhưng lại chọn nhập vào Danh mục kho.
    # Thiếu cột bắt buộc «Tên kho» -> chặn ngay, không tạo bản ghi nào.
    wb = _wb(ImportModule.COMPANY, [{"code": "C1", "name": "Cty 1"}])
    b = _batch(db, ImportModule.WAREHOUSE, ImportMode.APPLY)
    with pytest.raises(catalog_import.ImportValidationError):
        catalog_import.run(db, b, wb, apply=True)
    assert db.query(Warehouse).count() == 0


def test_file_khuyet_cot_tuy_chon_van_nhap_binh_thuong(db):
    # File chỉ có 2 cột BẮT BUỘC của NCC (Mã, Tên pháp lý), khuyết mọi cột tuỳ chọn
    # -> KHÔNG bị chặn oan (kiểm-tra-file không được nghiêm quá tay).
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Mã", "Tên pháp lý"])
    ws.append(["NCCMIN", "NCC tối giản"])
    b = _batch(db, ImportModule.SUPPLIER, ImportMode.APPLY)
    catalog_import.run(db, b, wb, apply=True)

    s = db.query(Supplier).filter(Supplier.code == "NCCMIN").first()
    assert s is not None and s.name == "NCC tối giản" and b.created_count == 1


# ── ghi thật từ bản chạy thử ──────────────────────────────────────────────────────
def test_ghi_that_tu_ban_chay_thu_dung_lai_dung_file(db):
    # «Ghi thật» tạo batch APPLY MỚI trỏ về đúng file của bản chạy thử (không upload lại).
    from app.modules.attachment.model import StoredFile
    sf = StoredFile(filename="x.xlsx", file_key="k", url="", content_type="", size=10,
                    created_by=1, updated_by=1)
    db.add(sf); db.commit(); db.refresh(sf)
    dry = ImportBatch(module=ImportModule.COMPANY, mode=ImportMode.DRY_RUN, filename="x.xlsx",
                      file_id=sf.id, status=ImportStatus.DONE, created_by=1, updated_by=1)
    db.add(dry); db.commit(); db.refresh(dry)

    created = service.commit_dry_run(db, dry, user_id=1)
    assert created.id != dry.id
    assert created.mode == ImportMode.APPLY
    assert created.file_id == sf.id and created.module == ImportModule.COMPANY


def test_thieu_ma_thi_tao_moi_voi_ma_tu_sinh(db):
    # Dòng có dữ liệu nhưng TRỐNG Mã -> tạo mới với mã hệ thống tự sinh (không bỏ qua).
    wb = _wb(ImportModule.COMPANY, [{"name": "Cty không mã"}])
    b = _batch(db, ImportModule.COMPANY, ImportMode.APPLY)
    catalog_import.run(db, b, wb, apply=True)

    c = db.query(Company).filter(Company.name == "Cty không mã").first()
    assert c is not None and c.code.startswith("CTY")
    assert b.created_count == 1 and b.skipped_count == 0


def test_sentinel_xoa_gia_tri_cua_o(db):
    # __/empty_value/__ ở một ô -> xóa giá trị trường đó (về rỗng) khi cập nhật.
    wb1 = _wb(ImportModule.COMPANY, [{"code": "CX", "name": "CX", "tax_code": "999"}])
    catalog_import.run(db, _batch(db, ImportModule.COMPANY, ImportMode.APPLY), wb1, apply=True)
    assert db.query(Company).filter(Company.code == "CX").first().tax_code == "999"

    wb2 = _wb(ImportModule.COMPANY,
              [{"code": "CX", "name": "CX", "tax_code": catalog_import.EMPTY_SENTINEL}])
    catalog_import.run(db, _batch(db, ImportModule.COMPANY, ImportMode.APPLY), wb2, apply=True)
    assert db.query(Company).filter(Company.code == "CX").first().tax_code == ""


def test_xoa_dong_bang_cot_hanh_dong_va_revert_tao_lai(db):
    from app.modules.import_tool import service
    catalog_import.run(db, _batch(db, ImportModule.COMPANY, ImportMode.APPLY),
                       _wb(ImportModule.COMPANY, [{"code": "DELME", "name": "Xóa tôi"}]), apply=True)
    assert db.query(Company).filter(Company.code == "DELME").first() is not None

    # File có cột "Hành động" = __/delete/__ -> xóa bản ghi theo Mã.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Mã", "Tên công ty", "Hành động"])
    ws.append(["DELME", "Xóa tôi", catalog_import.DELETE_SENTINEL])
    b = _batch(db, ImportModule.COMPANY, ImportMode.APPLY)
    catalog_import.run(db, b, wb, apply=True)
    assert b.deleted_count == 1
    assert db.query(Company).filter(Company.code == "DELME").first() is None

    # Hoàn tác -> tạo lại bản ghi đã xóa (giữ nguyên Mã).
    res = service.revert_batch(db, b, user_id=1)
    assert res["ok"]
    again = db.query(Company).filter(Company.code == "DELME").first()
    assert again is not None and again.name == "Xóa tôi"


def test_alias_nhan_cot_ngan_cua_file_export(db):
    # File EXPORT dùng nhãn ngắn «Ngày QĐ (có sẵn)» / «Hoạt động» — vẫn khớp nhờ
    # alias, không rơi mất std_days/std_days_unavail như lỗi round-trip đã gặp.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Mã", "Tên", "Ngày QĐ (có sẵn)", "Ngày QĐ (không sẵn)", "Ngày áp dụng", "Hoạt động"])
    ws.append(["PLX", "Nhóm X", "100", "150", "26/08/2026", "Có"])
    b = _batch(db, ImportModule.ITEM_GROUP, ImportMode.APPLY)
    catalog_import.run(db, b, wb, apply=True)

    g = db.query(ItemGroup).filter(ItemGroup.code == "PLX").first()
    assert g is not None
    assert g.std_days == "100" and g.std_days_unavail == "150"
    assert g.apply_date == "26/08/2026" and g.is_active is True


def test_reimport_thieu_ma_cap_nhat_theo_ten_khong_tao_trung(db):
    # Phân loại có name UNIQUE: re-import dòng thiếu Mã -> CẬP NHẬT bản cũ theo tên,
    # không đẻ bản trùng (lỗi 1062 trước đây) — import lần 2 phải là update.
    row = [{"name": "Nhóm trùng tên", "std_days": "100"}]
    b1 = _batch(db, ImportModule.ITEM_GROUP, ImportMode.APPLY)
    catalog_import.run(db, b1, _wb(ImportModule.ITEM_GROUP, row), apply=True)
    assert b1.created_count == 1

    b2 = _batch(db, ImportModule.ITEM_GROUP, ImportMode.APPLY)
    catalog_import.run(db, b2, _wb(ImportModule.ITEM_GROUP, row), apply=True)
    assert b2.updated_count == 1 and b2.error_count == 0
    assert db.query(ItemGroup).filter(ItemGroup.name == "Nhóm trùng tên").count() == 1


def test_trung_rang_buoc_chi_hong_dong_khong_sap_batch(db):
    # Mã KHÁC nhưng tên TRÙNG (name unique) -> chỉ dòng đó lỗi, batch vẫn DONE
    # (savepoint từng dòng), bản ghi lỗi không được tạo.
    catalog_import.run(db, _batch(db, ImportModule.ITEM_GROUP, ImportMode.APPLY),
                       _wb(ImportModule.ITEM_GROUP, [{"code": "IG_A", "name": "T1"}]), apply=True)
    b = _batch(db, ImportModule.ITEM_GROUP, ImportMode.APPLY)
    catalog_import.run(db, b, _wb(ImportModule.ITEM_GROUP, [{"code": "IG_B", "name": "T1"}]), apply=True)
    assert b.error_count == 1 and b.status == ImportStatus.DONE
    assert db.query(ItemGroup).filter(ItemGroup.code == "IG_B").first() is None


def test_roundtrip_phong_ban_xuat_ma_import_lai_giai_ma_dung(db):
    # Export Phòng ban đổi id -> MÃ (công ty / phòng cấp trên / trưởng bộ phận),
    # import lại giải mã đúng về id. Trưởng bộ phận theo MÃ NV nên trùng tên vẫn ổn.
    from app.modules.export_log import service as exp_service
    from app.modules.export_log.registry import EXPORT_ADAPTERS

    db.add(Company(code="C_RT", name="Cty RT", created_by=1, updated_by=1))
    db.add(Employee(code="MGR_RT", full_name="Trưởng RT", created_by=1, updated_by=1))
    db.add(Department(code="D_PARENT", name="PB cha", created_by=1, updated_by=1))
    db.commit()
    comp = db.query(Company).filter(Company.code == "C_RT").first()
    mgr = db.query(Employee).filter(Employee.code == "MGR_RT").first()
    parent = db.query(Department).filter(Department.code == "D_PARENT").first()
    child = Department(code="D_CHILD", name="PB con", kind=2, company_id=comp.id,
                       parent=parent.id, manager_id=mgr.id, created_by=1, updated_by=1)
    db.add(child); db.commit()

    cols = EXPORT_ADAPTERS["department"]["columns"]
    ref_maps = {c.ref: exp_service._code_map(db, c.ref, Department) for c in cols if c.ref}

    def cell(it, c):
        v = ref_maps[c.ref].get(getattr(it, c.key, 0) or 0, "") if c.ref else getattr(it, c.key, "")
        return exp_service._csv_cell(c, v)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([c.label for c in cols])
    ws.append([cell(child, c) for c in cols])

    # Xóa refs của D_CHILD rồi import file vừa xuất -> phải khôi phục đúng id.
    child.company_id = child.parent = child.manager_id = 0
    child.kind = 1
    db.commit()
    catalog_import.run(db, _batch(db, ImportModule.DEPARTMENT, ImportMode.APPLY), wb, apply=True)
    db.refresh(child)
    assert child.company_id == comp.id
    assert child.parent == parent.id
    assert child.manager_id == mgr.id and child.kind == 2


def test_precheck_headers_chan_sai_bang_ngay_tu_dau(db):
    # Bước upload gọi precheck_headers: đúng bảng cho qua, sai bảng ném lỗi ngay.
    from app.modules.import_tool import tasks
    wb = _wb(ImportModule.COMPANY, [{"code": "C1", "name": "Cty 1"}])
    tasks.precheck_headers(ImportModule.COMPANY, wb)   # đúng bảng -> không ném
    with pytest.raises(catalog_import.ImportValidationError):
        tasks.precheck_headers(ImportModule.WAREHOUSE, wb)   # sai bảng -> ném
